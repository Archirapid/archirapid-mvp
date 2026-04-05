"""
Exportador Excel para proyecto ArchiRapid.
Incluye pestaña 'Desglose de Materiales' con PEM detallado.
Uso: bytes_xlsx = generate_project_excel(name, layout, pem_breakdown, financials)
"""
from __future__ import annotations
import io
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _thin_border() -> "Border":
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill(color: str = '1A237E') -> "PatternFill":
    return PatternFill('solid', fgColor=color)


def _write_header_row(ws, row: int, headers: list, bg: str = '1A237E') -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row, col, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = _header_fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thin_border()


def generate_project_excel(
    project_name: str,
    babylon_layout: List[Dict],
    pem_breakdown: Dict[str, Any],
    financials: Dict[str, Any] | None = None,
) -> bytes:
    """
    Genera un Excel con tres hojas:
      - Hoja 1: Resumen del Proyecto
      - Hoja 2: Distribucion de Habitaciones
      - Hoja 3: Desglose de Materiales (PEM)

    Args:
        project_name:   nombre del proyecto
        babylon_layout: lista de rooms desde babylon_modified_layout
        pem_breakdown:  resultado de calculate_pem_breakdown()
        financials:     dict opcional con total_cost, total_area, style

    Returns:
        bytes del archivo Excel (.xlsx)

    Raises:
        ImportError: si openpyxl no está instalado
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl no instalado. Ejecutar: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Proyecto"

    _write_header_row(ws1, 1, ["ArchiRapid — Resumen de Proyecto", "", "", ""], bg='1A237E')
    ws1.merge_cells("A1:D1")

    rows_resumen = [
        ("Proyecto", project_name),
        ("Superficie total",
         f"{sum(float(r.get('width') or 0) * float(r.get('depth') or 0) for r in babylon_layout):.1f} m²"),
    ]
    if financials:
        rows_resumen += [
            ("Coste estimado", f"EUR {financials.get('total_cost', 0):,.0f}"),
            ("Estilo", financials.get('style', '--')),
        ]
    rows_resumen.append(
        ("PEM Materiales (asignados en 3D)", f"EUR {pem_breakdown.get('total_pem', 0):,.2f}")
    )

    for i, (label, value) in enumerate(rows_resumen, start=2):
        ws1.cell(i, 1, label).font = Font(bold=True)
        ws1.cell(i, 2, value)
        for col in (1, 2):
            ws1.cell(i, col).border = _thin_border()

    ws1.column_dimensions['A'].width = 34
    ws1.column_dimensions['B'].width = 24

    # ── Hoja 2: Habitaciones ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Distribucion Habitaciones")
    headers2 = ["Habitacion", "Zona", "Ancho (m)", "Fondo (m)", "Area (m2)", "Material"]
    _write_header_row(ws2, 1, headers2, bg='283593')

    for i, room in enumerate(babylon_layout, start=2):
        w = float(room.get('width') or 0)
        d = float(room.get('depth') or 0)
        ws2.cell(i, 1, room.get('name', '--'))
        ws2.cell(i, 2, room.get('zone', '--'))
        ws2.cell(i, 3, round(w, 2))
        ws2.cell(i, 4, round(d, 2))
        ws2.cell(i, 5, round(w * d, 2))
        ws2.cell(i, 6, room.get('material_id', '--'))
        for col in range(1, 7):
            ws2.cell(i, col).border = _thin_border()

    for col, w_ in zip(range(1, 7), [22, 14, 12, 12, 12, 24]):
        ws2.column_dimensions[get_column_letter(col)].width = w_

    # ── Hoja 3: Desglose de Materiales (PEM) ─────────────────────────────────
    ws3 = wb.create_sheet("Desglose de Materiales")

    _write_header_row(ws3, 1, ["Desglose PEM por Material", "", "", "", ""], bg='1B5E20')
    ws3.merge_cells("A1:E1")

    # Resumen por material
    ws3["A3"] = "RESUMEN POR MATERIAL"
    ws3["A3"].font = Font(bold=True, color='1B5E20')

    _write_header_row(ws3, 4,
                      ["Material", "Precio/m2", "Total m2", "Subtotal EUR", "Normativa CTE"],
                      bg='2E7D32')
    row_n = 5
    for mat_data in pem_breakdown.get('by_material', {}).values():
        ws3.cell(row_n, 1, mat_data['name'])
        ws3.cell(row_n, 2, f"EUR {mat_data['price_m2']:.2f}")
        ws3.cell(row_n, 3, mat_data['total_m2'])
        ws3.cell(row_n, 4, f"EUR {mat_data['subtotal_eur']:,.2f}")
        ws3.cell(row_n, 5, mat_data.get('normativa', ''))
        for col in range(1, 6):
            ws3.cell(row_n, col).border = _thin_border()
        row_n += 1

    row_n += 1
    ws3.cell(row_n, 3, "TOTAL PEM:").font = Font(bold=True)
    total_cell = ws3.cell(row_n, 4, f"EUR {pem_breakdown.get('total_pem', 0):,.2f}")
    total_cell.font = Font(bold=True, color='1B5E20')

    # Detalle por habitacion
    row_n += 2
    ws3.cell(row_n, 1, "DETALLE POR HABITACION")
    ws3.cell(row_n, 1).font = Font(bold=True, color='1B5E20')
    row_n += 1
    _write_header_row(ws3, row_n,
                      ["Habitacion", "Zona", "Material", "Area m2", "Subtotal EUR"],
                      bg='388E3C')
    row_n += 1
    for item in pem_breakdown.get('by_room', []):
        ws3.cell(row_n, 1, item['room_name'])
        ws3.cell(row_n, 2, item['zone'])
        ws3.cell(row_n, 3, item['material_name'])
        ws3.cell(row_n, 4, item['area_m2'])
        ws3.cell(row_n, 5, f"EUR {item['subtotal_eur']:,.2f}")
        for col in range(1, 6):
            ws3.cell(row_n, col).border = _thin_border()
        row_n += 1

    # Aviso habitaciones sin material
    if pem_breakdown.get('rooms_without_material'):
        row_n += 1
        ws3.cell(row_n, 1, "Sin material asignado:")
        ws3.cell(row_n, 1).font = Font(italic=True, color='E65100')
        ws3.cell(row_n, 2, ", ".join(pem_breakdown['rooms_without_material']))

    for col, w_ in zip(range(1, 6), [24, 14, 26, 12, 18]):
        ws3.column_dimensions[get_column_letter(col)].width = w_

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
