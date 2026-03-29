import streamlit as st
import json
import os
from dotenv import load_dotenv
from groq import Groq

from .data_model import create_example_design, HouseDesign, Plot, RoomType, RoomInstance


def _get_groq_key() -> str:
    """Obtiene GROQ_API_KEY desde st.secrets o env."""
    for k in ("GROQ_API_KEY",):
        try:
            v = st.secrets[k]
            if v and str(v).strip():
                return str(v).strip()
        except Exception:
            pass
        v = os.getenv(k, "")
        if v.strip():
            return v.strip()
    return None

# ================================================
# GENERADOR ZIP COMPLETO DEL PROYECTO
# ================================================
def generar_zip_proyecto(req, design_data, plot_data, partidas, subsidy_total, energy_label):
    import io, zipfile, json
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    zip_buffer = io.BytesIO()
    fecha = datetime.now().strftime("%Y-%m-%d")
    proyecto_nombre = f"ArchiRapid_{plot_data.get('title','Proyecto').replace(' ','_')}_{fecha}"
    total_cost = design_data.get('total_cost', 0)
    total_area = design_data.get('total_area', 0)
    style = req.get('style', 'Moderno')
    rooms = design_data.get('rooms', [])
    # Sumar coste de chimenea si el estilo la incluye (sincronizado con Step 4)
    _STYLE_CHIMNEY_ZIP = {'Montaña': 4500, 'Rural': 4500, 'Clásico': 3500}
    total_cost += _STYLE_CHIMNEY_ZIP.get(style, 0)

    # ----------------------------------------
    # GENERAR MEMORIA COMPLETA CON GROQ
    # ----------------------------------------
    memoria_ia = ""
    try:
        from groq import Groq as _Groq
        _client = _Groq(api_key=_get_groq_key())

        habitaciones_list = "\n".join([f"- {r['name']}: {r['area_m2']:.1f} m²" for r in rooms])
        extras_list = [k for k, v in req.get('extras', {}).items() if v]
        energy_list = [k for k, v in req.get('energy', {}).items() if v]

        prompt_memoria = f"""Eres un arquitecto técnico español experto en redacción de proyectos básicos de vivienda unifamiliar.
Redacta una MEMORIA DESCRIPTIVA Y CONSTRUCTIVA profesional y detallada para el siguiente proyecto.
Usa terminología técnica arquitectónica española. Sé específico y profesional.
Al final de cada sección importante añade esta nota: "[MVP ArchiRapid: sección orientativa, requiere desarrollo completo por arquitecto colegiado]"

DATOS DEL PROYECTO:
- Parcela: {plot_data.get('title','N/A')} | Ref. catastral: {plot_data.get('catastral_ref','N/A')}
- Superficie parcela: {plot_data.get('total_m2',0):.0f} m² | Edificable: {plot_data.get('buildable_m2',0):.0f} m²
- Ubicación: Lat {plot_data.get('lat','N/A')}, Lon {plot_data.get('lon','N/A')}
- Superficie construida: {total_area:.1f} m²
- Estilo arquitectónico: {style}
- Tipo cimentación: {req.get('foundation_type','zapatas corridas')}
- Tipo cubierta: {req.get('roof_type','a dos aguas')}
- Calificación energética: {energy_label}
- Dormitorios: {req.get('bedrooms',3)} | Baños: {req.get('bathrooms',2)}
- Extras: {', '.join(extras_list) if extras_list else 'ninguno'}
- Sistemas sostenibles: {', '.join(energy_list) if energy_list else 'ninguno'}

DISTRIBUCIÓN DE ESPACIOS:
{habitaciones_list}
TOTAL: {total_area:.1f} m²

Redacta EXACTAMENTE estas secciones con el nivel de detalle de un proyecto básico real:

1. MEMORIA DESCRIPTIVA
1.1. AGENTES (promotor: datos del cliente ArchiRapid, proyectista: ArchiRapid IA + arquitecto colegiado pendiente de visado)
1.2. INFORMACIÓN PREVIA Y EMPLAZAMIENTO (describe la parcela, entorno, accesos, orientación sur)
1.3. DESCRIPCIÓN DEL PROYECTO (objeto, programa de necesidades, solución adoptada, descripción general)
1.4. CUADRO DE SUPERFICIES (tabla con todas las estancias y sus m²)
1.5. PRESTACIONES CTE (HE ahorro energía, HS salubridad, SE seguridad estructural, SI incendio, SUA accesibilidad)

2. MEMORIA CONSTRUCTIVA
2.1. SUSTENTACIÓN Y CIMENTACIÓN (describe el tipo elegido: {req.get('foundation_type','zapatas')})
2.2. SISTEMA ESTRUCTURAL (hormigón armado, muros de carga según estilo {style})
2.3. SISTEMA ENVOLVENTE (cubierta {req.get('roof_type','dos aguas')}, fachadas estilo {style}, carpintería exterior)
2.4. PARTICIONES INTERIORES Y ACABADOS
2.5. INSTALACIONES PREVISTAS (fontanería, electricidad, climatización, {'paneles solares, ' if req.get('energy',{}).get('solar') else ''}{'aerotermia, ' if req.get('energy',{}).get('aerotermia') else ''}saneamiento)

3. CUMPLIMIENTO NORMATIVA
3.1. CTE DB-HE AHORRO DE ENERGÍA (calificación {energy_label})
3.2. CTE DB-HS SALUBRIDAD
3.3. CTE DB-SE SEGURIDAD ESTRUCTURAL  
3.4. NORMATIVA URBANÍSTICA (edificabilidad 33%, retranqueos, alturas)

Máximo 1800 palabras. Profesional, técnico, directo."""

        _resp = _client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt_memoria}],
            temperature=0.2,
            max_tokens=2500
        )
        memoria_ia = _resp.choices[0].message.content.strip()
    except Exception as _e:
        memoria_ia = f"[Memoria IA no disponible: {_e}. Ver datos en 03_Datos_Proyecto.json]"

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:

        # ── 1. MEMORIA COMPLETA PDF ───────────────────────────
        solar_si    = 'Sí' if req.get('energy', {}).get('solar') else 'No'
        aero_si     = 'Sí' if req.get('energy', {}).get('aerotermia') else 'No'
        geo_si      = 'Sí' if req.get('energy', {}).get('geotermia') else 'No'
        insul_si    = 'Sí' if req.get('energy', {}).get('insulation') else 'No'
        rain_si     = 'Sí' if req.get('energy', {}).get('rainwater') else 'No'
        domo_si     = 'Sí' if req.get('energy', {}).get('domotic') else 'No'

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.colors import HexColor, white
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                            Table, TableStyle, PageBreak)
            from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

            pdf_buf = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buf, pagesize=A4,
                rightMargin=2*cm, leftMargin=2*cm,
                topMargin=2.5*cm, bottomMargin=2*cm)

            C_DARK  = HexColor('#1A252F')
            C_BLUE  = HexColor('#2C3E50')
            C_GREEN = HexColor('#27AE60')
            C_ORA   = HexColor('#E67E22')
            C_LIGHT = HexColor('#ECF0F1')
            C_MVP   = HexColor('#9B59B6')
            C_GREY  = HexColor('#7F8C8D')
            C_ROW   = HexColor('#F8F9FA')

            SS = getSampleStyleSheet()
            def sty(name, **kw):
                return ParagraphStyle(name, parent=SS['Normal'], **kw)

            s_body  = sty('body',  fontSize=9.5, leading=14,  spaceAfter=5,  alignment=TA_JUSTIFY, textColor=C_BLUE)
            s_h1    = sty('h1',    fontSize=13,  fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=5, textColor=white)
            s_h2    = sty('h2',    fontSize=10.5,fontName='Helvetica-Bold', spaceBefore=8,  spaceAfter=3, textColor=C_BLUE)
            s_mvp   = sty('mvp',   fontSize=8,   fontName='Helvetica-Oblique', textColor=C_MVP, leftIndent=8, spaceAfter=8)
            s_aviso = sty('aviso', fontSize=9,   fontName='Helvetica-Bold',    textColor=C_ORA, alignment=TA_CENTER)
            s_small = sty('small', fontSize=8,   textColor=C_GREY, alignment=TA_CENTER)
            s_ctr   = sty('ctr',   fontSize=14,  fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=C_DARK, spaceAfter=6)
            s_sub   = sty('sub',   fontSize=11,  alignment=TA_CENTER, textColor=C_BLUE, spaceAfter=4)

            story = []

            # ── PORTADA
            story.append(Spacer(1, 1*cm))
            hdr = Table([[Paragraph('<font color="white"><b>ArchiRapid</b> — Proyecto de Vivienda</font>',
                          sty('ph', fontSize=18, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=white))]],
                        colWidths=[17*cm])
            hdr.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),C_DARK),('ROWPADDING',(0,0),(-1,-1),18)]))
            story.append(hdr)
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph('PROYECTO BÁSICO DE VIVIENDA UNIFAMILIAR', s_ctr))
            story.append(Paragraph(plot_data.get('title','Parcela').upper(), s_sub))
            story.append(Paragraph(f"Ref. Catastral: {plot_data.get('catastral_ref','N/A')}", s_small))
            story.append(Spacer(1, 0.5*cm))

            kpis = [
                ['SUPERFICIE', 'PRESUPUESTO', 'CALIF. ENERGÉTICA', 'ESTILO'],
                [f"{total_area:.0f} m²", f"€{total_cost:,}", energy_label, style],
            ]
            tk = Table(kpis, colWidths=[4.25*cm]*4)
            tk.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),C_BLUE),('TEXTCOLOR',(0,0),(-1,0),white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                ('FONTNAME',(0,1),(-1,1),'Helvetica-Bold'),('FONTSIZE',(0,1),(-1,1),13),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('ROWPADDING',(0,0),(-1,-1),8),('GRID',(0,0),(-1,-1),0.4,C_LIGHT),
            ]))
            story.append(tk)
            story.append(Spacer(1, 0.4*cm))
            story.append(Paragraph(f"Generado por ArchiRapid MVP | {fecha} | www.archirapid.es", s_small))
            story.append(Paragraph("⚠️ DOCUMENTO ORIENTATIVO — Requiere visado por arquitecto colegiado", s_aviso))
            story.append(PageBreak())

            # ── ÍNDICE
            def sec_header(txt):
                t = Table([[Paragraph(f'<font color="white"><b>{txt}</b></font>',
                    sty('sh', fontSize=11, fontName='Helvetica-Bold', textColor=white))]],
                    colWidths=[17*cm])
                t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),C_BLUE),('ROWPADDING',(0,0),(-1,-1),8)]))
                return t

            idx_rows = [
                ['ÍNDICE',''],
                ['I.   Memoria Descriptiva','3'],
                ['II.  Memoria Constructiva','4'],
                ['III. Cumplimiento CTE','5'],
                ['IV.  Cuadro de Superficies','6'],
                ['V.   Presupuesto por Partidas','7'],
                ['VI.  Eficiencia Energética','8'],
                ['VII. Sistemas Sostenibles','9'],
                ['VIII.Vistas 3D (archivos adjuntos)','—'],
                ['IX.  Aviso Legal MVP','—'],
            ]
            ti = Table(idx_rows, colWidths=[14*cm, 3*cm])
            ti.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),C_DARK),('TEXTCOLOR',(0,0),(-1,0),white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),10),
                ('ROWPADDING',(0,0),(-1,-1),7),('LINEBELOW',(0,0),(-1,-1),0.3,C_LIGHT),
                ('ALIGN',(1,0),(1,-1),'CENTER'),
            ]))
            story.append(ti)
            story.append(PageBreak())

            # ── CUERPO IA
            if memoria_ia:
                for linea in memoria_ia.split('\n'):
                    l = linea.strip()
                    if not l:
                        story.append(Spacer(1,0.15*cm))
                    elif l.startswith('[MVP'):
                        story.append(Paragraph(l, s_mvp))
                    elif (l[:2] in ('1.','2.','3.') and len(l)>3):
                        story.append(Spacer(1,0.2*cm))
                        story.append(sec_header(l))
                        story.append(Spacer(1,0.1*cm))
                    elif l[:4].replace('.','').isdigit() and '.' in l[:5]:
                        story.append(Paragraph(f'<b>{l}</b>', s_h2))
                    else:
                        story.append(Paragraph(l, s_body))
            story.append(PageBreak())

            # ── CUADRO SUPERFICIES
            story.append(sec_header('IV. CUADRO DE SUPERFICIES'))
            story.append(Spacer(1,0.2*cm))
            sup_h = [['ESTANCIA','USO','m²']]
            for r in rooms:
                c = r.get('code','').lower()
                uso = ('Húmedo' if any(x in c for x in ['bano','aseo'])
                  else 'Noche'   if 'dorm' in c
                  else 'Día'     if any(x in c for x in ['salon','cocina','comedor'])
                  else 'Servicio' if any(x in c for x in ['garaje','garage'])
                  else 'Exterior' if any(x in c for x in ['porche','terraza'])
                  else 'Auxiliar')
                sup_h.append([r['name'], uso, f"{r['area_m2']:.1f}"])
            sup_h += [['','',''],
                      ['TOTAL CONSTRUIDO','',f'{total_area:.1f} m²'],
                      ['Máx. edificable','' ,f"{plot_data.get('buildable_m2',0):.0f} m²"],
                      ['Ocupación edificable','',
                       f"{(total_area/max(plot_data.get('buildable_m2',1),1)*100):.1f}%"]]
            ts = Table(sup_h, colWidths=[9*cm,4.5*cm,3.5*cm])
            ts.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),C_DARK),('TEXTCOLOR',(0,0),(-1,0),white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                ('ROWPADDING',(0,0),(-1,-1),5),('GRID',(0,0),(-1,-1),0.3,C_LIGHT),
                ('BACKGROUND',(0,-3),(-1,-1),C_LIGHT),('FONTNAME',(0,-3),(-1,-1),'Helvetica-Bold'),
                ('ROWBACKGROUNDS',(0,1),(-1,-4),[white,C_ROW]),
            ]))
            story.append(ts)
            story.append(Paragraph('[MVP: superficies orientativas. El proyecto definitivo incluirá superficies útiles, construidas y computables según normativa municipal aplicable]', s_mvp))
            story.append(PageBreak())

            # ── PRESUPUESTO
            story.append(sec_header('V. PRESUPUESTO POR PARTIDAS'))
            story.append(Spacer(1,0.2*cm))
            pr_h = [['PARTIDA','COSTE','%','DESCRIPCIÓN']]
            for p in partidas:
                pr_h.append([p[0],p[1],p[2],p[3]])
            pr_h += [['','','',''],
                     ['TOTAL EJECUCIÓN MATERIAL',f'€{total_cost:,}','100%',''],
                     ['Subvenciones estimadas',f'-€{subsidy_total:,}','','IDAE + CCAA'],
                     ['COSTE NETO',f'€{total_cost-subsidy_total:,}','','']]
            tp = Table(pr_h, colWidths=[6*cm,3*cm,2*cm,6*cm])
            tp.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),C_DARK),('TEXTCOLOR',(0,0),(-1,0),white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8.5),
                ('ROWPADDING',(0,0),(-1,-1),5),('GRID',(0,0),(-1,-1),0.3,C_LIGHT),
                ('BACKGROUND',(0,-3),(-1,-1),C_LIGHT),('FONTNAME',(0,-3),(-1,-1),'Helvetica-Bold'),
                ('ROWBACKGROUNDS',(0,1),(-1,-4),[white,C_ROW]),
            ]))
            story.append(tp)
            story.append(Paragraph('[Presupuesto orientativo a €1.200–1.800/m² según estilo y calidad (mercado España 2025). El definitivo requiere mediciones detalladas por arquitecto y aparejador colegiado]', s_mvp))
            story.append(PageBreak())

            # ── EFICIENCIA ENERGÉTICA
            story.append(sec_header('VI. EFICIENCIA ENERGÉTICA'))
            story.append(Spacer(1,0.3*cm))
            letra = energy_label if energy_label else 'B'
            colores_ee = {'A':'#2ECC71','B':'#27AE60','C':'#F39C12','D':'#E74C3C','E':'#C0392B'}
            col_ee = HexColor(colores_ee.get(letra,'#27AE60'))
            ee_rows = [
                ['Calificación Energética', letra],
                ['Zona Climática (estimada)','D3 — Clima mediterráneo interior'],
                ['Demanda calefacción (est.)','≤ 30 kWh/m²·año'],
                ['Demanda refrigeración (est.)','≤ 15 kWh/m²·año'],
                ['Emisiones CO₂ (est.)','< 10 kg CO₂/m²·año'],
                ['Aislamiento fachada','≥ 8 cm poliestireno expandido'],
                ['Carpintería exterior','Aluminio RPT + doble acristalamiento bajo emisivo'],
                ['Paneles solares',solar_si],
                ['Aerotermia',aero_si],
                ['Geotermia',geo_si],
                ['Aislamiento reforzado',insul_si],
                ['Recogida agua de lluvia',rain_si],
                ['Domótica / gestión energética',domo_si],
            ]
            tee = Table(ee_rows, colWidths=[10*cm,7*cm])
            tee.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(0,-1),C_LIGHT),
                ('FONTNAME',(0,0),(-1,-1),'Helvetica'),
                ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),9),
                ('ROWPADDING',(0,0),(-1,-1),6),
                ('GRID',(0,0),(-1,-1),0.3,C_LIGHT),
                ('BACKGROUND',(1,0),(1,0),col_ee),
                ('TEXTCOLOR',(1,0),(1,0),white),
                ('FONTNAME',(1,0),(1,0),'Helvetica-Bold'),
                ('FONTSIZE',(1,0),(1,0),22),
                ('ALIGN',(1,0),(1,0),'CENTER'),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[white,C_ROW]),
            ]))
            story.append(tee)
            story.append(Spacer(1,0.3*cm))
            story.append(Paragraph('[MVP: certificación energética orientativa calculada por IA. La CEE oficial requiere software reconocido (HULC/CE3X) y técnico competente habilitado]', s_mvp))
            story.append(PageBreak())

            # ── SISTEMAS SOSTENIBLES
            story.append(sec_header('VII. SISTEMAS SOSTENIBLES Y EXTRAS'))
            story.append(Spacer(1,0.3*cm))
            extras_list2 = [k for k, v in req.get('extras', {}).items() if v]
            sos_rows = [['SISTEMA','INCLUIDO','DESCRIPCIÓN']]
            sos_items = [
                ('Paneles Solares Fotovoltaicos', solar_si,
                 f"Orientación Sur 30°. Potencia estimada: {int(total_area*0.05):.0f} kWp"),
                ('Aerotermia', aero_si,
                 'Bomba de calor aire-agua. COP estimado 3.5. Calefacción + ACS'),
                ('Geotermia', geo_si,
                 'Captación geotérmica horizontal. Altísima eficiencia estacional'),
                ('Aislamiento Reforzado', insul_si,
                 'Fachada ventilada + SATE. U ≤ 0.25 W/m²K'),
                ('Recogida Agua de Lluvia', rain_si,
                 f"Depósito estimado {int(plot_data.get('total_m2',200)*0.1):.0f} litros. Riego y usos no potables"),
                ('Domótica', domo_si,
                 'Control inteligente climatización, iluminación y seguridad'),
            ]
            for nom, inc, desc in sos_items:
                color_inc = C_GREEN if inc == 'Sí' else C_GREY
                sos_rows.append([nom, inc, desc])
            tsos = Table(sos_rows, colWidths=[5*cm,2.5*cm,9.5*cm])
            tsos.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),C_DARK),('TEXTCOLOR',(0,0),(-1,0),white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8.5),
                ('ROWPADDING',(0,0),(-1,-1),6),('GRID',(0,0),(-1,-1),0.3,C_LIGHT),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[white,C_ROW]),
                ('ALIGN',(1,0),(1,-1),'CENTER'),
            ]))
            story.append(tsos)
            if extras_list2:
                story.append(Spacer(1,0.3*cm))
                story.append(Paragraph(f'<b>Extras incluidos:</b> {", ".join(extras_list2)}', s_body))
            story.append(Paragraph('[MVP: sistemas orientativos. Dimensionado definitivo requiere proyecto de instalaciones por ingeniero especialista]', s_mvp))
            story.append(PageBreak())

            # ── PLANO 2D EMBEBIDO
            import streamlit as _st2
            plano_bytes = (_st2.session_state.get('final_floor_plan')
                        or _st2.session_state.get('current_floor_plan'))
            if plano_bytes:
                try:
                    from reportlab.platypus import Image as RLImage
                    from reportlab.lib.units import cm
                    import io as _io2
                    story.append(sec_header('VIII. PLANO DE DISTRIBUCIÓN EN PLANTA'))
                    story.append(Spacer(1, 0.3*cm))
                    story.append(Paragraph(
                        'Plano de distribución generado automáticamente. '
                        'El proyecto de ejecución incluirá planta acotada, '
                        'alzados N/S/E/O, secciones y detalles constructivos a escala.',
                        s_body))
                    story.append(Spacer(1, 0.2*cm))
                    img_buf = _io2.BytesIO(plano_bytes)
                    rl_img = RLImage(img_buf, width=16*cm, height=12*cm,
                                     kind='proportional')
                    story.append(rl_img)
                    story.append(Spacer(1, 0.2*cm))
                    story.append(Paragraph(
                        '[MVP ArchiRapid: plano orientativo generado por IA. '
                        'El proyecto definitivo incluirá planos técnicos visados '
                        'a escala 1:50 / 1:100 con cotas, superficies y referencias]',
                        s_mvp))
                    story.append(PageBreak())
                except Exception as _img_e:
                    story.append(Paragraph(
                        f'[Plano no disponible en PDF: {_img_e}. Ver 04_Plano_2D.png adjunto]',
                        s_mvp))

            # ── AVISO LEGAL
            story.append(sec_header('IX. AVISO LEGAL — DOCUMENTO MVP'))
            story.append(Spacer(1,0.3*cm))
            aviso_txt = """
<b>Este documento ha sido generado automáticamente por ArchiRapid (MVP v2.1)</b>
mediante inteligencia artificial a partir de los datos del usuario.<br/><br/>
<b>NO tiene validez legal</b> para presentación ante el Ayuntamiento, Colegio de
Arquitectos ni para inicio de obra sin revisión y <b>visado por arquitecto colegiado</b>.<br/><br/>
Tiene carácter <b>orientativo y comercial</b>, destinado a mostrar las capacidades del
sistema y servir de base de trabajo para el arquitecto que desarrolle el proyecto definitivo.<br/><br/>
<b>Cuando el sistema esté completamente operativo incluirá:</b><br/>
- Planos técnicos a escala: planta acotada, alzados N/S/E/O, secciones, detalles constructivos<br/>
- Pliego de condiciones técnicas completo<br/>
- Estado de mediciones por unidades de obra (m³ hormigón, m² enfoscado, ml cerrajería...)<br/>
- Estudio geotécnico y levantamiento topográfico<br/>
- Estudio básico de seguridad y salud<br/>
- Proyectos de instalaciones: electricidad BT, fontanería, climatización, telecomunicaciones<br/>
- Certificado de eficiencia energética oficial (HULC/CE3X)<br/>
- Libro del edificio y plan de mantenimiento<br/><br/>
<b>ArchiRapid | www.archirapid.es | info@archirrapid.es</b>
"""
            av_t = Table([[Paragraph(aviso_txt, sty('avb', fontSize=9, textColor=C_DARK,
                           leading=13))]],colWidths=[17*cm])
            av_t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,-1),HexColor('#FEF9E7')),
                ('BOX',(0,0),(-1,-1),1.5,C_ORA),
                ('ROWPADDING',(0,0),(-1,-1),16),
            ]))
            story.append(av_t)

            doc.build(story)
            zf.writestr(f"{proyecto_nombre}/01_MEMORIA_PROYECTO.pdf", pdf_buf.getvalue())

        except ImportError:
            hab_txt = "\n".join([f"  - {r['name']}: {r['area_m2']:.1f} m²" for r in rooms])
            zf.writestr(f"{proyecto_nombre}/01_Memoria_Descriptiva.txt",
                f"MEMORIA DESCRIPTIVA\nArchiRapid MVP | {fecha}\n\n{memoria_ia}\n\nSUPERFICIES:\n{hab_txt}\nTOTAL: {total_area:.1f} m²\n\nINSTALAR: pip install reportlab")
        except Exception as _pdf_e:
            zf.writestr(f"{proyecto_nombre}/01_Memoria_Error.txt",
                f"Error PDF: {_pdf_e}\n\n{memoria_ia}")


        # 2. ESTADO DE MEDICIONES — Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mediciones y Presupuesto"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="1A252F")
        tfill = PatternFill("solid", fgColor="27AE60")

        ws.merge_cells("A1:F1")
        ws["A1"] = f"ESTADO DE MEDICIONES — {plot_data.get('title','Proyecto').upper()} | ArchiRapid MVP"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = hfill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F2")
        ws["A2"] = f"Ref: {plot_data.get('catastral_ref','N/A')} | {total_area:.0f} m² | {style} | {fecha}"
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A4:F4")
        ws["A4"] = "DISTRIBUCIÓN DE ESPACIOS"
        ws["A4"].font = Font(bold=True, color="FFFFFF")
        ws["A4"].fill = PatternFill("solid", fgColor="2C3E50")

        for col, h in enumerate(["Espacio","Código","Área (m²)","Ancho (m)","Fondo (m)","€/m²"], 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = hf; c.fill = hfill

        for ri, r in enumerate(rooms, 6):
            ws.cell(row=ri, column=1, value=r['name'])
            ws.cell(row=ri, column=2, value=r.get('code',''))
            ws.cell(row=ri, column=3, value=round(r['area_m2'],1))
            ws.cell(row=ri, column=4, value=round(r.get('width',0),2))
            ws.cell(row=ri, column=5, value=round(r.get('depth',0),2))
            ws.cell(row=ri, column=6, value=design_data.get('cost_per_m2', 1500))

        rt = len(rooms)+6
        ws.cell(row=rt, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rt, column=1).fill = tfill
        ws.cell(row=rt, column=3, value=round(total_area,1)).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rt, column=3).fill = tfill

        rp = rt+3
        ws.merge_cells(f"A{rp}:F{rp}")
        ws.cell(row=rp, column=1, value="PARTIDAS PRESUPUESTARIAS").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rp, column=1).fill = PatternFill("solid", fgColor="2C3E50")
        for col, h in enumerate(["Partida","Coste","% Total","Descripción","",""], 1):
            c = ws.cell(row=rp+1, column=col, value=h)
            c.font = hf; c.fill = hfill
        for ri2, (partida, coste, pct, desc) in enumerate(partidas, rp+2):
            ws.cell(row=ri2, column=1, value=partida)
            ws.cell(row=ri2, column=2, value=coste)
            ws.cell(row=ri2, column=3, value=pct)
            ws.cell(row=ri2, column=4, value=desc)
        rf = rp+2+len(partidas)
        ws.cell(row=rf, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rf, column=1).fill = tfill
        ws.cell(row=rf, column=2, value=f"€{total_cost:,}").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rf, column=2).fill = tfill
        ws.cell(row=rf+1, column=1, value="Subvenciones")
        ws.cell(row=rf+1, column=2, value=f"-€{subsidy_total:,}")
        ws.cell(row=rf+2, column=1, value="COSTE NETO").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rf+2, column=1).fill = tfill
        ws.cell(row=rf+2, column=2, value=f"€{total_cost-subsidy_total:,}").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=rf+2, column=2).fill = tfill
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['D'].width = 40

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        zf.writestr(f"{proyecto_nombre}/02_Mediciones_Presupuesto.xlsx", excel_buffer.getvalue())

        # 3. DATOS COMPLETOS JSON
        catastro_data = {
            "proyecto": proyecto_nombre, "fecha": fecha,
            "parcela": {
                "ref_catastral": plot_data.get('catastral_ref','N/A'),
                "descripcion": plot_data.get('title','N/A'),
                "superficie_total_m2": plot_data.get('total_m2',0),
                "edificable_m2": plot_data.get('buildable_m2',0),
                "lat": plot_data.get('lat','N/A'), "lon": plot_data.get('lon','N/A')
            },
            "vivienda": {
                "superficie_m2": total_area, "estilo": style,
                "dormitorios": req.get('bedrooms',0), "banos": req.get('bathrooms',0),
                "calificacion_energetica": energy_label,
                "cimientos": req.get('foundation_type','N/A'),
                "tejado": req.get('roof_type','N/A')
            },
            "presupuesto": {
                "total": total_cost, "subvenciones": subsidy_total,
                "neto": total_cost-subsidy_total, "coste_m2": 1500
            },
            "habitaciones": [{"nombre": r['name'], "codigo": r.get('code',''), "area_m2": r['area_m2']} for r in rooms],
            "sostenibilidad": req.get('energy', {}),
            "aviso": "MVP ArchiRapid. Requiere validacion arquitecto colegiado."
        }
        zf.writestr(f"{proyecto_nombre}/03_Datos_Proyecto.json", json.dumps(catastro_data, ensure_ascii=False, indent=2))

        # 4. PLANO PNG
        import streamlit as _st
        plano = _st.session_state.get('current_floor_plan') or _st.session_state.get('final_floor_plan')
        if plano:
            zf.writestr(f"{proyecto_nombre}/04_Plano_2D.png", plano)

        # 5. VISTAS 3D BABYLON — capturas de perspectivas
        vistas_3d = _st.session_state.get('babylon_captures', {})
        if vistas_3d:
            nombres_legibles = {
                'sur_fachada_principal': '05a_Vista_Sur_Fachada_Principal.png',
                'norte':                 '05b_Vista_Norte.png',
                'este':                  '05c_Vista_Este.png',
                'oeste':                 '05d_Vista_Oeste.png',
                'planta_cenital':        '05e_Vista_Planta_Cenital.png',
            }
            for key, filename in nombres_legibles.items():
                if key in vistas_3d:
                    # Las capturas llegan como data:image/png;base64,...
                    img_data = vistas_3d[key]
                    if ',' in img_data:
                        img_data = img_data.split(',')[1]
                    import base64 as _b64
                    try:
                        zf.writestr(
                            f"{proyecto_nombre}/Vistas_3D/{filename}",
                            _b64.b64decode(img_data)
                        )
                    except Exception:
                        pass
        else:
            # Nota explicativa si no hay capturas
            zf.writestr(
                f"{proyecto_nombre}/Vistas_3D/INSTRUCCIONES.txt",
                "Para incluir las vistas 3D en el ZIP:\n"
                "1. Abre el Editor 3D (Paso 3)\n"
                "2. Pulsa el boton '📸 Capturar Vistas' en la barra de herramientas\n"
                "3. Espera el mensaje de confirmacion\n"
                "4. Vuelve a Documentacion y descarga el ZIP\n"
            )

        # 6. LAYOUT BABYLON JSON
        babylon_layout = _st.session_state.get('babylon_modified_layout')
        if babylon_layout:
            zf.writestr(f"{proyecto_nombre}/06_Layout_3D_Editable.json",
                json.dumps(babylon_layout, ensure_ascii=False, indent=2))

        # 6. README
        zf.writestr(f"{proyecto_nombre}/LEEME.txt", f"""PAQUETE DOCUMENTAL — {proyecto_nombre}
ArchiRapid MVP | {fecha}

CONTENIDO:
  01_Memoria_Descriptiva.txt       — Memoria proyecto básico
  02_Mediciones_Presupuesto.xlsx   — Mediciones y partidas presupuestarias
  03_Datos_Proyecto.json           — Datos completos (catastro, vivienda, costes)
  04_Plano_2D.png                  — Plano distribución en planta
  05_Layout_3D.json                — Modelo 3D editable (si fue modificado)

PRÓXIMOS PASOS:
  1. Arquitecto colegiado para visado
  2. Estudio geotécnico del terreno
  3. Proyecto Básico al Ayuntamiento
  4. Proyecto de Ejecución completo
  5. Contrato constructor

AVISO LEGAL: MVP orientativo. Sin validez legal sin firma de arquitecto colegiado.
www.archirapid.es | info@archirapid.es
""" )

    zip_buffer.seek(0)
    return zip_buffer.getvalue(), f"{proyecto_nombre}.zip"

# ---------------------------------------------------------------------------
# CONSTANTES REUTILIZABLES
# ---------------------------------------------------------------------------
PANEL_MIN_M2 = 3



# ============================================================
# FUENTE ÚNICA DE VERDAD - ARQUITECTURA CENTRALIZADA
# ============================================================

def get_current_design_data():
    """
    Retorna los datos del diseño actual desde la fuente correcta.
    ORDEN DE PRIORIDAD:
    1. babylon_modified_layout (si existe) - diseño editado en 3D
    2. ai_room_proposal - diseño inicial de la IA
    """
    _req_gcd = st.session_state.get("ai_house_requirements", {})
    _style_gcd = _req_gcd.get("style", "Moderno")
    _COST_BY_STYLE = {
        "Moderno": 1500, "Mediterráneo": 1500, "Contemporáneo": 1550,
        "Ecológico": 1600, "Rural": 1650, "Montaña": 1750, "Clásico": 1800,
    }
    COST_PER_M2 = _COST_BY_STYLE.get(_style_gcd, 1500)

    # 1. Intentar cargar desde Babylon (prioridad)
    babylon_data = st.session_state.get("babylon_modified_layout")
    
    if babylon_data:
        rooms = []
        total_area = 0
        for room in babylon_data:
            try:
                area = float(room.get('new_area', room.get('original_area', 10)))
                total_area += area
                rooms.append({
                    'name': room.get('name', 'Espacio'),
                    'code': room.get('code', room.get('name', 'generic')),
                    'area_m2': area,
                    'x': room.get('x', 0),
                    'z': room.get('z', 0),
                    'width': room.get('width', 0),
                    'depth': room.get('depth', 0)
                })
            except (ValueError, TypeError):
                continue
        return {
            'rooms': rooms,
            'total_area': round(total_area, 1),
            'total_cost': int(total_area * COST_PER_M2),
            'cost_per_m2': COST_PER_M2,
            'source': 'babylon',
            'modified': True
        }
    
    # 2. Si no hay Babylon, usar propuesta IA
    req = st.session_state.get("ai_house_requirements", {})
    proposal = req.get("ai_room_proposal", {})
    
    if proposal:
        rooms = []
        total_area = 0
        for code, area in proposal.items():
            try:
                area_float = float(area) if isinstance(area, (str, int, float)) else 10
                total_area += area_float
                rooms.append({
                    'name': code,
                    'code': code,
                    'area_m2': area_float
                })
            except (ValueError, TypeError):
                continue
        return {
            'rooms': rooms,
            'total_area': round(total_area, 1),
            'total_cost': int(total_area * COST_PER_M2),
            'cost_per_m2': COST_PER_M2,
            'source': 'ai_proposal',
            'modified': False
        }
    
    # 3. Fallback
    return {
        'rooms': [],
        'total_area': 0,
        'total_cost': 0,
        'cost_per_m2': COST_PER_M2,
        'source': 'none',
        'modified': False
    }


def main():
    # ============================================
    # 🔗 CONEXIÓN CON DATOS DE LA FINCA COMPRADA
    # ============================================
    
    # Obtener ID de la finca si viene del panel de cliente
    design_plot_id = st.session_state.get("design_plot_id")
    
    if design_plot_id:
        # Consultar datos de la finca en la BD
        try:
            from modules.marketplace.utils import db_conn
            conn = db_conn()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, title, catastral_ref, m2, superficie_edificable, 
                       lat, lon, owner_email
                FROM plots 
                WHERE id = ?
            """, (design_plot_id,))
            
            plot_row = cursor.fetchone()
            conn.close()
            
            if plot_row:
                # Guardar datos de la parcela en session_state
                st.session_state["design_plot_data"] = {
                    'id': plot_row[0],
                    'title': plot_row[1],
                    'catastral_ref': plot_row[2],
                    'total_m2': plot_row[3] or 0,
                    'buildable_m2': plot_row[4] if plot_row[4] else ((plot_row[3] * 0.33) if plot_row[3] else 0),
                    'lat': plot_row[5],
                    'lon': plot_row[6],
                    'owner_email': plot_row[7]
                }
                
                # Pre-configurar superficie objetivo según edificabilidad
                if "ai_house_requirements" not in st.session_state:
                    max_buildable = st.session_state["design_plot_data"]['buildable_m2']
                    st.session_state["ai_house_requirements"] = {
                        "target_area_m2": min(120.0, max_buildable),  # 120m² o el máximo edificable
                        "max_buildable_m2": max_buildable,  # NUEVO: límite máximo
                        "budget_limit": None,
                        "bedrooms": 3,
                        "bathrooms": 2,
                        "wants_pool": False,
                        "wants_porch": True,
                        "wants_garage": False,
                        "wants_outhouse": False,
                        "max_floors": 1,
                        "style": "moderna",
                        "materials": ["hormigón"],
                        "notes": "",
                        "orientation": "Sur",
                        "roof_type": "A dos aguas",
                        "energy_rating": "B",
                        "accessibility": False,
                        "sustainable_materials": []
                    }
            else:
                st.warning("⚠️ No se encontraron datos de la finca. Usando valores por defecto.")
        
        except Exception as e:
            st.error(f"❌ Error cargando datos de la finca: {e}")
            import traceback
            st.code(traceback.format_exc())
    
    # ============================================
    
    # Si hay una finca nueva seleccionada, resetear el diseñador desde cero
    # Esto evita que quede atrapado en un paso anterior de otra sesión
    if design_plot_id and st.session_state.get("_last_design_plot_id") != design_plot_id:
        st.session_state["ai_house_step"] = 1
        st.session_state["_last_design_plot_id"] = design_plot_id
        # Limpiar diseño anterior para que no contamine el nuevo
        st.session_state.pop("ai_room_proposal", None)
        st.session_state.pop("babylon_modified_layout", None)
        st.session_state.pop("current_floor_plan", None)
        st.session_state.pop("babylon_editor_used", None)
    
    # Inicializar el paso si no existe
    if "ai_house_step" not in st.session_state:
        st.session_state["ai_house_step"] = 1
    
    ai_house_step = st.session_state["ai_house_step"]
    
    # Título principal
    st.title("🏠 Diseñador de Vivienda con IA (MVP) v2.1 🔧")
    
    # Barra de progreso visual
    _step_labels = ["", "Requisitos", "Ajustes", "Editor 3D", "Resumen", "Servicios", "Pago"]
    _prog_cols = st.columns(6)
    for _si, _slabel in enumerate(_step_labels[1:], 1):
        with _prog_cols[_si - 1]:
            _active = _si == ai_house_step
            _done   = _si < ai_house_step
            _color  = "#F59E0B" if _active else ("#22C55E" if _done else "#374151")
            _weight = "800" if _active else "600"
            st.markdown(
                f'<div style="text-align:center;padding:6px 2px;border-radius:8px;'
                f'background:{"rgba(245,158,11,0.15)" if _active else "transparent"};'
                f'border:{"2px solid #F59E0B" if _active else "1px solid #374151"};'
                f'font-size:11px;font-weight:{_weight};color:{_color};">'
                f'{"✓ " if _done else f"{_si}. "}{_slabel}</div>',
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # Llamar a la función correspondiente según el paso
    if ai_house_step == 1:
        render_step1()
    elif ai_house_step == 2:
        render_step2()
    elif ai_house_step == 3:
        render_step3_editor()
    elif ai_house_step == 4:
        render_step4_resumen()
    elif ai_house_step == 5:
        render_step5_docs()
    elif ai_house_step == 6:
        render_step6_pago()

    # CRÍTICO: Salir para no ejecutar código del marketplace
    return


def recalculate_layout(modified_rooms: list, house_shape: str = "Rectangular") -> dict:
    """
    Recibe habitaciones con áreas modificadas desde Babylon.
    Redistribuye el layout completo usando generate_layout().
    
    Args:
        modified_rooms: [{'code': str, 'name': str, 'area_m2': float}, ...]
        house_shape: Forma de la planta
    
    Returns:
        dict: {
            'success': bool,
            'layout': [{'x', 'z', 'width', 'depth', 'name', 'code', 'area_m2'}, ...],
            'total_width': float,
            'total_depth': float,
            'error': str (solo si success=False)
        }
    """
    try:
        from .architect_layout import generate_layout
        
        # Validar y limpiar datos de entrada
        rooms_clean = []
        for r in modified_rooms:
            try:
                area = float(r.get('area_m2', r.get('new_area', 10)))
                area = max(2.0, area)  # mínimo 2m²
                rooms_clean.append({
                    'code': str(r.get('code', r.get('name', 'espacio'))),
                    'name': str(r.get('name', 'Espacio')),
                    'area_m2': area
                })
            except (ValueError, TypeError):
                continue
        
        if not rooms_clean:
            return {'success': False, 'error': 'No hay habitaciones válidas'}
        
        # Redistribuir layout completo
        layout = generate_layout(rooms_clean, house_shape)
        
        if not layout:
            return {'success': False, 'error': 'generate_layout devolvió vacío'}
        
        # Calcular dimensiones totales
        total_width = max(item['x'] + item['width'] for item in layout)
        total_depth = max(item['z'] + item['depth'] for item in layout)
        
        return {
            'success': True,
            'layout': layout,
            'total_width': round(total_width, 2),
            'total_depth': round(total_depth, 2)
        }
    
    except Exception as e:
        import traceback
        return {
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }


def get_final_design():
    """
    Retorna el diseño FINAL del usuario.
    ORDEN DE PRIORIDAD:
    1. babylon_modified_layout (si modificó en 3D)
    2. Sliders Paso 2 (desde session_state)
    3. ai_room_proposal original
    
    Returns:
        dict: {
            'rooms': [{'name': str, 'area_m2': float, ...}],
            'total_area': float,
            'total_cost': float,
            'source': 'babylon' | 'step2_sliders' | 'ai_original'
        }
    """
    import streamlit as st
    _req_gfd = st.session_state.get("ai_house_requirements", {})
    _style_gfd = _req_gfd.get("style", "Moderno")
    _COST_BY_STYLE_GFD = {
        "Moderno": 1500, "Mediterráneo": 1500, "Contemporáneo": 1550,
        "Ecológico": 1600, "Rural": 1650, "Montaña": 1750, "Clásico": 1800,
    }
    COST_PER_M2 = _COST_BY_STYLE_GFD.get(_style_gfd, 1500)

    # PRIORIDAD 1: Babylon (diseño modificado en 3D)
    babylon_data = st.session_state.get("babylon_modified_layout")
    if babylon_data:
        rooms = []
        total_area = 0
        for room in babylon_data:
            try:
                area = float(room.get('new_area', room.get('original_area', 10)))
                total_area += area
                rooms.append({
                    'name': room.get('name', 'Espacio'),
                    'code': room.get('name', 'generic'),
                    'area_m2': area,
                    'index': room.get('index', 0)
                })
            except (ValueError, TypeError):
                continue
        
        return {
            'rooms': rooms,
            'total_area': round(total_area, 1),
            'total_cost': int(total_area * COST_PER_M2),
            'source': 'babylon'
        }
    
    # PRIORIDAD 2: Sliders Paso 2 (valores ajustados por usuario)
    req = st.session_state.get("ai_house_requirements", {})
    proposal = req.get("ai_room_proposal", {})
    
    if proposal:
        rooms = []
        total_area = 0
        
        for i, (code, original_area) in enumerate(proposal.items()):
            # Leer del slider si existe
            slider_key = f"step2_slider_{i}"
            area = st.session_state.get(slider_key)
            
            if area is None:
                # Si no hay slider, usar área original
                try:
                    area = float(original_area)
                except (ValueError, TypeError):
                    area = 10
            
            total_area += area
            rooms.append({
                'name': code,
                'code': code,
                'area_m2': area,
                'index': i
            })
        
        # Determinar si es modificado o original
        source = 'step2_sliders' if any(
            st.session_state.get(f"step2_slider_{i}") is not None 
            for i in range(len(proposal))
        ) else 'ai_original'
        
        return {
            'rooms': rooms,
            'total_area': round(total_area, 1),
            'total_cost': int(total_area * COST_PER_M2),
            'source': source
        }
    
    # FALLBACK: Vacío
    return {
        'rooms': [],
        'total_area': 0,
        'total_cost': 0,
        'source': 'none'
    }

def render_step1():
    """Paso 1: Configurador de vivienda - Estilo Mercedes Benz"""
    
    # Obtener datos de la finca
    plot_data = st.session_state.get("design_plot_data", {})
    max_buildable = plot_data.get('buildable_m2', 0)

    # ============================================
    # HERO BANNER PROFESIONAL
    # ============================================
    st.markdown("""
    <div style='
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 40px 30px;
        border-radius: 16px;
        text-align: center;
        margin-bottom: 24px;
        border: 1px solid rgba(52,152,219,0.3);
    '>
        <h1 style='
            color: white;
            font-size: 2.4em;
            margin: 0 0 10px 0;
            font-weight: 700;
            letter-spacing: -0.5px;
        '>🏡 Diseña Tu Vivienda Sostenible</h1>
        <p style='
            color: rgba(255,255,255,0.75);
            font-size: 1.1em;
            margin: 0 0 20px 0;
        '>Inteligencia Artificial + Editor 3D + Normativa CTE · Todo en un solo flujo</p>
        <div style='display: flex; justify-content: center; gap: 30px; flex-wrap: wrap;'>
            <div style='color: #2ECC71; font-size: 0.95em;'>✅ Distribución automática</div>
            <div style='color: #3498DB; font-size: 0.95em;'>🏗️ Editor 3D profesional</div>
            <div style='color: #E67E22; font-size: 0.95em;'>📐 Planos constructivos</div>
            <div style='color: #9B59B6; font-size: 0.95em;'>💰 Presupuesto en tiempo real</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # BANNER FINCA (si existe)
    if plot_data:
        col1, col2, col3 = st.columns(3)
        col1.metric("📍 Tu Finca", plot_data.get('title', 'Tu parcela'))
        col2.metric("📏 Superficie total", f"{plot_data.get('total_m2', 0):.0f} m²")
        col3.metric("🏗️ Máx. edificable (33%)", f"{max_buildable:.0f} m²")
        st.markdown("---")
    
    # ============================================
    # PASO A: PRESUPUESTO
    # ============================================
    st.subheader("💰 ¿Cuál es tu presupuesto?")
    
    budget = st.select_slider(
        "Selecciona tu rango de inversión",
        options=[50000, 75000, 100000, 125000, 150000, 175000, 200000, 
                 250000, 300000, 400000, 500000],
        value=150000,
        format_func=lambda x: f"€{x:,.0f}"
    )
    
    # Mensaje orientativo según presupuesto
    if budget < 100000:
        st.info("🏠 Casa básica sostenible: 60-80 m² · 2 dormitorios · Eficiente")
    elif budget < 150000:
        st.info("🏡 Casa confortable: 80-120 m² · 3 dormitorios · Muy eficiente")
    elif budget < 250000:
        st.success("🏘️ Casa amplia premium: 120-180 m² · 4 dormitorios · Alta eficiencia")
    else:
        st.success("🏰 Casa de lujo sostenible: 180m²+ · 5+ dormitorios · Máxima eficiencia")
    
    st.markdown("---")
    
    # ============================================
    # PASO B: ESTILO DE VIVIENDA - CARDS CON FOTOS
    # ============================================
    st.subheader("🎨 ¿Qué estilo te gusta?")

    def _style_img_b64(filename):
        import base64 as _b64, os as _os
        path = _os.path.join("assets", "estilos", filename)
        if _os.path.exists(path):
            with open(path, "rb") as _f:
                data = _b64.b64encode(_f.read()).decode()
            ext = filename.rsplit(".", 1)[-1].lower()
            mime = "image/png" if ext == "png" else "image/jpeg"
            return f"data:{mime};base64,{data}"
        return ""

    styles_data = {
        "Ecológico":    {"desc": "Materiales naturales, mínimo impacto ambiental", "img": _style_img_b64("ecologico.jpg")},
        "Rural":        {"desc": "Piedra, madera, integrado en el paisaje",        "img": _style_img_b64("rural.jpg")},
        "Moderno":      {"desc": "Líneas limpias, grandes ventanales, minimalista", "img": _style_img_b64("moderno.jpg")},
        "Montaña":      {"desc": "Refugio alpino, tejados inclinados, madera y piedra", "img": _style_img_b64("montana.jpg")},
        "Playa":        {"desc": "Abierto, ventilado, colores claros, terrazas",   "img": _style_img_b64("playa.jpg")},
        "Clásico":      {"desc": "Elegante, simétrico, materiales nobles",          "img": _style_img_b64("clasico.jpg")},
        "Andaluz":      {"desc": "Patio central, cerámica, cal, frescor natural",  "img": _style_img_b64("andaluz.jpg")},
        "Contemporáneo":{"desc": "Vanguardista, tecnológico, sostenible",           "img": _style_img_b64("contemporaneo.jpg")},
    }

    # Inicializar selección
    if 'selected_style_key' not in st.session_state:
        st.session_state['selected_style_key'] = 'Moderno'

    # Render cards en filas de 4
    style_keys = list(styles_data.keys())
    for row_start in range(0, len(style_keys), 4):
        cols = st.columns(4)
        for col, style_key in zip(cols, style_keys[row_start:row_start+4]):
            data = styles_data[style_key]
            is_selected = st.session_state['selected_style_key'] == style_key
            border_color = "#3498DB" if is_selected else "rgba(255,255,255,0.1)"
            bg_color = "rgba(52,152,219,0.15)" if is_selected else "rgba(255,255,255,0.03)"
            check = "✅ " if is_selected else ""

            with col:
                st.markdown(f"""
                <div style='\
                    border: 2px solid {border_color};\
                    border-radius: 12px;\
                    overflow: hidden;\
                    background: rgba(20,30,48,0.95);\
                    margin-bottom: 8px;\
                    cursor: pointer;\
                '>
                    <img src='{data["img"]}' style='width:100%; height:130px; object-fit:cover;'>
                    <div style='padding: 8px 10px;'>
                        <p style='margin:0; font-weight:700; font-size:0.9em; color:white; text-shadow: 1px 1px 3px rgba(0,0,0,0.9);'>{check}{style_key}</p>
                        <p style='margin:2px 0 0 0; font-size:0.75em; color:rgba(255,255,255,0.85); text-shadow: 1px 1px 2px rgba(0,0,0,0.9);'>{data["desc"]}</p>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if st.button(
                    "✓ Seleccionar" if is_selected else "Elegir",
                    key=f"style_btn_{style_key}",
                    use_container_width=True,
                    type="primary" if is_selected else "secondary"
                ):
                    st.session_state['selected_style_key'] = style_key
                    st.rerun()

    selected_style = st.session_state['selected_style_key']
    st.caption(f"✨ Seleccionado: **{selected_style}** — {styles_data[selected_style]['desc']}")
    st.info(f"💡 Este estilo se aplicará automáticamente al editor 3D. Dentro del editor podrás ajustar los materiales de fachada si lo deseas.", icon=None)
    st.markdown("---")
    
    # ============================================
    # PASO C: HABITACIONES Y BAÑOS
    # ============================================
    st.subheader("🛏️ Habitaciones y Baños")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        bedrooms = st.number_input(
            "🛏️ Dormitorios",
            min_value=1, max_value=6,
            value=3, step=1
        )
    
    with col2:
        bathrooms = st.number_input(
            "🚿 Baños",
            min_value=1, max_value=4,
            value=2, step=1
        )
    
    with col3:
        floors = st.selectbox(
            "Plantas",
            ["1 Planta", "2 Plantas", "Planta Baja + Semisótano", "2 Plantas + Semisótano"],
            index=0,
            help="Cada planta genera su propio plano de distribución"
        )
        
        # Avisos según selección
        if floors == "2 Plantas":
            st.caption("⚠️ Se diseñarán 2 planos: Planta Baja y Planta Alta")
        elif "Semisótano" in floors:
            st.caption("⚠️ Incluye plano de semisótano (garage/bodega/zona técnica)")
    
    st.markdown("---")
    
    # ============================================
    # PASO C2: FORMA Y TEJADO
    # ============================================
    st.subheader("🏗️ Forma y Tejado")
    
    col1, col2 = st.columns(2)
    
    with col1:
        house_shape = st.selectbox(
            "Forma de la planta",
            [
                "Cuadrada (más económica)",
                "Rectangular (más común)",
                "En L (para parcelas irregulares)",
                "Irregular / Personalizada"
            ],
            index=1,
            help="La forma afecta al coste de cimentación y construcción"
        )
        
        # Info de coste según forma
        shape_costs = {
            "Cuadrada (más económica)": ("Mínimo perímetro = mínimo coste", "✅"),
            "Rectangular (más común)": ("Equilibrio perfecto coste/funcionalidad", "✅"),
            "En L (para parcelas irregulares)": ("15-20% más cara que rectangular", "⚠️"),
            "Irregular / Personalizada": ("20-30% más cara, requiere arquitecto", "⚠️")
        }
        
        msg, icon = shape_costs[house_shape]
        st.caption(f"{icon} {msg}")
    
    with col2:
        roof_type = st.selectbox(
            "Tipo de tejado",
            [
                "Dos aguas (clásico, eficiente)",
                "Cuatro aguas (para zonas con mucha lluvia)",
                "Plana/Transitable (zona seca, terraza)",
                "A un agua (moderno, minimalista)",
                "Invertida (colecta agua lluvia)"
            ],
            index=0,
            help="El tejado afecta al coste, aislamiento y recogida de agua"
        )
    
    # Guardar selecciones
    if 'request' not in st.session_state:
        st.session_state['request'] = {}
    
    st.session_state['request']['house_shape'] = house_shape
    st.session_state['request']['roof_type'] = roof_type
    
    st.markdown("---")
    
    # ============================================
    # PASO D: EXTRAS - CARDS VISUALES
    # ============================================
    st.subheader("🌟 Extras")
    st.caption("Selecciona los elementos adicionales que quieres incluir")

    extras_data = [
        {"key": "garage",     "label": "Garaje",         "desc": "2 plazas cubiertas",     "color": "#1a3a5c", "accent": "#3498DB", "default": True},
        {"key": "pool",       "label": "Piscina",        "desc": "8×4m con depuradora",    "color": "#0d3d56", "accent": "#00BCD4", "default": False},
        {"key": "porch",      "label": "Porche/Terraza", "desc": "Cubierto, suelo ext.",   "color": "#1a3d2b", "accent": "#2ECC71", "default": True},
        {"key": "bodega",     "label": "Bodega",         "desc": "9m² climatizada",         "color": "#3d1a2b", "accent": "#9B59B6", "default": False},
        {"key": "huerto",     "label": "Huerto",         "desc": "30m² con riego",          "color": "#1e3d1a", "accent": "#27AE60", "default": False},
        {"key": "caseta",     "label": "Casa de Aperos", "desc": "15m² exterior",           "color": "#3d2e1a", "accent": "#E67E22", "default": False},
        {"key": "accessible", "label": "Accesible",      "desc": "Rampas y baño adaptado", "color": "#1a2d3d", "accent": "#5DADE2", "default": False},
        {"key": "office",     "label": "Despacho",       "desc": "10m² insonorizado",       "color": "#2d1a3d", "accent": "#8E44AD", "default": False},
    ]

    # Inicializar estado
    for extra in extras_data:
        if f"extra_{extra['key']}" not in st.session_state:
            st.session_state[f"extra_{extra['key']}"] = extra['default']

    # Render en 2 filas de 4
    for row_start in range(0, len(extras_data), 4):
        cols = st.columns(4)
        for col, extra in zip(cols, extras_data[row_start:row_start+4]):
            key = f"extra_{extra['key']}"
            is_on = st.session_state[key]
            bg = extra['color'] if not is_on else extra['color']
            border = extra['accent'] if is_on else "rgba(255,255,255,0.1)"
            opacity = "1" if is_on else "0.55"
            indicator = f"<span style='color:{extra['accent']}; font-weight:900;'>●</span>" if is_on else "<span style='color:rgba(255,255,255,0.3); font-weight:900;'>○</span>"

            with col:
                st.markdown(f"""
                <div style='\
                    border: 2px solid {border};\
                    border-radius: 10px;\
                    padding: 10px 12px;\
                    background: {bg};\
                    margin-bottom: 6px;\
                    opacity: {opacity};\
                '>
                    <div style='display:flex; justify-content:space-between; align-items:center;'>
                        <p style='margin:0; font-weight:700; font-size:0.88em; color:white;'>{extra['label']}</p>
                        {indicator}
                    </div>
                    <p style='margin:3px 0 0 0; font-size:0.72em; color:rgba(255,255,255,0.75);'>{extra['desc']}</p>
                </div>
                """, unsafe_allow_html=True)

                if st.button(
                    "✓ Incluido" if is_on else "Añadir",
                    key=f"btn_extra_{extra['key']}",
                    use_container_width=True,
                    type="primary" if is_on else "secondary"
                ):
                    st.session_state[key] = not is_on
                    st.rerun()

    # Extraer valores para uso posterior
    has_garage     = st.session_state["extra_garage"]
    has_pool       = st.session_state["extra_pool"]
    has_porch      = st.session_state["extra_porch"]
    has_bodega     = st.session_state["extra_bodega"]
    has_huerto     = st.session_state["extra_huerto"]
    has_caseta     = st.session_state["extra_caseta"]
    has_accessible = st.session_state["extra_accessible"]
    has_office     = st.session_state["extra_office"]

    st.markdown("---")
    
    # ============================================
    # PASO E: ENERGÍA Y SOSTENIBILIDAD
    # ============================================
    st.subheader("⚡ Energía y Sostenibilidad")
    st.caption("Reduce tu factura energética hasta un 90%")
    
    col1, col2 = st.columns(2)
    
    with col1:
        solar = st.checkbox("☀️ Paneles Solares", value=True,
            help="Autoconsumo eléctrico. Ahorro: €1,200/año")
        aerotermia = st.checkbox("🌡️ Aerotermia",
            help="Calefacción/frío eficiente. Ahorro: €800/año")
        geotermia = st.checkbox("🌍 Geotermia",
            help="Temperatura constante del suelo")
    
    with col2:
        rainwater = st.checkbox("💧 Recuperación Agua Lluvia", value=True,
            help="Ahorro hasta 40% en agua")
        insulation = st.checkbox("🌿 Aislamiento Natural",
            help="Lana de roca, corcho, cáñamo")
        domotic = st.checkbox("🏠 Domótica",
            help="Control inteligente del hogar")
    
    # Calcular ahorro estimado
    ahorro_anual = 0
    if solar: ahorro_anual += 1200
    if aerotermia: ahorro_anual += 800
    if rainwater: ahorro_anual += 300
    if insulation: ahorro_anual += 400
    if geotermia: ahorro_anual += 1000
    
    if ahorro_anual > 0:
        st.success(f"💰 Ahorro energético estimado: **€{ahorro_anual:,}/año** · Retorno inversión: {int(15000/ahorro_anual)} años")
    
    st.markdown("---")
    
    # ============================================
    # PASO F: NOTAS ESPECIALES
    # ============================================
    st.subheader("📝 Algo más que quieras añadir")
    
    special_notes = st.text_area(
        "Cuéntanos más detalles",
        placeholder="Ej: Quiero una bodega grande para eventos, necesito espacio para animales, me gusta mucho la luz natural...",
        height=80,
        label_visibility="collapsed"
    )
    
    st.markdown("---")
    
    # ============================================
    # PASO G: CIMIENTOS E INSTALACIONES
    # ============================================
    st.subheader("🏗️ Cimientos e Instalaciones Básicas")
    st.caption("Calculamos automáticamente lo mínimo necesario para abaratar costes")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        foundation_type = st.selectbox(
            "Tipo de cimentación",
            [
                "Zapatas aisladas (mínimo, más barato)",
                "Losa de hormigón (suelos blandos)",
                "Pilotes (terrenos difíciles)",
                "Recomendación automática IA"
            ],
            index=3
        )
    
    with col2:
        st.markdown("**Suministro de agua**")
        water_municipal = st.checkbox("🏘️ Red municipal (si disponible)", value=False)
        water_rain = st.checkbox("🌧️ Depósito agua lluvia", value=True, 
            help="Depósito 5,000-10,000L. Ahorro 40% en agua")
        water_well = st.checkbox("⛏️ Pozo propio", value=False)
        water_solar = st.checkbox("☀️ Calentador solar ACS", value=True,
            help="Agua caliente sanitaria solar. Ahorro €400/año")
        
        # Calcular combinación
        water_systems = []
        if water_municipal: water_systems.append("red municipal")
        if water_rain: water_systems.append("depósito lluvia")
        if water_well: water_systems.append("pozo")
        if water_solar: water_systems.append("solar ACS")
        
        if not water_systems:
            st.warning("⚠️ Selecciona al menos un sistema de agua")
        else:
            st.caption(f"✅ Combinación: {' + '.join(water_systems)}")
    
    with col3:
        st.markdown("**Saneamiento**")
        sew_municipal = st.checkbox("🏘️ Red municipal (si disponible)", value=False,
            key="sew_municipal")
        sew_septic = st.checkbox("♻️ Fosa séptica ecológica", value=True,
            help="Autónomo, sin conexión a red")
        sew_phyto = st.checkbox("🌿 Fitodepuración", value=False,
            help="Máxima sostenibilidad. Filtra con plantas naturales")
        sew_biodigestor = st.checkbox("⚗️ Biodigestor", value=False,
            help="Genera biogás aprovechable para cocina")
        
        sewage_systems = []
        if sew_municipal: sewage_systems.append("red municipal")
        if sew_septic: sewage_systems.append("fosa séptica")
        if sew_phyto: sewage_systems.append("fitodepuración")
        if sew_biodigestor: sewage_systems.append("biodigestor")
        
        if not sewage_systems:
            st.warning("⚠️ Selecciona al menos un sistema de saneamiento")
        else:
            st.caption(f"✅ Combinación: {' + '.join(sewage_systems)}")
    
    st.markdown("---")
    
    # ============================================
    # BOTÓN DISEÑAR
    # ============================================
    
    # Calcular m² recomendados según presupuesto
    cost_per_m2 = 1100  # €/m² promedio sostenible
    max_m2_budget = int(budget * 0.85 / cost_per_m2)  # 85% para construcción
    
    if max_buildable > 0:
        recommended_m2 = min(max_m2_budget, int(max_buildable * 0.9))
    else:
        recommended_m2 = max_m2_budget
    
    # Calcular presupuesto cimentación automáticamente
    foundation_cost = int(recommended_m2 * 180)  # €180/m² cimentación
    
    st.info(f"💰 Presupuesto estimado de cimentación: **€{foundation_cost:,}** ({int(foundation_cost/budget*100)}% del presupuesto total) · Incluido en el presupuesto global")
    
    # Resumen antes del botón - diseño profesional
    extras_activos = []
    if has_garage: extras_activos.append("Garaje")
    if has_pool: extras_activos.append("Piscina")
    if has_porch: extras_activos.append("Porche")
    if has_bodega: extras_activos.append("Bodega")
    if has_huerto: extras_activos.append("Huerto")
    if has_caseta: extras_activos.append("Casa Aperos")
    if has_office: extras_activos.append("Despacho")

    extras_html = "".join([
        f"<span style='background:rgba(255,255,255,0.15); padding:3px 10px; "
        f"border-radius:20px; margin:3px; display:inline-block; font-size:12px;'>"
        f"{e}</span>"
        for e in extras_activos
    ]) if extras_activos else "<span style='opacity:0.6; font-size:12px;'>Sin extras seleccionados</span>"

    ahorro_html = (
        f"<div style='margin-top:12px; padding:8px 16px; background:rgba(46,204,113,0.2); "
        f"border-radius:8px; border:1px solid rgba(46,204,113,0.4); font-size:13px;'>"
        f"⚡ Ahorro energético estimado: <b>€{ahorro_anual:,}/año</b></div>"
    ) if ahorro_anual > 0 else ""

    resumen_html = (
        "<div style='background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);"
        "padding: 24px; border-radius: 16px; color: white;"
        "border: 1px solid rgba(255,255,255,0.1);"
        "box-shadow: 0 8px 32px rgba(0,0,0,0.3);'>"
        "<p style='margin:0 0 16px 0; font-size:13px; opacity:0.6;"
        "letter-spacing:2px; text-transform:uppercase;'>Resumen de tu proyecto</p>"
        "<div style='display:grid; grid-template-columns: repeat(4,1fr); gap:12px; margin-bottom:16px;'>"
        "<div style='background:rgba(255,255,255,0.07); padding:14px; border-radius:10px;"
        "text-align:center; border:1px solid rgba(255,255,255,0.1);'>"
        "<div style='font-size:11px; opacity:0.6; margin-bottom:4px;'>PRESUPUESTO</div>"
        f"<div style='font-size:20px; font-weight:700; color:#2ECC71;'>€{budget:,}</div></div>"
        "<div style='background:rgba(255,255,255,0.07); padding:14px; border-radius:10px;"
        "text-align:center; border:1px solid rgba(255,255,255,0.1);'>"
        "<div style='font-size:11px; opacity:0.6; margin-bottom:4px;'>SUPERFICIE</div>"
        f"<div style='font-size:20px; font-weight:700; color:#3498DB;'>{recommended_m2} m²</div></div>"
        "<div style='background:rgba(255,255,255,0.07); padding:14px; border-radius:10px;"
        "text-align:center; border:1px solid rgba(255,255,255,0.1);'>"
        "<div style='font-size:11px; opacity:0.6; margin-bottom:4px;'>HABITACIONES</div>"
        f"<div style='font-size:20px; font-weight:700; color:#E67E22;'>{bedrooms} dorm · {bathrooms} baños</div></div>"
        "<div style='background:rgba(255,255,255,0.07); padding:14px; border-radius:10px;"
        "text-align:center; border:1px solid rgba(255,255,255,0.1);'>"
        "<div style='font-size:11px; opacity:0.6; margin-bottom:4px;'>ESTILO</div>"
        f"<div style='font-size:14px; font-weight:700; color:#9B59B6;'>{selected_style.split(' ')[-1]}</div>"
        "</div></div>"
        "<div style='margin-bottom:8px; font-size:12px; opacity:0.6;'>EXTRAS INCLUIDOS</div>"
        f"<div style='margin-bottom:4px;'>{extras_html}</div>"
        f"{ahorro_html}"
        "</div>"
    )

    st.markdown(resumen_html, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        design_button = st.button(
            "🤖 DISEÑAR MI CASA CON IA",
            type="primary",
            use_container_width=True
        )
    
    if design_button:
        # Recopilar todos los datos
        req = {
            "budget": budget,
            "style": selected_style,
            "bedrooms": bedrooms,
            "bathrooms": bathrooms,
            "floors": floors,
            "target_area_m2": recommended_m2,
            "max_buildable_m2": max_buildable,
            "house_shape": house_shape,
            "roof_type": roof_type,
            "foundation_type": foundation_type,
            "water_systems": water_systems,
            "sewage_systems": sewage_systems,
            "water_details": {
                "municipal": water_municipal,
                "rain_tank": water_rain,
                "well": water_well,
                "solar_acs": water_solar
            },
            "sewage_details": {
                "municipal": sew_municipal,
                "septic": sew_septic,
                "phyto": sew_phyto,
                "biodigestor": sew_biodigestor
            },
            "extras": {
                "garage": has_garage,
                "pool": has_pool,
                "porch": has_porch,
                "bodega": has_bodega,
                "huerto": has_huerto,
                "caseta": has_caseta,
                "accessible": has_accessible,
                "office": has_office
            },
            "energy": {
                "solar": solar,
                "aerotermia": aerotermia,
                "geotermia": geotermia,
                "rainwater": rainwater,
                "insulation": insulation,
                "domotic": domotic
            },
            "special_notes": special_notes,
            "estimated_savings": ahorro_anual
        }
        
        st.session_state["ai_house_requirements"] = req
        
        # Llamar a IA
        _generate_ai_proposal(req)

def _normalize_ai_proposal(proposal: dict, energy_list: list) -> dict:
    """Normaliza los nombres y entradas del JSON generado por la IA.

    - Agrega o agrupa cualquier clave que mencione paneles o solar en una sola
      entrada con código **paneles_solares**.
    - Si el cliente pidió energía solar y no aparece ningún panel, inserta una
      entrada mínima con {PANEL_MIN_M2} m² para garantizar su visualización.
    - Devuelve el diccionario modificado (la operación se realiza in‑place).
    """
    panel_area = 0.0
    keys_to_remove = []

    for k, v in list(proposal.items()):
        if 'panel' in k.lower() or 'solar' in k.lower():
            try:
                panel_area += float(v)
            except Exception:
                # valores no numéricos se ignoran
                pass
            keys_to_remove.append(k)

    # quitar también claves de sistemas energéticos no espaciales
    # si el usuario marcó cualquier energía distinta de solar, eliminamos
    # posibles entradas generadas por la IA que mencionen el término.
    for en in energy_list:
        if en != 'solar':
            for k in list(proposal.keys()):
                if en.lower() in k.lower():
                    proposal.pop(k, None)

    # eliminar las claves antiguas asociadas a paneles
    for k in keys_to_remove:
        proposal.pop(k, None)

    if panel_area > 0:
        proposal['paneles_solares'] = panel_area
    elif 'solar' in energy_list:
        # el usuario quería paneles pero la IA no generó ninguno
        proposal['paneles_solares'] = PANEL_MIN_M2

    return proposal


def _generate_ai_proposal(req):
    """Genera propuesta de distribución con IA"""
    with st.spinner("🤖 La IA está diseñando tu casa..."):
        try:
            from dotenv import load_dotenv
            from groq import Groq
            from pathlib import Path
            import os
            import json
            
            # Cargar API key
            groq_api_key = _get_groq_key()

            if not groq_api_key:
                st.error("❌ GROQ_API_KEY no encontrada")
                return

            client = Groq(api_key=groq_api_key)
            
            # Construir prompt simplificado
            extras_list = [k for k, v in req['extras'].items() if v]
            energy_list = [k for k, v in req['energy'].items() if v]
            
            prompt = f"""Diseña una vivienda '{req['style']}'de {req['target_area_m2']}m² con:
- {req['bedrooms']} dormitorios (1 principal + {req['bedrooms']-1} secundarios)
- {req['bathrooms']} baños
- Extras: {', '.join(extras_list) if extras_list else 'ninguno'}
- Energía/Sostenibilidad: {', '.join(energy_list) if energy_list else 'ninguno'}
- **IMPORTANTE**: No transformes ninguna energía/sostenibilidad salvo los paneles en habitaciones. Es decir, no crees espacios llamados "aerotermia", "geotermia", "domótica", "rainwater", "insulation" ni similares. Esas tecnologías sólo deben aparecer en el análisis escrito, nunca en el plano ni como habitaciones.
- REGLA ESTRICTA: Si en la lista aparece "solar" (paneles solares), incluye EXACTAMENTE UNO con code "paneles_solares" en el JSON. NUNCA dos entradas de paneles.
- RESTRICCIÓN ECONÓMICA ABSOLUTA: La suma de TODOS los m² del JSON NO puede superar {req['target_area_m2']} m². El cliente tiene €{req['budget']:,} de presupuesto. Si no caben todos los extras, reduce sus m² proporcionalmente hasta que la suma total sea <= {req['target_area_m2']} m². Esto es innegociable.

PETICIONES ESPECIALES DEL CLIENTE (OBLIGATORIO INCLUIR):
{req.get('special_notes', 'ninguna')}

IMPORTANTE - Si el cliente menciona:
- "chimenea" → añadir coste €3,000-5,000 en descripción
- "suelo radiante" → añadir €80/m² extra al presupuesto
- "domótica" → añadir €5,000-15,000 en sistemas inteligentes
- Cualquier extra especial → incluirlo en el análisis y presupuesto
NUNCA ignorar las peticiones especiales del cliente.

Responde SOLO con un JSON válido (sin markdown) con habitaciones y m². Ejemplo:
{{
  "salon": 22,
  "cocina": 14,
  "dormitorio_principal": 16,
  "dormitorio": 11,
  "bano": 6,
  "bano_2": 5,
  "porche": 18,
  "garaje": 20
}}
"""
            
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3
            )
            
            # Parsear respuesta
            response_text = response.choices[0].message.content.strip()
            
            # Limpiar markdown si existe
            if "```json" in response_text:
                start = response_text.find("```json") + 7
                end = response_text.find("```", start)
                response_text = response_text[start:end].strip()
            elif "```" in response_text:
                start = response_text.find("```") + 3
                end = response_text.find("```", start)
                response_text = response_text[start:end].strip()
            
            ai_proposal = json.loads(response_text)

            # Normalizar resultados y asegurar paneles solares
            ai_proposal = _normalize_ai_proposal(ai_proposal, energy_list)
            
            # Guardar propuesta
            req["ai_room_proposal"] = ai_proposal
            st.session_state["ai_house_requirements"] = req
            
            st.success("✅ ¡Casa diseñada! Pasando al Paso 2...")
            st.session_state["ai_house_step"] = 2
            st.rerun()
            
        except Exception as e:
            st.error(f"❌ Error generando diseño: {e}")
            import traceback
            st.code(traceback.format_exc())

def render_step2():
    """Paso 2: Editor visual unificado - Layout 2 columnas profesional"""
    
    # Leer datos del proyecto para el banner (lectura segura, sin modificar)
    _req_header = st.session_state.get("ai_house_requirements", {})
    _budget_h = _req_header.get("budget", 0)
    _style_h = _req_header.get("style", "")
    _style_h_short = _style_h.split(" ")[-1] if _style_h else "—"
    _area_h = _req_header.get("target_area_m2", 0)
    _beds_h = _req_header.get("bedrooms", 0)
    _baths_h = _req_header.get("bathrooms", 0)

    _banner_html = (
        "<div style='background: linear-gradient(135deg, #1a1a2e 0%, #0f3460 100%);"
        "padding: 16px 24px; border-radius: 12px; color: white; margin-bottom: 16px;"
        "border: 1px solid rgba(255,255,255,0.1);'>"
        "<div style='display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:12px;'>"
        "<div>"
        "<div style='font-size:11px; opacity:0.5; letter-spacing:2px; text-transform:uppercase;'>Tu proyecto</div>"
        "<div style='font-size:20px; font-weight:700; margin-top:2px;'>Ajusta tu diseño</div>"
        "<div style='font-size:12px; opacity:0.6; margin-top:2px;'>El plano se actualiza automáticamente</div>"
        "</div>"
        "<div style='display:flex; gap:16px; flex-wrap:wrap;'>"
        f"<div style='text-align:center;'><div style='font-size:10px; opacity:0.5;'>PRESUPUESTO</div>"
        f"<div style='font-size:16px; font-weight:700; color:#2ECC71;'>€{_budget_h:,}</div></div>"
        f"<div style='text-align:center;'><div style='font-size:10px; opacity:0.5;'>SUPERFICIE</div>"
        f"<div style='font-size:16px; font-weight:700; color:#3498DB;'>{_area_h} m²</div></div>"
        f"<div style='text-align:center;'><div style='font-size:10px; opacity:0.5;'>HABITACIONES</div>"
        f"<div style='font-size:16px; font-weight:700; color:#E67E22;'>{_beds_h}d · {_baths_h}b</div></div>"
        f"<div style='text-align:center;'><div style='font-size:10px; opacity:0.5;'>ESTILO</div>"
        f"<div style='font-size:16px; font-weight:700; color:#9B59B6;'>{_style_h_short}</div></div>"
        "</div></div></div>"
    )
    st.markdown(_banner_html, unsafe_allow_html=True)
    
    # Botones navegación
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("← Paso 1", key="back_to_1", use_container_width=True):
            st.session_state["ai_house_step"] = 1
            st.rerun()
    with col2:
        if st.button("Paso 3: Editor 3D →", key="go_to_3", type="primary", use_container_width=True):
            st.session_state["ai_house_step"] = 3
            st.rerun()
    
    st.markdown("---")
    
    # ============================================
    # VALIDAR DATOS
    # ============================================
    req = st.session_state.get("ai_house_requirements", {})
    proposal = req.get("ai_room_proposal", {})
    
    if not proposal:
        st.warning("Primero completa el Paso 1")
        if st.button("← Volver al Paso 1"):
            st.session_state["ai_house_step"] = 1
            st.rerun()
        return
    
    # ================================================
    # USAR DATOS CENTRALIZADOS (referencia)
    # ================================================
    design_data = get_current_design_data()
    base_total_cost = design_data['total_cost']
    
    # Si hay diseño de Babylon, usar esas áreas en los sliders
    # El cable #076 ya sincronizó ai_room_proposal con new_area de Babylon
    if design_data['modified']:
        st.info(f"📐 **Diseño editado en 3D activo**: {design_data['total_area']}m² · "
                f"€{base_total_cost:,} · Los sliders reflejan tu diseño 3D")
    
    # proposal ya tiene los valores de Babylon gracias al cable #076
    # No necesitamos sobreescribir nada más aquí

    # Datos de parcela
    plot_data = st.session_state.get("design_plot_data", {})
    max_buildable = plot_data.get('buildable_m2', 400.0)
    
    # ============================================
    # TIPOS DE HABITACIÓN CON PRECIOS REALES
    # ============================================
    from .data_model import HouseDesign, Plot, RoomType, RoomInstance
    
    room_types = {
        "salon_cocina": RoomType(code="salon_cocina", name="Salón-Cocina", min_m2=20, max_m2=50, base_cost_per_m2=1500),
        "salon": RoomType(code="salon", name="Salón", min_m2=15, max_m2=40, base_cost_per_m2=1500),
        "cocina": RoomType(code="cocina", name="Cocina", min_m2=8, max_m2=20, base_cost_per_m2=1600),
        "dormitorio_principal": RoomType(code="dormitorio_principal", name="Dormitorio Principal", min_m2=12, max_m2=25, base_cost_per_m2=1600),
        "dormitorio": RoomType(code="dormitorio", name="Dormitorio", min_m2=8, max_m2=15, base_cost_per_m2=1400),
        "bano": RoomType(code="bano", name="Baño", min_m2=4, max_m2=8, base_cost_per_m2=1500),
        "bodega": RoomType(code="bodega", name="Bodega", min_m2=6, max_m2=20, base_cost_per_m2=1000),
        "piscina": RoomType(code="piscina", name="Piscina", min_m2=20, max_m2=60, base_cost_per_m2=1200),
        "paneles_solares": RoomType(code="paneles_solares", name="Paneles Solares", min_m2=3, max_m2=15, base_cost_per_m2=350),
        "garaje": RoomType(code="garaje", name="Garaje", min_m2=15, max_m2=40, base_cost_per_m2=1000),
        "porche": RoomType(code="porche", name="Porche/Terraza", min_m2=8, max_m2=40, base_cost_per_m2=700),
        "bomba_agua": RoomType(code="bomba_agua", name="Instalaciones", min_m2=2, max_m2=8, base_cost_per_m2=1500),
        "accesibilidad": RoomType(code="accesibilidad", name="Zona Accesible", min_m2=0, max_m2=10, base_cost_per_m2=1500),
        "pasillo": RoomType(code="pasillo", name="Pasillo/Distribuidor", min_m2=5, max_m2=20, base_cost_per_m2=1200),
        "huerto": RoomType(code="huerto", name="Huerto", min_m2=10, max_m2=100, base_cost_per_m2=40),
        "despacho": RoomType(code="despacho", name="Despacho", min_m2=8, max_m2=20, base_cost_per_m2=1400),
    }
    
    # ============================================
    # CREAR DISEÑO CON MAPEO INTELIGENTE
    # ============================================
    plot = Plot(
        id=plot_data.get('id', 'temp'),
        area_m2=plot_data.get('total_m2', 400.0),
        buildable_ratio=0.33
    )
    design = HouseDesign(plot)
    
    for code, area in proposal.items():
        if not isinstance(area, (int, float)):
            continue
        
        room_type = None
        code_lower = code.lower()
        
        if code in room_types:
            room_type = room_types[code]
        elif 'salon' in code_lower and 'cocina' not in code_lower:
            room_type = room_types['salon']
        elif 'cocina' in code_lower and 'salon' not in code_lower:
            room_type = room_types['cocina']
        elif 'salon' in code_lower and 'cocina' in code_lower:
            room_type = room_types['salon_cocina']
        elif 'dormitorio' in code_lower and 'principal' in code_lower:
            room_type = room_types['dormitorio_principal']
        elif 'dormitorio' in code_lower:
            room_type = room_types['dormitorio']
        elif 'bano' in code_lower or 'baño' in code_lower:
            room_type = room_types['bano']
        elif 'bodega' in code_lower:
            room_type = room_types['bodega']
        elif 'piscina' in code_lower:
            room_type = room_types['piscina']
        elif 'paneles' in code_lower or 'solar' in code_lower:
            room_type = room_types['paneles_solares']
        elif 'garaje' in code_lower or 'garage' in code_lower:
            room_type = room_types['garaje']
        elif 'porche' in code_lower or 'terraza' in code_lower:
            room_type = room_types['porche']
        elif 'bomba' in code_lower or 'instalac' in code_lower:
            room_type = room_types['bomba_agua']
        elif 'accesib' in code_lower:
            room_type = room_types['accesibilidad']
        elif 'pasillo' in code_lower or 'distrib' in code_lower:
            room_type = room_types['pasillo']
        elif 'huerto' in code_lower:
            room_type = room_types['huerto']
        elif 'despacho' in code_lower or 'oficina' in code_lower:
            room_type = room_types['despacho']
        else:
            room_type = RoomType(
                code=code,
                name=code.replace("_", " ").title(),
                min_m2=5, max_m2=50,
                base_cost_per_m2=1000
            )
        
        # Deduplicar: tipos que solo pueden aparecer una vez (evita dobles despacho/garaje/etc.)
        _SINGLE_TYPES = {
            'despacho', 'garaje', 'bodega', 'piscina', 'paneles_solares',
            'salon', 'cocina', 'salon_cocina', 'pasillo', 'porche',
            'huerto', 'bomba_agua', 'accesibilidad',
        }
        _existing_codes = [r.room_type.code for r in design.rooms]
        if room_type.code in _SINGLE_TYPES and room_type.code in _existing_codes:
            # Ya existe — mantener el de mayor área
            _idx = _existing_codes.index(room_type.code)
            if float(area) > design.rooms[_idx].area_m2:
                design.rooms[_idx].area_m2 = float(area)
        else:
            design.rooms.append(RoomInstance(
                room_type=room_type,
                area_m2=float(area)
            ))

    # ============================================
    # LAYOUT 2 COLUMNAS PRINCIPALES
    # ============================================
    col_left, col_right = st.columns([4, 6])
    
    # ============================================
    # COLUMNA IZQUIERDA: CONTROLES
    # ============================================
    with col_left:
        
        budget = req.get('budget', 150000)
        
        # Placeholder para métricas (se calculan después de sliders y checkboxes)
        metrics_placeholder = st.empty()
        
        st.markdown("---")
        
        # AJUSTAR HABITACIONES
        st.markdown("#### Ajustar Habitaciones")
        
        # Primero identificar qué extras están marcados para eliminar
        optional_codes = ['piscina', 'garaje', 'porche', 'bodega',
                         'huerto', 'paneles', 'bomba', 'accesib', 'despacho', 'office']
        
        # Pre-calcular qué rooms se van a eliminar
        preview_remove = []
        for i, room in enumerate(design.rooms):
            code = room.room_type.code.lower()
            if any(x in code for x in optional_codes):
                keep = st.session_state.get(f"keep_{i}", True)
                if not keep:
                    preview_remove.append(i)
        
        # Aviso antes de los sliders
        st.warning("""
        ⚠️ **DISEÑO PRELIMINAR**

        Los sliders te permiten ajustar áreas para estimar presupuesto.

        El diseño arquitectónico FINAL se crea en el **Editor 3D (Paso 3)**.

        Las medidas aquí son aproximadas y pueden variar en el editor 3D 
        debido a las dimensiones reales de construcción.
        """)
        
        # Mostrar sliders SOLO para rooms que NO se van a eliminar
        for i, room in enumerate(design.rooms):
            if room.area_m2 < 2:
                continue
            if i in preview_remove:
                continue  # No mostrar slider si va a ser eliminado
            
            cost_per_m2 = room.room_type.base_cost_per_m2
            
            # Leer valor previo desde session_state si existe
            slider_key = f"step2_slider_{i}"
            saved_area = st.session_state.get(slider_key, float(room.area_m2))
            saved_area = max(float(room.room_type.min_m2), 
                       min(float(room.room_type.max_m2), saved_area))

            # Mostrar PRIMERO: nombre y precio
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**{room.room_type.name}**")
            with col2:
                current_cost = saved_area * cost_per_m2
                st.markdown(f"<span style='color:#2ECC71; font-weight:bold;'>€{current_cost:,.0f}</span>", unsafe_allow_html=True)

            # Luego el slider
            new_area = st.slider(
                f"{saved_area:.1f} m²",
                min_value=float(room.room_type.min_m2),
                max_value=float(room.room_type.max_m2),
                value=saved_area,
                step=0.5,
                key=slider_key,
                label_visibility="collapsed"
            )

            # Calcular coste con valor ACTUAL del slider
            new_cost = new_area * cost_per_m2
            old_cost = float(room.area_m2) * cost_per_m2
            diff = new_cost - old_cost

            if abs(diff) > 10:
                if diff > 0:
                    st.caption(f"💰 +€{diff:,.0f}")
                else:
                    st.caption(f"💰 {diff:,.0f}")

            design.rooms[i].area_m2 = new_area
            
            # CABLE SLIDER → SESSION_STATE: persistir valor del slider
            req = st.session_state.get("ai_house_requirements", {})
            if "ai_room_proposal" in req:
                req["ai_room_proposal"][room.room_type.code] = new_area
                st.session_state["ai_house_requirements"] = req
        
        st.markdown("---")
        
        # ELIMINAR EXTRAS
        st.markdown("#### Eliminar Extras (Ahorrar)")
        
        optional_codes = ['piscina', 'garaje', 'porche', 'bodega',
                         'huerto', 'paneles', 'bomba', 'accesib', 'despacho', 'office']
        
        rooms_to_remove = []
        # Construir lista unificada: habitaciones opcionales + sistemas energéticos
        ENERGY_COSTS = {
            'aerotermia': ('🌡️ Aerotermia · €8,000',    8000),
            'geotermia':  ('🌍 Geotermia · €12,000',    12000),
            'rainwater':  ('💧 Agua Lluvia · €3,500',    3500),
            'insulation': ('🌿 Aislamiento · €2,000',    2000),
            'domotic':    ('🏠 Domótica · €5,000',       5000),
        }
        energy_selected = req.get('energy', {})
        energy_cost_total = 0
        energy_keep = {}

        # Grid 3 columnas — todos los extras juntos
        optional_rooms = [(i, room) for i, room in enumerate(design.rooms)
                          if any(x in room.room_type.code.lower() for x in optional_codes)]
        energy_items = [(k, label, cost) for k, (label, cost) in ENERGY_COSTS.items()
                        if energy_selected.get(k)]
        total_items = len(optional_rooms) + len(energy_items)
        cols = st.columns(3)
        col_idx = 0
        for i, room in optional_rooms:
            cost = room.area_m2 * room.room_type.base_cost_per_m2
            with cols[col_idx % 3]:
                keep = st.checkbox(
                    f"{room.room_type.name} · €{cost:,.0f}",
                    value=True,
                    key=f"keep_{i}"
                )
                if not keep:
                    rooms_to_remove.append(i)
            col_idx += 1
        for key, label, cost in energy_items:
            with cols[col_idx % 3]:
                keep_e = st.checkbox(label, value=True, key=f"keep_energy_{key}")
                energy_keep[key] = keep_e
                if keep_e:
                    energy_cost_total += cost
            col_idx += 1
        st.session_state['energy_cost_total'] = energy_cost_total
        st.session_state['energy_keep'] = energy_keep

        # ── EXTRAS DE ESTILO (determinados por el estilo elegido, no editables) ──
        _style_now = req.get('style', '')
        _STYLE_EXTRAS = {
            'Montaña': {'label': '🔥 Chimenea (Estilo Montaña)', 'cost': 4500},
            'Rural':   {'label': '🔥 Chimenea (Estilo Rural)',   'cost': 4500},
            'Clásico': {'label': '🔥 Chimenea (Estilo Clásico)', 'cost': 3500},
        }
        chimney_cost = 0
        if _style_now in _STYLE_EXTRAS:
            _se = _STYLE_EXTRAS[_style_now]
            chimney_cost = _se['cost']
            with cols[col_idx % 3]:
                st.markdown(
                    f"<div style='background:rgba(231,76,60,0.12); border:1px solid #E74C3C; "
                    f"border-radius:6px; padding:8px 10px; font-size:13px; color:white;'>"
                    f"<b>{_se['label']}</b><br>"
                    f"<span style='color:#E74C3C; font-weight:bold;'>€{_se['cost']:,}</span> "
                    f"<span style='font-size:11px; opacity:0.7;'>· incluido por estilo</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
        st.session_state['chimney_cost'] = chimney_cost

        # Eliminar desmarcados
        for idx in sorted(rooms_to_remove, reverse=True):
            design.rooms.pop(idx)

        # RECALCULAR MÉTRICAS FINALES (después de sliders y checkboxes)
        total_area_final = sum([r.area_m2 for r in design.rooms])
        total_cost_final = sum([r.area_m2 * r.room_type.base_cost_per_m2 for r in design.rooms])
        foundation_cost = int(total_area_final * 180)
        installation_cost = int(total_area_final * 150)
        energy_cost_total = st.session_state.get('energy_cost_total', 0)
        chimney_cost = st.session_state.get('chimney_cost', 0)
        total_with_extras = total_cost_final + foundation_cost + installation_cost + energy_cost_total + chimney_cost
        budget_pct = total_with_extras / budget * 100
        
        if budget_pct <= 90:
            b_icon = "✅"
            b_color = "normal"
        elif budget_pct <= 100:
            b_icon = "⚠️"
            b_color = "normal"
        else:
            b_icon = "❌"
            b_color = "inverse"
        
        # Rellenar placeholder con métricas actualizadas
        with metrics_placeholder.container():
            m1, m2 = st.columns(2)
            m1.metric(
                "Presupuesto",
                f"€{total_with_extras:,.0f}",
                delta=f"{b_icon} {budget_pct:.0f}% de €{budget:,.0f}",
                delta_color=b_color
            )
            m2.metric(
                "Superficie",
                f"{total_area_final:.0f} m²",
                delta=f"Máx: {max_buildable:.0f} m²"
            )
            savings = req.get('estimated_savings', 0)
            if savings > 0:
                st.success(f"Ahorro energético: €{savings:,}/año")
        
        # WARNING presupuesto superado
        if budget_pct > 110:
            exceso = int(total_with_extras - budget)
            st.error(
                f"🚨 **Presupuesto superado**: El diseño actual cuesta **€{total_with_extras:,}** "                f"pero tu presupuesto es **€{budget:,}**. "                f"Exceso: **€{exceso:,}**. "                f"Reduce habitaciones con los sliders o vuelve al Paso 1 para ajustar tu presupuesto."
            )
            if st.button("← Volver al Paso 1 y ajustar presupuesto", key="btn_back_budget"):
                st.session_state["ai_house_step"] = 1
                st.rerun()
        elif budget_pct > 100:
            exceso = int(total_with_extras - budget)
            st.warning(
                f"⚠️ El diseño actual supera ligeramente tu presupuesto en **€{exceso:,}**. "                f"Ajusta alguna habitación con los sliders."
            )
        
        st.markdown("---")
        
        if 'current_floor_plan' in st.session_state:
            st.download_button(
                label="Descargar Plano PNG",
                data=st.session_state['current_floor_plan'],
                file_name="plano_distribucion.png",
                mime="image/png",
                use_container_width=True
            )
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Paso 1", use_container_width=True):
                st.session_state["ai_house_step"] = 1
                st.rerun()
        with col2:
            if st.button("Paso 3 →", type="primary", use_container_width=True):
                st.session_state["ai_house_step"] = 3
                st.rerun()
    
    # ============================================
    # COLUMNA DERECHA: PLANO + IA
    # ============================================
    with col_right:
        
        # TABLA RESUMEN
        st.markdown("#### Tu Distribución")
        
        import pandas as pd
        table_data = []
        for room in design.rooms:
            price = room.room_type.base_cost_per_m2
            cost = room.area_m2 * price
            table_data.append({
                "Espacio": room.room_type.name,
                "m²": f"{room.area_m2:.0f}",
                "€/m²": f"€{price:,}",
                "Total": f"€{cost:,.0f}"
            })
        
        # Añadir cimentación e instalaciones
        table_data.append({
            "Espacio": "Cimentación",
            "m²": f"{total_area_final:.0f}",
            "€/m²": "€180",
            "Total": f"€{foundation_cost:,}"
        })
        table_data.append({
            "Espacio": "Instalaciones",
            "m²": f"{total_area_final:.0f}",
            "€/m²": "€150",
            "Total": f"€{installation_cost:,}"
        })
        if energy_cost_total > 0:
            table_data.append({
                "Espacio": "Sistemas sostenibles",
                "m²": "-",
                "€/m²": "-",
                "Total": f"€{energy_cost_total:,}"
            })
        
        df = pd.DataFrame(table_data)
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        st.markdown("---")

        # Plano 2D preliminar (sin pestañas)
        st.subheader("📐 Plano 2D Preliminar")

        # Firma del diseño actual — detecta si el plano está desactualizado
        _cur_sig = "|".join(
            f"{r.room_type.code}:{r.area_m2:.1f}"
            for r in sorted(design.rooms, key=lambda x: x.room_type.code)
        )

        if st.button("Generar Plano 2D", type="primary", use_container_width=True):
            try:
                from .floor_plan_svg import FloorPlanSVG
                planner = FloorPlanSVG(design)
                img_bytes = planner.generate()
                st.session_state['current_floor_plan'] = img_bytes
                st.session_state['current_design'] = design
                st.session_state['floor_plan_signature'] = _cur_sig
                st.success("Plano generado correctamente")
                st.rerun()
            except Exception as e:
                st.error(f"Error generando plano: {e}")
                import traceback
                st.code(traceback.format_exc())

        if 'current_floor_plan' in st.session_state:
            _saved_sig = st.session_state.get('floor_plan_signature', '')
            if _saved_sig != _cur_sig:
                st.warning("⚠️ Has modificado habitaciones o superficies. Pulsa **Generar Plano 2D** para actualizar el plano.")
            st.image(
                st.session_state['current_floor_plan'],
                caption="Plano profesional con medidas reales",
                use_container_width=True
            )
        else:
            st.info("Pulsa 'Generar Plano 2D' para ver tu distribución")
        
        st.markdown("---")
        
        # ANÁLISIS IA
        if 'current_floor_plan' in st.session_state:
            st.markdown("#### Análisis del Arquitecto IA")
            
            with st.spinner("Analizando..."):
                try:
                    from groq import Groq
                    import os
                    from dotenv import load_dotenv
                    from pathlib import Path
                    
                    client = Groq(api_key=_get_groq_key())
                    
                    rooms_summary = "\n".join([
                        f"- {r.room_type.name}: {r.area_m2:.0f} m² ({r.area_m2:.1f}m × {r.area_m2/max(r.area_m2**0.5,1):.1f}m aprox)"
                        for r in design.rooms
                    ])
                    ENERGY_LABELS = {
                        'aerotermia': 'Aerotermia', 'geotermia': 'Geotermia',
                        'rainwater': 'Recuperación agua lluvia',
                        'insulation': 'Aislamiento natural', 'domotic': 'Domótica',
                        'solar': 'Paneles solares',
                    }
                    energy_sel = req.get('energy', {})
                    systems_summary = ", ".join([
                        ENERGY_LABELS[k] for k in ENERGY_LABELS if energy_sel.get(k)
                    ]) or "ninguno"
                    
                    response = client.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[{"role": "user", "content": f"""Eres arquitecto experto en vivienda sostenible española.

Analiza esta distribución:
{rooms_summary}
TOTAL: {total_area_final:.0f} m²
PRESUPUESTO: €{budget:,}
ESTILO: {req.get('style', 'No especificado')}
SISTEMAS SOSTENIBLES SELECCIONADOS: {systems_summary}

Evalúa brevemente:
1. ¿Las medidas son correctas para una vivienda real? (mínimos CTE)
2. ¿Algún espacio demasiado grande o pequeño?
3. ¿El presupuesto es realista?
4. Una sugerencia de mejora concreta
5. En 1 frase por sistema: comenta profesionalmente cada sistema sostenible seleccionado indicando que se instalará en obra (no en el plano)

Máximo 200 palabras. Usa: ✅ para correcto, ⚠️ para mejorable, ❌ para problema."""}],
                        temperature=0.3,
                        max_tokens=300
                    )
                    
                    st.markdown(response.choices[0].message.content)
                
                except Exception as e:
                    st.warning(f"Análisis IA no disponible: {e}")

def render_step3_editor():
    """Paso 3: Editor 3D con Babylon.js"""
    
    st.header("Paso 3 – Diseña tu Casa en 3D")
    st.caption("Visualiza, ajusta y personaliza tu vivienda antes de construirla")
    
    # Validar datos
    req = st.session_state.get("ai_house_requirements", {})
    if not req or not req.get("ai_room_proposal"):
        st.warning("⚠️ Completa primero el Paso 1 y 2")
        if st.button("← Volver al Paso 2"):
            st.session_state["ai_house_step"] = 2
            st.rerun()
        return
    
    # Botón volver arriba
    if st.button("← Volver al Paso 2", key="back_to_2_top", use_container_width=True):
        st.session_state["ai_house_step"] = 2
        st.rerun()
    
    st.markdown("---")
    
    # Información del editor
    st.info("""
    🏠 **Tu casa en 3D — muévela, gírala, personalízala**

    - 🚶 **Vista Calle** — ve cómo queda desde el exterior
    - 🔝 **Vista Planta** — distribución desde arriba
    - 🖱️ Selecciona cualquier habitación para ver sus medidas
    - ⤢ Ajusta el tamaño de cada estancia
    - 🏠 Activa el tejado, los cimientos o el cerramiento
    - ✅ El sistema comprueba automáticamente la normativa CTE
    """)
    
    # Botón abrir editor - DESTACADO
    st.markdown("### 🏗️ Diseña tu casa")
    
    if st.button("🏠 Construir mi Casa — Ver en 3D", type="primary", use_container_width=True, key="open_babylon"):
        
        # Obtener forma de casa
        house_shape = st.session_state.get('request', {}).get('house_shape', 'Rectangular (más común)')
        
        from .babylon_editor import generate_babylon_html
        
        # Obtener layout
        from .architect_layout import generate_layout
        
        rooms_data = []
        for room_data in req.get("ai_room_proposal", {}).items():
            rooms_data.append({
                'code': room_data[0],
                'name': room_data[0],
                'area_m2': room_data[1]
            })
        
        layout_result = generate_layout(rooms_data, house_shape)
        
        # Calcular dimensiones
        if layout_result:
            all_x = [item['x'] + item['width'] for item in layout_result]
            all_z = [item['z'] + item['depth'] for item in layout_result]
            total_width = max(all_x)
            total_depth = max(all_z)
        else:
            total_width = 10
            total_depth = 10

        st.session_state["babylon_total_width"]   = total_width
        st.session_state["babylon_total_depth"]   = total_depth
        st.session_state["babylon_initial_layout"] = layout_result

        roof_type = st.session_state.get('request', {}).get('roof_type', 'Dos aguas (clásico, eficiente)')
        plot_data = st.session_state.get("design_plot_data", {})
        plot_area_m2 = float(plot_data.get('total_m2', 0) or 0)
        req_data = st.session_state.get("ai_house_requirements", {})
        foundation_type = req_data.get("foundation_type", "Losa de hormigón (suelos blandos)")
        house_style = req_data.get("style", "Moderno")
        _arch_cost_m2 = int(st.session_state.get("arch_cost_per_m2", 1600))
        html_editor = generate_babylon_html(layout_result, total_width, total_depth, roof_type, plot_area_m2, foundation_type, house_style, cost_per_m2=_arch_cost_m2)
        
        # Guardar HTML en session_state para renderizar embebido
        st.session_state["babylon_html"] = html_editor
        st.session_state["babylon_editor_used"] = True
        st.rerun()

    # Modo Estudio: botón para abrir Babylon en nueva pestaña (preserva contexto del portal)
    if st.session_state.get("estudio_mode") and st.session_state.get("babylon_html"):
        if st.button("🔗 Abrir Editor 3D en Nueva Pestaña", key="open_babylon_newtab", use_container_width=True):
            import base64 as _b64mod
            import streamlit.components.v1 as _cv1_bt
            _b64html = _b64mod.b64encode(st.session_state["babylon_html"].encode("utf-8")).decode("utf-8")
            _cv1_bt.html(
                f"""<script>
                var html = atob('{_b64html}');
                var blob = new Blob([html], {{type:'text/html'}});
                var url = URL.createObjectURL(blob);
                window.open(url, '_blank');
                </script>""",
                height=0,
            )

    # Renderizar editor embebido
    if st.session_state.get("babylon_html"):
        import streamlit.components.v1 as components
        st.info(
            "💡 Edita tu casa en 3D. "
            "Pulsa **📸 Capturar Vistas** para guardar 5 fotos (se descargan automáticamente). "
            "Pulsa **💾 Guardar JSON** para exportar el layout editado."
        )
        components.html(
            st.session_state["babylon_html"],
            height=700,
            scrolling=False
        )
    
    # ── Planos Técnicos MEP ──────────────────────────────────────────────────
    _mep_rooms = (st.session_state.get("babylon_modified_layout")
                  or st.session_state.get("babylon_initial_layout"))
    if _mep_rooms:
        with st.expander("📐 Planos Técnicos MEP — Descargar por capa", expanded=False):
            st.caption("Planos independientes para cada instalación. Generados automáticamente a partir del layout actual.")
            try:
                from .floor_plan_svg import generate_mep_plan_png as _gen_mep
                import json as _json
                if isinstance(_mep_rooms, str):
                    _mep_rooms = _json.loads(_mep_rooms)
                _tw = st.session_state.get("babylon_total_width")
                _td = st.session_state.get("babylon_total_depth")
                _mep_layers = [
                    ("sewage",     "🚽 Saneamiento"),
                    ("water",      "💧 Agua"),
                    ("electrical", "⚡ Eléctrico"),
                    ("rainwater",  "🌧️ Canalones"),
                    ("domotics",   "📡 Domótica"),
                ]
                _cols = st.columns(5)
                for _i, (_lid, _lbl) in enumerate(_mep_layers):
                    with _cols[_i]:
                        try:
                            _png = _gen_mep(_mep_rooms, _lid, _tw, _td)
                            if _png:
                                st.image(_png, use_container_width=True)
                                st.download_button(
                                    f"⬇️ {_lbl}", _png,
                                    file_name=f"plano_mep_{_lid}.png",
                                    mime="image/png",
                                    key=f"dl_mep_{_lid}",
                                    use_container_width=True,
                                )
                        except Exception as _le:
                            st.caption(f"⚠️ {_lbl}: {_le}")
            except Exception as _me:
                st.warning(f"No se pudieron generar los planos MEP: {_me}")

    # ── Cálculo CTE HS-5 Saneamiento ─────────────────────────────────────────
    if _mep_rooms:
        with st.expander("🚽 Instalación de Saneamiento — Cálculo CTE HS-5", expanded=False):
            try:
                from .mep_hs5 import render_mep_hs5_panel as _render_hs5
                _hs5_req = st.session_state.get("ai_house_requirements", {})
                _render_hs5(_mep_rooms, _hs5_req)
            except Exception as _hs5e:
                st.warning(f"No se pudo calcular el saneamiento: {_hs5e}")

    # Botón continuar DESPUÉS del editor
    st.markdown("### ✅ Ya terminé de diseñar")
    
    col1, col2 = st.columns([1, 2])
    with col2:
        if st.button("Continuar a Documentación (Paso 4) →", type="primary", use_container_width=True, key="go_to_4_bottom"):
            st.session_state["ai_house_step"] = 4
            st.rerun()


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _zip_images_dict(images_dict, thumb=False):
    """Return ZIP bytes containing png images taken from data‑URLs.

    If *thumb* is True the filenames get a ``_thumb`` suffix.
    """
    import io, zipfile, base64
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, dataurl in images_dict.items():
            parts = dataurl.split(',', 1)
            if len(parts) != 2:
                continue
            data = base64.b64decode(parts[1])
            fname = name + ("_thumb.png" if thumb else ".png")
            zf.writestr(fname, data)
    buf.seek(0)
    return buf.read()


def _process_babylon_return(editor_return: str):
    """Parse JSON returned by the Babylon editor component.

    *editor_return* is the string delivered to Streamlit by
    ``Streamlit.setComponentValue``.  It must contain a JSON object
    mapping view names to ``data:image/png;base64,...`` strings.

    Returns ``(captures_dict, thumbnails_dict)`` where ``thumbnails_dict``
    maps the same keys to base64-encoded PNG thumbnails.  An empty
    dict may be returned for thumbnails if the operation fails.
    """
    import json
    # editor_return may already be a dict if the JS sent an object.
    if isinstance(editor_return, str):
        caps = json.loads(editor_return)
    elif isinstance(editor_return, dict):
        caps = editor_return
    else:
        raise ValueError(f"Unexpected editor_return type: {type(editor_return)}")
    thumbs = {}
    try:
        from PIL import Image
        import io, base64
        for name, dataurl in caps.items():
            parts = dataurl.split(',', 1)
            if len(parts) != 2:
                continue
            imgdata = base64.b64decode(parts[1])
            img = Image.open(io.BytesIO(imgdata))
            img.thumbnail((300, 300))
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            thumbs[name] = base64.b64encode(buf.getvalue()).decode('ascii')
    except Exception:
        # if pillow isn't available or processing fails just ignore
        thumbs = {}
    return caps, thumbs


def render_step3():
    """Paso 3: Documentación completa y monetización"""

    st.header("Paso 4 – Tu Proyecto Completo")
    st.caption("Documentación técnica, eficiencia energética y siguiente paso.")

    # ── BLOQUE WOW: referencia visual del estilo + proyecto real más parecido ──
    try:
        _req_wow = st.session_state.get("ai_house_requirements", {})
        _style_wow = _req_wow.get("style", "Moderno")
        _area_wow = get_current_design_data().get("total_area", 0)

        import base64 as _b64s, os as _os2
        def _sw(fn):
            p = _os2.path.join("assets", "estilos", fn)
            if _os2.path.exists(p):
                with open(p, "rb") as _f: d = _b64s.b64encode(_f.read()).decode()
                ext = fn.rsplit(".",1)[-1].lower()
                return f"data:{'image/png' if ext=='png' else 'image/jpeg'};base64,{d}"
            return ""
        _STYLE_IMGS = {
            "Ecológico":     _sw("ecologico.jpg"),
            "Rural":         _sw("rural.jpg"),
            "Moderno":       _sw("moderno.jpg"),
            "Montaña":       _sw("montana.jpg"),
            "Playa":         _sw("playa.jpg"),
            "Clásico":       _sw("clasico.jpg"),
            "Andaluz":       _sw("andaluz.jpg"),
            "Contemporáneo": _sw("contemporaneo.jpg"),
        }
        _STYLE_DESCS = {
            "Ecológico": "Materiales naturales, mínimo impacto ambiental",
            "Rural": "Piedra, madera, integrado en el paisaje",
            "Moderno": "Líneas limpias, grandes ventanales, minimalista",
            "Montaña": "Refugio alpino, tejados inclinados, madera y piedra",
            "Playa": "Abierto, ventilado, colores claros, terrazas",
            "Clásico": "Elegante, simétrico, materiales nobles",
            "Andaluz": "Patio central, cerámica, cal, frescor natural",
            "Contemporáneo": "Vanguardista, tecnológico, sostenible",
        }

        col_wow1, col_wow2 = st.columns([3, 2])

        with col_wow1:
            st.markdown(f"### 🏠 Así podría quedar tu casa estilo **{_style_wow}**")
            _img_url = _STYLE_IMGS.get(_style_wow, _STYLE_IMGS["Moderno"])
            st.markdown(
                f'<img src="{_img_url}" style="width:100%;height:280px;object-fit:cover;'
                f'border-radius:12px;display:block;">',
                unsafe_allow_html=True
            )
            st.caption(f"✨ {_STYLE_DESCS.get(_style_wow, '')}  ·  Tu diseño: **{_area_wow:.0f} m²**")

        with col_wow2:
            st.markdown("### 📐 Proyecto profesional más parecido")
            try:
                from modules.marketplace.utils import db_conn as _db_conn
                from modules.marketplace.marketplace import get_project_display_image as _get_img
                import os as _os, base64 as _b64
                _conn_w = _db_conn()
                _cur_w = _conn_w.cursor()
                _cur_w.execute("""
                    SELECT id, title, m2_construidos, area_m2, price, architect_name
                    FROM projects
                    WHERE (m2_construidos > 0 OR area_m2 > 0)
                    ORDER BY ABS(COALESCE(m2_construidos, area_m2, 0) - ?) ASC
                    LIMIT 1
                """, (float(_area_wow),))
                _row_w = _cur_w.fetchone()
                _conn_w.close()

                if _row_w:
                    _pid, _ptitle, _pm2c, _pm2a, _pprice, _parch = _row_w
                    _pm2 = _pm2c or _pm2a or 0
                    _thumb_w = _get_img(_pid, image_type='main')
                    if isinstance(_thumb_w, str) and _os.path.exists(_thumb_w):
                        with open(_thumb_w, 'rb') as _f:
                            _raw = _f.read()
                        _ext = _thumb_w.rsplit('.', 1)[-1].lower()
                        _mime = "image/png" if _ext == "png" else "image/jpeg"
                        _b64str = _b64.b64encode(_raw).decode()
                        st.markdown(
                            f'<img src="data:{_mime};base64,{_b64str}" '
                            f'style="width:100%;height:150px;object-fit:cover;border-radius:8px;display:block;margin-bottom:8px;">',
                            unsafe_allow_html=True
                        )
                    st.markdown(f"**{_ptitle}**")
                    st.caption(f"📐 {_pm2:.0f} m²  ·  💰 €{_pprice:,.0f}" if _pprice else f"📐 {_pm2:.0f} m²")
                    if _parch:
                        st.caption(f"👨‍💼 {_parch}")
                    st.markdown(
                        '<div style="background:#EBF5FB;border-radius:6px;padding:8px 10px;'
                        'font-size:12px;color:#1A5276;margin:6px 0;">💡 Planos, memoria técnica y CAD listos para descargar.</div>',
                        unsafe_allow_html=True
                    )
                    if st.button("Ver proyecto completo →", key="wow_ver_proyecto", type="primary", use_container_width=True):
                        st.query_params["selected_project_v2"] = str(_pid)
                        st.rerun()
                else:
                    st.info("Explora proyectos en el marketplace.")
            except Exception:
                st.info("Explora proyectos en el marketplace.")
    except Exception:
        pass  # Bloque silencioso — si falla no rompe nada

    st.markdown("---")

    # Verificar si se usó el editor 3D

    # Verificar si se usó el editor 3D
    if st.session_state.get("babylon_editor_used", False):
        st.warning("""
        ⚠️ **DISEÑO MODIFICADO MANUALMENTE**
        
        Este proyecto ha sido editado con el editor 3D. Los cambios realizados:
        - Requieren validación por arquitecto colegiado
        - NO garantizan cumplimiento de normativa CTE
        - Pueden afectar presupuesto y plazos
        
        Nuestro equipo revisará el diseño antes de visar si contrata este servicio con nosotros.
        """)

    # Sincronización Babylon → Paso 4
    st.markdown("---")
    babylon_json = st.file_uploader(
        "📐 Si modificaste el diseño en el editor 3D, sube el JSON aquí:",
        type=['json'],
        key="babylon_sync_step4",
        help="El JSON se descarga al hacer 'Guardar Cambios' en Babylon"
    )

    if babylon_json:
        import json as _json
        try:
            babylon_data = _json.load(babylon_json)
            # Guardar en session_state
            st.session_state["babylon_modified_layout"] = babylon_data
            
            # CABLE BABYLON → ai_room_proposal
            # Actualizar propuesta con new_area de Babylon
            # para que los sliders del Paso 2 reflejen el diseño editado
            _req_sync = st.session_state.get("ai_house_requirements", {})
            if "ai_room_proposal" in _req_sync:
                for _room in babylon_data:
                    try:
                        _name = _room.get('name', '')
                        _new_area = float(_room.get('new_area', _room.get('original_area', 0)))
                        if _name and _new_area > 0:
                            _req_sync["ai_room_proposal"][ _name ] = round(_new_area, 1)
                    except (ValueError, TypeError):
                        continue
                st.session_state["ai_house_requirements"] = _req_sync
            
            # Recalcular con conversión segura
            total_area = 0
            original_area = 0

            _area_loaded = sum(float(r.get('new_area', r.get('original_area', 0))) for r in babylon_data)
            st.success(f"✅ Diseño cargado: {len(babylon_data)} habitaciones | {_area_loaded:.1f}m²")
            # NO hacer st.rerun() aquí
        except Exception as e:
            st.error(f"❌ Error: {e}")

    # Validar datos
    req = st.session_state.get("ai_house_requirements", {})
    proposal = req.get("ai_room_proposal", {})
    
    if not proposal:
        st.warning("Primero completa el Paso 1 y 2")
        if st.button("← Volver al inicio"):
            st.session_state["ai_house_step"] = 1
            st.rerun()
        return
    
    # Calcular datos del diseño
    budget = req.get('budget', 150000)
    style = req.get('style', 'No especificado')
    energy = req.get('energy', {})
    water_systems = req.get('water_systems', [])
    sewage_systems = req.get('sewage_systems', [])
    
    # Usar diseño FINAL via get_current_design_data (fuente única de verdad)
    design_data = get_current_design_data()
    total_area = design_data['total_area']
    rooms = design_data['rooms']
    # ── CAPTURAS 3D ──────────────────────────────────────────────────────────────
    st.markdown("### 📸 Vistas 3D de tu diseño")
    if st.session_state.get('babylon_captures'):
        caps_doc = st.session_state['babylon_captures']
        st.success(f"✅ {len(caps_doc)} vistas 3D vinculadas — se incluirán en tu proyecto")
        cols_caps = st.columns(min(len(caps_doc), 5))
        _cap_labels = {
            'sur_fachada_principal': 'Fachada Sur',
            'norte': 'Vista Norte',
            'este': 'Vista Este',
            'oeste': 'Vista Oeste',
            'planta_cenital': 'Planta Cenital',
        }
        for i, (k, dataurl) in enumerate(caps_doc.items()):
            with cols_caps[i % 5]:
                st.image(dataurl, caption=_cap_labels.get(k, k), use_container_width=True)
        zip_caps = _zip_images_dict(caps_doc, thumb=False)
        st.download_button(
            label="📁 Descargar vistas 3D (ZIP)",
            data=zip_caps,
            file_name="vistas_3d.zip",
            mime="application/zip"
        )
    else:
        st.info(
            "Si capturaste vistas en el editor 3D, sube aquí el ZIP para incluirlas "
            "en la documentación final y en la carpeta del proyecto."
        )
        uploaded_caps = st.file_uploader(
            "Sube el ZIP de capturas (Vistas_3D_ArchiRapid.zip) o imágenes PNG sueltas",
            type=["zip", "png"],
            accept_multiple_files=True,
            key="captures_doc_uploader"
        )
        if uploaded_caps:
            import zipfile, io, base64 as _b64
            _view_labels = ['sur_fachada_principal', 'norte', 'este', 'oeste', 'planta_cenital']
            captures = {}
            for f in uploaded_caps:
                if f.name.lower().endswith('.zip'):
                    with zipfile.ZipFile(io.BytesIO(f.read())) as zf:
                        png_names = sorted(n for n in zf.namelist() if n.lower().endswith('.png'))
                        for i, png_name in enumerate(png_names):
                            key = _view_labels[i] if i < len(_view_labels) else f"vista_{i+1}"
                            b64 = _b64.b64encode(zf.read(png_name)).decode()
                            captures[key] = f"data:image/png;base64,{b64}"
                elif f.name.lower().endswith('.png'):
                    i = len(captures)
                    key = _view_labels[i] if i < len(_view_labels) else f"vista_{i+1}"
                    b64 = _b64.b64encode(f.read()).decode()
                    captures[key] = f"data:image/png;base64,{b64}"
            if captures:
                st.session_state['babylon_captures'] = captures
                st.rerun()
    st.markdown("---")

    # Indicador visual de origen
    if design_data['modified']:
        st.success("🏗️ **Diseño desde Editor 3D** - Versión personalizada")
    else:
        st.info("🤖 **Diseño generado por IA** - Propuesta original")
    # ── Costes reales mercado 2025 (sincronizado con _get_financials) ──────────
    _BASE_M2_S3 = {
        "Moderno": 1500, "Mediterráneo": 1500, "Contemporáneo": 1550,
        "Ecológico": 1600, "Rural": 1650, "Montaña": 1750, "Clásico": 1800,
    }
    _FLOOR_FACTOR_S3 = {
        "1 Planta": 1.00, "2 Plantas": 1.05,
        "Planta Baja + Semisótano": 1.18, "2 Plantas + Semisótano": 1.22,
    }
    _floors_s3    = req.get("floors", "1 Planta")
    _base_m2_s3   = _BASE_M2_S3.get(req.get('style', 'Moderno'), 1500)
    _floor_fac_s3 = _FLOOR_FACTOR_S3.get(_floors_s3, 1.00)
    PEM_s3        = int(total_area * _base_m2_s3 * _floor_fac_s3)
    _honorarios_s3 = int(PEM_s3 * 0.09)
    _licencias_s3  = int(PEM_s3 * 0.04)
    _STYLE_CHIMNEY = {'Montaña': 4500, 'Rural': 4500, 'Clásico': 3500}
    chimney_cost  = _STYLE_CHIMNEY.get(req.get('style', ''), 0)
    # Variables de compatibilidad
    construction_cost = PEM_s3
    foundation_cost   = int(PEM_s3 * 0.10)
    installation_cost = int(PEM_s3 * 0.13)
    architecture_cost = _honorarios_s3 + _licencias_s3
    total_cost = int(PEM_s3 * 1.13) + 1200 + chimney_cost
    
    # Calcular subvenciones
    subsidy = 0
    if energy.get('solar'): subsidy += 3000
    if energy.get('aerotermia'): subsidy += 5000
    if energy.get('geotermia'): subsidy += 8000
    if energy.get('insulation'): subsidy += 2000
    if energy.get('rainwater'): subsidy += 1000
    if energy.get('domotic'): subsidy += 500
    subsidy_total = min(subsidy, int(total_cost * 0.40))
    
    # ============================================
    # RESUMEN EJECUTIVO - ARRIBA
    # ============================================
    st.markdown("### Tu Proyecto en Números")
    
    col1, col2, col3, col4 = st.columns(4)
    
    col1.metric(
        "Coste Total Estimado",
        f"€{total_cost:,}",
        delta=f"vs €{int(total_area * 3500):,} piso ciudad",
        delta_color="normal"
    )
    col2.metric(
        "Subvenciones Estimadas",
        f"€{subsidy_total:,}",
        delta=f"Coste neto: €{total_cost - subsidy_total:,}",
        delta_color="normal"
    )
    col3.metric(
        "Superficie Total",
        f"{total_area:.0f} m²",
        delta=f"Estilo: {style.split(' ')[-1]}"
    )
    col4.metric(
        "Ahorro vs Piso Ciudad",
        f"€{int(total_area * 3500) - (total_cost - subsidy_total):,}",
        delta="Ahorro real con subvenciones"
    )
    
    st.markdown("---")

    # ============================================
    # MAPA SATÉLITE CON HUELLA DE LA CASA
    # ============================================
    _plot_map = st.session_state.get("design_plot_data", {})
    _lat_m = _plot_map.get('lat')
    _lon_m = _plot_map.get('lon')

    if _lat_m and _lon_m:
        import folium as _folium
        import math as _math
        import streamlit.components.v1 as _cmp

        # Área real desde design_data (inmune a sobrescritura por bloque babylon_json)
        _total_area_map = design_data.get('total_area') or 80

        # Dimensiones reales desde Babylon (session_state), JSON subido, o aproximación
        _tw = st.session_state.get("babylon_total_width", 0)
        _td = st.session_state.get("babylon_total_depth", 0)
        _dim_source = "real"
        if not (_tw > 0 and _td > 0):
            _babylon_lay = st.session_state.get("babylon_modified_layout")
            if _babylon_lay:
                try:
                    _tw = max(r.get('x', 0) + r.get('width', 0) for r in _babylon_lay if r.get('width'))
                    _td = max(r.get('z', 0) + r.get('depth', 0) for r in _babylon_lay if r.get('depth'))
                except Exception:
                    _tw = _td = 0
        if not (_tw > 0 and _td > 0):
            _tw = _math.sqrt(_total_area_map * 1.3)
            _td = _math.sqrt(_total_area_map / 1.3)
            _dim_source = "estimada"

        # Metros → grados (WGS84)
        _dlat = _td / 111000
        _dlon = _tw / (111000 * _math.cos(_math.radians(float(_lat_m))))

        _m_map = _folium.Map(
            location=[float(_lat_m), float(_lon_m)],
            zoom_start=18,
            tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
            attr='Esri'
        )
        _folium.Rectangle(
            bounds=[
                [float(_lat_m) - _dlat / 2, float(_lon_m) - _dlon / 2],
                [float(_lat_m) + _dlat / 2, float(_lon_m) + _dlon / 2],
            ],
            color='#0055FF',
            fill=True,
            fill_color='#0055FF',
            fill_opacity=0.35,
            weight=2,
            tooltip=f"Tu casa: {_tw:.1f}m × {_td:.1f}m ({_total_area_map:.0f} m²)"
        ).add_to(_m_map)
        _folium.Marker(
            [float(_lat_m), float(_lon_m)],
            icon=_folium.Icon(color='red', icon='home', prefix='fa'),
            tooltip=_plot_map.get('title', 'Tu parcela')
        ).add_to(_m_map)

        st.markdown("### Ubicación de tu proyecto")
        _cmp.html(_m_map._repr_html_(), height=420)
        st.caption(f"Huella {_dim_source} de la casa ({_tw:.1f}m × {_td:.1f}m) sobre imagen satélite de la parcela.")
        st.markdown("---")

    # ============================================
    # DOS COLUMNAS: ENERGÍA + PRESUPUESTO
    # ============================================
    col_left, col_right = st.columns([1, 1])
    
    with col_left:
        
        # EFICIENCIA ENERGÉTICA
        st.markdown("### Eficiencia Energética")
        
        # Calcular calificación
        energy_score = 0
        if energy.get('solar'): energy_score += 25
        if energy.get('aerotermia'): energy_score += 20
        if energy.get('geotermia'): energy_score += 20
        if energy.get('insulation'): energy_score += 15
        if energy.get('rainwater'): energy_score += 10
        if energy.get('domotic'): energy_score += 10
        
        if energy_score >= 60:
            rating = "A"
            rating_color = "#2ECC71"
            rating_text = "Máxima eficiencia energética"
        elif energy_score >= 40:
            rating = "B"
            rating_color = "#27AE60"
            rating_text = "Alta eficiencia energética"
        elif energy_score >= 20:
            rating = "C"
            rating_color = "#F39C12"
            rating_text = "Eficiencia media"
        else:
            rating = "D"
            rating_color = "#E74C3C"
            rating_text = "Eficiencia básica"
        
        # Badge de calificación
        st.markdown(f"""
        <div style='background: {rating_color}; padding: 20px; border-radius: 10px; 
                    text-align: center; color: white; margin-bottom: 15px;'>
            <h1 style='margin:0; font-size: 60px;'>{rating}</h1>
            <p style='margin:0; font-size: 16px;'>{rating_text}</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Detalles energéticos
        consumo_base = total_area * 120  # kWh/año casa normal
        ahorro_pct = energy_score * 0.9
        consumo_real = int(consumo_base * (1 - ahorro_pct/100))
        ahorro_euros = int((consumo_base - consumo_real) * 0.18)
        co2_evitado = int((consumo_base - consumo_real) * 0.25 / 1000 * 10) / 10
        
        e1, e2 = st.columns(2)
        e1.metric("Consumo estimado", f"{consumo_real:,} kWh/año",
                 delta=f"-{ahorro_pct:.0f}% vs media")
        e2.metric("Ahorro energético", f"€{ahorro_euros:,}/año",
                 delta="vs casa convencional")
        
        st.metric("CO₂ evitado", f"{co2_evitado} ton/año",
                 delta="Contribución al medioambiente")
        
        # Sistemas instalados
        st.markdown("**Sistemas sostenibles incluidos:**")
        systems = []
        if energy.get('solar'): systems.append("☀️ Paneles solares fotovoltaicos")
        if energy.get('aerotermia'): systems.append("🌡️ Aerotermia (calefacción/frío)")
        if energy.get('geotermia'): systems.append("🌍 Geotermia")
        if energy.get('insulation'): systems.append("🌿 Aislamiento natural")
        if energy.get('rainwater'): systems.append("💧 Recuperación agua lluvia")
        if energy.get('domotic'): systems.append("🏠 Domótica inteligente")
        if water_systems: systems.append(f"💧 Agua: {' + '.join(water_systems)}")
        if sewage_systems: systems.append(f"♻️ Saneamiento: {' + '.join(sewage_systems)}")
        
        for s in systems:
            st.markdown(f"- {s}")
        
        st.markdown("---")
        
        # SUBVENCIONES
        st.markdown("### Subvenciones Disponibles")
        
        subsidy_data = []
        if energy.get('solar'):
            subsidy_data.append(("☀️ Paneles solares (IDAE)", "€3,000"))
        if energy.get('aerotermia'):
            subsidy_data.append(("🌡️ Aerotermia (NextGen EU)", "€5,000"))
        if energy.get('geotermia'):
            subsidy_data.append(("🌍 Geotermia (NextGen EU)", "€8,000"))
        if energy.get('insulation'):
            subsidy_data.append(("🌿 Aislamiento (IDAE)", "€2,000"))
        if energy.get('rainwater'):
            subsidy_data.append(("💧 Agua lluvia (CC.AA.)", "€1,000"))
        
        if subsidy_data:
            for name, amount in subsidy_data:
                col_s1, col_s2 = st.columns([3, 1])
                col_s1.markdown(f"- {name}")
                col_s2.markdown(f"**{amount}**")
            
            st.success(f"Total estimado: **€{subsidy_total:,}** (hasta 40% del coste)")
            st.caption("Sujeto a convocatorias vigentes. Consulta con nuestro equipo.")
        else:
            st.info("Activa sistemas sostenibles en el Paso 1 para acceder a subvenciones")
    
    with col_right:
        
        # PRESUPUESTO DETALLADO
        st.markdown("### Presupuesto por Partidas")
        
        import pandas as pd
        
        _style_s3 = req.get('style', 'Moderno')
        partidas = [
            ("1. Geotécnico y topografía",  f"€1.200",              "0%",  "Estudio del terreno previo a proyecto"),
            ("2. Movimiento de tierras",     f"€{int(PEM_s3*0.03):,}", "3%", "Excavación, nivelación y transporte"),
            ("3. Cimentación",              f"€{int(PEM_s3*0.10):,}", "10%", "Zapatas / losa según geotécnico"),
            ("4. Estructura y forjados",     f"€{int(PEM_s3*0.20):,}", "20%", "Hormigón armado, pilares, forjado"),
            ("5. Cerramientos y fachada",    f"€{int(PEM_s3*0.12):,}", "12%", f"Fachada estilo {_style_s3}"),
            ("6. Cubierta",                 f"€{int(PEM_s3*0.07):,}", "7%",  f"Cubierta {req.get('roof_type','inclinada')}"),
            ("7. Carpintería exterior",      f"€{int(PEM_s3*0.06):,}", "6%",  "Ventanas, puertas acceso, persianas"),
            ("8. Particiones interiores",    f"€{int(PEM_s3*0.05):,}", "5%",  "Tabiquería, trasdosados, pladur"),
            ("9. Instalaciones",            f"€{int(PEM_s3*0.13):,}", "13%", "Eléctrica, fontanería, climatización"),
            ("10. Acabados interiores",      f"€{int(PEM_s3*0.12):,}", "12%", "Suelos, techos, pintura interior"),
            ("11. Baños y cocina",           f"€{int(PEM_s3*0.05):,}", "5%",  "Sanitarios, grifería, equip. básico cocina"),
            ("12. Urbanización parcela",     f"€{int(PEM_s3*0.03):,}", "3%",  "Acceso, cerramiento, jardinería básica"),
            ("— Honorarios técnicos",       f"€{_honorarios_s3:,}",   "9%",  "Arquitecto, aparejador, coord. seguridad"),
            ("— Licencias y tasas",         f"€{_licencias_s3:,}",    "4%",  "Licencia de obras e ICIO municipal"),
        ]
        if chimney_cost > 0:
            partidas.append((
                "★ Chimenea / hogar",
                f"€{chimney_cost:,}", "—",
                f"Chimenea de leña/biomasa — estilo {_style_s3}"
            ))
        
        df = pd.DataFrame(partidas, 
                         columns=["Partida", "Coste", "%", "Descripción"])
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        # Total
        st.markdown(f"""
        <div style='background: #2C3E50; padding: 15px; border-radius: 8px; 
                    color: white; text-align: center;'>
            <h3 style='margin:0;'>TOTAL: €{total_cost:,}</h3>
            <p style='margin:5px 0 0 0; font-size:14px;'>
                Neto con subvenciones: €{total_cost - subsidy_total:,}
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # PLANO FINAL (siempre desde get_final_design)
        st.markdown("### Tu Plano")

        final_rooms_for_plan = get_final_design()
        plan_source = final_rooms_for_plan.get('source', 'ai_original')
        plan_label = {
            'babylon': '🏗️ Plano desde Editor 3D',
            'step2_sliders': '📊 Plano con ajustes del Paso 2',
            'ai_original': '🤖 Plano propuesta IA'
        }.get(plan_source, 'Plano de distribución')

        if st.button("🗺️ Generar Plano Final", type="primary", use_container_width=True, key="gen_plan_paso4"):
            try:
                from .architect_layout import generate_layout
                from .floor_plan_svg import FloorPlanSVG
                from .data_model import HouseDesign, Plot, RoomType, RoomInstance

                house_shape = st.session_state.get('request', {}).get('house_shape', 'Rectangular')
                rooms_for_layout = []
                for r in final_rooms_for_plan['rooms']:
                    rooms_for_layout.append({
                        'code': r.get('code', r.get('name', 'espacio')),
                        'name': r.get('name', 'Espacio'),
                        'area_m2': float(r.get('area_m2', 10))
                    })

                layout = generate_layout(rooms_for_layout, house_shape)

                plot = Plot(id='final', area_m2=500, buildable_ratio=0.33)
                design_for_plan = HouseDesign(plot)
                for r in final_rooms_for_plan['rooms']:
                    rt = RoomType(
                        code=r.get('code', 'espacio'),
                        name=r.get('name', 'Espacio'),
                        min_m2=2, max_m2=200,
                        base_cost_per_m2=1000
                    )
                    design_for_plan.rooms.append(RoomInstance(room_type=rt, area_m2=float(r.get('area_m2', 10))))

                planner = FloorPlanSVG(design_for_plan)
                img_bytes = planner.generate()
                st.session_state['final_floor_plan'] = img_bytes
                st.success(f"✅ Plano generado desde: {plan_label}")

            except Exception as e:
                st.error(f"❌ Error generando plano: {e}")
                import traceback
                st.code(traceback.format_exc())

        if 'final_floor_plan' in st.session_state:
            st.image(
                st.session_state['final_floor_plan'],
                caption=plan_label,
                use_container_width=True
            )
        elif 'current_floor_plan' in st.session_state:
            st.image(
                st.session_state['current_floor_plan'],
                caption="Plano del Paso 2 (genera el plano final arriba)",
                use_container_width=True
            )
        else:
            st.info("Pulsa 'Generar Plano Final' para ver el plano correcto")

        st.markdown("---")
        
        # DESCARGAS
        st.markdown("### Descargas")

        # mostrar thumbnails si hay capturas previas
        if st.session_state.get('babylon_captures'):
            st.markdown("#### 📷 Vistas 3D capturadas")
            try:
                st.image(list(st.session_state['babylon_captures'].values()), width=100)
            except Exception:
                pass
            # descarga rápida de todas las vistas
            zip_all = _zip_images_dict(st.session_state['babylon_captures'], thumb=False)
            st.download_button(
                label="📁 Descargar vistas 3D (ZIP)",
                data=zip_all,
                file_name="vistas_3d.zip",
                mime="application/zip",
                use_container_width=True
            )
            if st.session_state.get('babylon_captures_thumb'):
                zip_th = _zip_images_dict({k: f"data:image/png;base64,{b64}" for k,b64 in st.session_state['babylon_captures_thumb'].items()}, thumb=True)
                st.download_button(
                    label="📁 Descargar miniaturas (ZIP)",
                    data=zip_th,
                    file_name="miniaturas_3d.zip",
                    mime="application/zip",
                    use_container_width=True
                )

        dl1, dl2 = st.columns(2)
        
        with dl1:
            plan_to_download = st.session_state.get('final_floor_plan') or st.session_state.get('current_floor_plan')
            if plan_to_download:
                st.download_button(
                    label="Descargar Plano PNG",
                    data=plan_to_download,
                    file_name="plano_archirapid_final.png",
                    mime="image/png",
                    use_container_width=True
                )
        
        with dl2:
            # Excel presupuesto
            try:
                import io
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Presupuesto"
                
                # Cabecera
                ws['A1'] = "PRESUPUESTO ARCHIRAPID"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A2'] = f"Superficie: {total_area:.0f} m² | Estilo: {style}"
                
                # Headers
                headers = ["Partida", "Coste (€)", "% Total", "Descripción"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="2C3E50")
                
                # Datos
                for row, (partida, coste, pct, desc) in enumerate(partidas, 5):
                    ws.cell(row=row, column=1, value=partida)
                    ws.cell(row=row, column=2, value=coste)
                    ws.cell(row=row, column=3, value=pct)
                    ws.cell(row=row, column=4, value=desc)
                
                # Total
                ws.cell(row=len(partidas)+6, column=1, value="TOTAL").font = Font(bold=True)
                ws.cell(row=len(partidas)+6, column=2, value=f"€{total_cost:,}").font = Font(bold=True)
                ws.cell(row=len(partidas)+7, column=1, value="Con subvenciones").font = Font(bold=True)
                ws.cell(row=len(partidas)+7, column=2, value=f"€{total_cost-subsidy_total:,}").font = Font(bold=True, color="2ECC71")
                
                # Ajustar columnas
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 45
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="Descargar Excel",
                    data=excel_buffer.getvalue(),
                    file_name="presupuesto_archirapid.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.warning(f"Excel no disponible: {e}")
    
    st.markdown("---")

    # ── VISTA 3D BIM — preview gratuita antes del pago ───────────────────────
    try:
        from .viewer3d import Viewer3D
        from .data_model import HouseDesign, Plot, RoomType, RoomInstance
        import streamlit.components.v1 as _cmp_v3d
        from modules.ai_house_designer.ifc_export import rooms_to_svg as _rooms_to_svg

        _final_d_v3 = get_final_design()
        _rooms_v3 = _final_d_v3.get('rooms', [])
        if _rooms_v3:
            st.markdown("### 🏗️ Tu Casa en 3D — Vista Previa BIM")
            st.caption("Visualización interactiva de tu diseño · Arrastra para rotar · Rueda para zoom")

            _col_3d, _col_svg = st.columns([3, 2])

            with _col_3d:
                _plot_v3 = Plot(id='preview', area_m2=500, buildable_ratio=0.33)
                _design_v3 = HouseDesign(_plot_v3)
                _roof_v3 = req.get('roof_type', 'Dos aguas (clásico, eficiente)')
                _shape_v3 = req.get('house_shape', 'Rectangular (más común)')
                setattr(_design_v3, 'request', {'house_shape': _shape_v3, 'roof_type': _roof_v3})
                for _r3 in _rooms_v3:
                    _rt3 = RoomType(
                        code=_r3.get('code', _r3.get('name', 'espacio')),
                        name=_r3.get('name', 'Espacio'),
                        min_m2=2, max_m2=200, base_cost_per_m2=1000
                    )
                    _design_v3.rooms.append(
                        RoomInstance(room_type=_rt3, area_m2=float(_r3.get('area_m2', 10)))
                    )
                _viewer_v3 = Viewer3D(_design_v3, roof_type=_roof_v3)
                _html_v3d = _viewer_v3.generate_html()
                _cmp_v3d.html(_html_v3d, height=520, scrolling=False)

            with _col_svg:
                _bim_src_v3 = (
                    st.session_state.get("babylon_modified_layout")
                    or [{"name": k, "area": float(v)}
                        for k, v in proposal.items()
                        if isinstance(v, (int, float))]
                )
                _svg_bim = _rooms_to_svg(_bim_src_v3, px=360)
                if _svg_bim:
                    st.markdown("**📐 Plano de distribución BIM**")
                    st.markdown(_svg_bim, unsafe_allow_html=True)
                    st.caption("Cada espacio = objeto IFC2x3 certificable · Compatible UE BIM Mandate")

            st.markdown("---")
    except Exception:
        pass  # nunca interrumpe el flujo principal

    # ============================================
    # CONSTRUCTORES ASOCIADOS - MONETIZACIÓN
    # ============================================
    st.markdown("### Constructores Asociados")
    st.caption("Solicita presupuesto real a constructores verificados de tu zona")
    
    constructors = [
        {
            "name": "Construcciones EcoVerde",
            "specialty": "Vivienda sostenible y bioconstrucción",
            "rating": "⭐⭐⭐⭐⭐",
            "projects": "127 proyectos",
            "zone": "Andalucía, Extremadura"
        },
        {
            "name": "BuildGreen España",
            "specialty": "Casas pasivas y certificación energética",
            "rating": "⭐⭐⭐⭐⭐",
            "projects": "89 proyectos",
            "zone": "Nacional"
        },
        {
            "name": "Construye Rural",
            "specialty": "Vivienda rural, piedra natural y madera",
            "rating": "⭐⭐⭐⭐",
            "projects": "203 proyectos",
            "zone": "Todo el territorio"
        }
    ]
    
    cols = st.columns(3)
    for col, constructor in zip(cols, constructors):
        with col:
            st.markdown(f"""
            <div style='border: 1px solid #ddd; border-radius: 10px; 
                        padding: 15px; height: 200px;'>
                <h4 style='color: #2C3E50; margin-top:0;'>{constructor["name"]}</h4>
                <p style='font-size:12px; color:#666;'>{constructor["specialty"]}</p>
                <p>{constructor["rating"]}</p>
                <p style='font-size:11px;'>📁 {constructor["projects"]}</p>
                <p style='font-size:11px;'>📍 {constructor["zone"]}</p>
            </div>
            """, unsafe_allow_html=True)
            
            st.button(
                f"Solicitar Presupuesto",
                key=f"constructor_{constructor['name']}",
                use_container_width=True,
                type="primary"
            )
    
    st.info("ArchiRapid cobra una comisión del 3% al constructor cuando se formaliza el contrato. Sin coste para el cliente.")
    
    st.markdown("---")
    
    # NAVEGACIÓN
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Volver al Paso 2", use_container_width=True):
            st.session_state["ai_house_step"] = 2
            st.rerun()
    with col2:
        if st.button("✅ He terminado el diseño — Ver opciones de proyecto", type="primary", use_container_width=True):
            st.session_state["mostrar_monetizacion"] = True
            st.rerun()

    # ================================================
    # BLOQUE MONETIZACIÓN — solo visible tras aprobar
    # ================================================
    if st.session_state.get("mostrar_monetizacion", False):

        st.markdown("---")
        st.markdown("## 📋 Tu Proyecto Está Listo")
        st.success("✅ Diseño aprobado. Selecciona los servicios que necesitas.")

        # Leer presupuesto real del proyecto
        req = st.session_state.get("ai_house_requirements", {})
        design_data = get_current_design_data()
        presupuesto_obra = design_data.get('total_cost', 200000)

        # ------------------------------------------
        # DOCUMENTACIÓN BASE
        # ------------------------------------------
        st.markdown("### 📄 Documentación del Proyecto")
        col_doc1, col_doc2 = st.columns(2)

        with col_doc1:
            st.markdown("""
            **📦 Paquete PDF Completo** *(1ª copia gratis)*
            - Memoria descriptiva
            - Plano 2D
            - Presupuesto por partidas
            - Calificación energética
            - Datos cimientos y extras
            """)
            pdf_copias = st.number_input("Copias adicionales PDF (150€/copia)", min_value=0, max_value=10, value=0, key="pdf_copias")

        with col_doc2:
            st.markdown("""
            **📐 Paquete CAD/BIM Profesional**
            - Todo lo del PDF
            - Archivos editables CAD
            - Modelo 3D exportable (.glb)
            - Apto para constructor y arquitecto
            """)
            cad_copias = st.number_input("Copias CAD (250€/copia)", min_value=0, max_value=10, value=0, key="cad_copias")

        coste_doc = int(presupuesto_obra * 0.01) + (pdf_copias * 150) + (cad_copias * 250)

        st.info(f"💰 **Proyecto completo (1% presupuesto): €{int(presupuesto_obra * 0.01):,}** + copias adicionales: €{(pdf_copias*150)+(cad_copias*250):,}")

        # ------------------------------------------
        # SERVICIOS TÉCNICOS
        # ------------------------------------------
        st.markdown("### 🔧 Servicios Técnicos Profesionales")
        st.caption("Todos gestionados por ArchiRapid con profesionales colegiados")

        servicios = {
            "visado": {"label": "📋 Visado del Proyecto (Colegio Arquitectos)", "precio": int(presupuesto_obra * 0.015), "desc": "Obligatorio para licencia de obra. ~1.5% presupuesto"},
            "revision": {"label": "🔍 Revisión y Validación Técnica", "precio": 450, "desc": "Arquitecto revisa y firma el diseño ArchiRapid"},
            "licencia": {"label": "🏛️ Gestión Licencia Municipal", "precio": 800, "desc": "Tramitación completa ante el Ayuntamiento"},
            "constructor": {"label": "👷 Búsqueda y Contrato Constructor", "precio": int(presupuesto_obra * 0.005), "desc": f"0.5% del presupuesto. Selección, negociación y contrato"},
            "fontaneria": {"label": "🚿 Proyecto Fontanería e Instalaciones", "precio": 650, "desc": "Planos de agua, saneamiento y climatización"},
            "electricidad": {"label": "⚡ Proyecto Eléctrico BT", "precio": 550, "desc": "Esquema unifilar, cuadro eléctrico, domótica"},
            "energia": {"label": "☀️ Auditoría Energética + CEE", "precio": 380, "desc": "Certificado Eficiencia Energética oficial"},
            "geotecnico": {"label": "🔬 Estudio Geotécnico del Terreno", "precio": 1200, "desc": "Obligatorio para cálculo de cimientos"},
            "topografia": {"label": "📏 Levantamiento Topográfico", "precio": 900, "desc": "Medición precisa de la parcela"},
            "seguridad": {"label": "🦺 Estudio Seguridad y Salud", "precio": 350, "desc": "Obligatorio para obras con proyecto"},
            "vr": {"label": "🥽 Tour Virtual 360° / Realidad Virtual", "precio": 290, "desc": "Recorrido inmersivo de tu vivienda antes de construir"},
        }

        seleccionados = {}
        coste_servicios = 0

        cols_svc = st.columns(2)
        for i, (key, svc) in enumerate(servicios.items()):
            col = cols_svc[i % 2]
            with col:
                checked = st.checkbox(
                    f"{svc['label']} — **€{svc['precio']:,}**",
                    key=f"svc_{key}",
                    help=svc['desc']
                )
                if checked:
                    seleccionados[key] = svc['precio']
                    coste_servicios += svc['precio']

        # ------------------------------------------
        # RESUMEN TOTAL
        # ------------------------------------------
        st.markdown("---")
        total_monetizacion = coste_doc + coste_servicios

        st.markdown(f"""
        <div style='background: linear-gradient(135deg, #1a1a2e, #2C3E50);
                    padding: 20px; border-radius: 12px; color: white; text-align: center;
                    border: 2px solid #27AE60;'>
            <h3 style='margin:0; color:#27AE60;'>TOTAL SERVICIOS SELECCIONADOS</h3>
            <h1 style='margin:10px 0; color:white;'>€{total_monetizacion:,}</h1>
            <p style='margin:0; font-size:13px; color:#aaa;'>
                Documentación: €{coste_doc:,} | Servicios técnicos: €{coste_servicios:,}
            </p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        col_pay1, col_pay2 = st.columns(2)
        with col_pay1:
            if st.button("💳 Pagar con Tarjeta (Simulado)", type="primary", use_container_width=True):
                st.session_state["pago_completado"] = True
                st.session_state["servicios_contratados"] = seleccionados
                st.session_state["total_pagado"] = total_monetizacion
                st.rerun()
        with col_pay2:
            if st.button("📞 Hablar con un Arquitecto", use_container_width=True):
                st.info("📧 Contacto: info@archirapid.es | ☎️ 900 XXX XXX")

        # ------------------------------------------
        # POST-PAGO — desbloquear descargas
        # ------------------------------------------
        if st.session_state.get("pago_completado", False):
            st.balloons()
            st.success(f"✅ Pago simulado de €{st.session_state.get('total_pagado',0):,} procesado. ¡Tu proyecto está listo!")

            # ── Calculadora de hipoteca con datos del diseño ─────────────────
            try:
                from modules.marketplace.hipoteca import render_calculadora
                _req_h    = st.session_state.get("ai_house_requirements", {})
                _plot_h   = st.session_state.get("design_plot_data", {})
                _precio_t = float(_plot_h.get("price") or 0)
                _m2_h     = float(_req_h.get("total_m2") or _req_h.get("m2_construidos") or 80)
                _style_h  = (_req_h.get("style") or "").lower()
                _coste_m2 = 1800 if "premium" in _style_h or "lujo" in _style_h else (1200 if "eco" in _style_h or "madera" in _style_h else 1400)
                _coste_obra = _m2_h * _coste_m2
                with st.expander("🏦 Calculadora de Financiación — ¿Cuánto pagarías al mes?", expanded=True):
                    render_calculadora(
                        precio_terreno=_precio_t,
                        coste_construccion=_coste_obra,
                        key_prefix="flow5"
                    )
            except Exception:
                pass

            st.markdown("### 📥 Descarga Tu Proyecto Completo")

            try:
                zip_bytes, zip_filename = generar_zip_proyecto(
                    req=st.session_state.get("ai_house_requirements", {}),
                    design_data=get_current_design_data(),
                    plot_data=st.session_state.get("design_plot_data", {}),
                    partidas=partidas,
                    subsidy_total=subsidy_total,
                    energy_label=rating if 'rating' in dir() else "B"
                )

                st.download_button(
                    label="📦 DESCARGAR TODO EL PROYECTO (ZIP)",
                    data=zip_bytes,
                    file_name=zip_filename,
                    mime="application/zip",
                    use_container_width=True,
                    type="primary"
                )
                st.caption("Incluye: Memoria descriptiva · Mediciones y presupuesto Excel · Plano 2D · Datos catastro · Layout 3D · Guía de próximos pasos")

                # ── Certificación criptográfica del ZIP ──────────────────────
                try:
                    from modules.marketplace.blockchain_cert import certify, cert_badge_html
                    _cert = certify(
                        zip_bytes=zip_bytes,
                        doc_name=zip_filename,
                        user_email=st.session_state.get("client_email", ""),
                        plot_id=str(st.session_state.get("design_plot_data", {}).get("id", ""))
                    )
                    st.markdown(cert_badge_html(_cert), unsafe_allow_html=True)
                except Exception:
                    pass  # falla silenciosamente, nunca interrumpe la descarga

                # ── Gemelo Digital BIM / IFC ──────────────────────────────────
                try:
                    from modules.ai_house_designer.ifc_export import generate_ifc, rooms_to_svg
                    _rooms_src = (
                        st.session_state.get("babylon_modified_layout")
                        or [{"name": k, "area": float(v)}
                            for k, v in req.get("ai_room_proposal", {}).items()
                            if isinstance(v, (int, float))]
                    )
                    if _rooms_src:
                        _pname_ifc = (req.get("nombre_proyecto") or "ArchiRapid_Proyecto").replace(" ", "_")
                        _ifc_bytes = generate_ifc(_rooms_src, project_name=_pname_ifc)
                        st.markdown("---")
                        st.markdown("""
<div style="background:linear-gradient(135deg,rgba(30,58,95,0.6),rgba(13,27,42,0.8));
            border:1px solid rgba(245,158,11,0.35);border-radius:12px;
            padding:16px 20px;margin-bottom:12px;">
  <div style="font-size:15px;font-weight:800;color:#F8FAFC;margin-bottom:4px;">
    🏗️ Gemelo Digital BIM/IFC
  </div>
  <div style="font-size:12px;color:#94A3B8;">
    Formato IFC2x3 · Compatible con FreeCAD · Archicad · Revit · Navisworks · BIMvision
  </div>
  <div style="margin-top:8px;font-size:11px;color:#64748B;">
    ✅ Subvencionable UE (BIM Mandate) · Cada habitación es un espacio IFC certificado
  </div>
</div>""", unsafe_allow_html=True)

                        st.download_button(
                            label="📐 Descargar Gemelo Digital (.ifc)",
                            data=_ifc_bytes,
                            file_name=f"{_pname_ifc}.ifc",
                            mime="application/x-step",
                            use_container_width=True,
                            help="Ábrelo en FreeCAD (gratis) o BIMvision para ver el modelo 3D técnico"
                        )
                        st.caption("IFC2x3 · FreeCAD · Archicad · Revit · Navisworks · BIMvision · UE BIM Mandate")
                except Exception:
                    pass  # nunca interrumpe el flujo principal

            except Exception as e:
                st.error(f"Error generando ZIP: {e}")
                import traceback
                st.code(traceback.format_exc())

            # ── Tablón de Obras — publicar para recibir ofertas de constructores ──
            try:
                st.markdown("---")
                st.markdown("### 🏗️ ¿Quieres recibir ofertas de constructores?")

                _tab_key = f"tablon_published_{req.get('nombre_proyecto','')}"
                if st.session_state.get(_tab_key):
                    st.success("✅ Proyecto publicado en el Tablón de Obras. Los constructores de tu zona ya pueden enviarte ofertas.")
                    st.info("Consulta las ofertas recibidas en tu **Panel de Cliente → Ofertas de Construcción**.")
                else:
                    st.markdown("""
<div style="background:rgba(34,197,94,0.07);border:1px solid rgba(34,197,94,0.25);
            border-radius:10px;padding:14px 18px;margin-bottom:12px;">
  <div style="font-weight:700;color:#F8FAFC;font-size:14px;">🏗️ Red de constructores ArchiRapid</div>
  <div style="color:#94A3B8;font-size:12px;margin-top:4px;">
    Publicamos tu proyecto (anónimo) en el tablón. Los constructores de tu provincia ven el
    desglose de presupuesto y te envían ofertas con precio, plazo y garantía.
    Tú comparas y eliges. Sin compromiso.
  </div>
</div>""", unsafe_allow_html=True)

                    _prov_tab   = (req.get("province") or
                                   st.session_state.get("design_plot_data", {}).get("province") or "")
                    _style_tab  = req.get("style", "")
                    _name_tab   = req.get("nombre_proyecto") or "Mi proyecto ArchiRapid"
                    _cemail_tab = (st.session_state.get("user_email") or
                                   st.session_state.get("client_email") or "")
                    _cname_tab  = (st.session_state.get("user_name") or
                                   st.session_state.get("client_name") or "Cliente")

                    # Selector de partidas a contratar (Fase 2 discriminación por especialidad)
                    from modules.marketplace.service_providers import _ESP_LABELS as _ESP_L
                    _partidas_opts = list(_ESP_L.keys())
                    _partidas_sel = st.multiselect(
                        "¿Qué partidas quieres contratar? (deja vacío = obra completa)",
                        options=_partidas_opts,
                        format_func=lambda x: _ESP_L.get(x, x),
                        key="tablon_partidas_sel",
                        help="Solo recibirás ofertas de profesionales con esas especialidades."
                    )

                    if st.button("🏗️ Publicar en Tablón de Obras — Recibir ofertas",
                                 type="primary", use_container_width=True, key="btn_tablon"):
                        try:
                            from modules.marketplace.service_providers import publish_to_tablon
                            _tid = publish_to_tablon(
                                client_email=_cemail_tab,
                                client_name=_cname_tab,
                                project_name=_name_tab,
                                province=_prov_tab,
                                style=_style_tab,
                                total_area=float(total_area),
                                total_cost=float(total_cost),
                                partidas_list=[(p[0], p[1], p[2], p[3]) for p in partidas],
                                partidas_solicitadas=_partidas_sel if _partidas_sel else None,
                            )
                            st.session_state[_tab_key] = True
                            try:
                                from modules.marketplace.email_notify import _send
                                _parts_str = ", ".join(_partidas_sel) if _partidas_sel else "obra completa"
                                _send(
                                    f"🏗️ <b>Nuevo proyecto en Tablón</b>\n"
                                    f"ID: {_tid}\nCliente: {_cname_tab} ({_cemail_tab})\n"
                                    f"Proyecto: {_name_tab} · {total_area:.0f} m²\n"
                                    f"Provincia: {_prov_tab} · Coste: €{total_cost:,}\n"
                                    f"Partidas: {_parts_str}"
                                )
                            except Exception:
                                pass
                            st.success("✅ Publicado. Los constructores de tu zona ya pueden enviarte ofertas.")
                            st.rerun()
                        except Exception as _etab:
                            st.error(f"Error publicando: {_etab}")
            except Exception:
                pass  # nunca interrumpe el flujo

            st.markdown("---")
            st.info("📬 Proyecto guardado en tu **Panel de Cliente**. El equipo ArchiRapid recibirá notificación inmediata.")


# ══════════════════════════════════════════════════════════════════════════════
# HELPER COMPARTIDO — cálculos financieros y energéticos (pasos 4/5/6)
# ══════════════════════════════════════════════════════════════════════════════

def _get_financials() -> dict:
    """Cálculo determinista de presupuesto, subvenciones y energía.
    Lee de session_state → siempre devuelve los mismos números independientemente
    del paso en el que se llame."""
    req         = st.session_state.get("ai_house_requirements", {})
    design_data = get_current_design_data()
    total_area  = design_data.get("total_area", 0) or 0
    rooms       = design_data.get("rooms", [])
    style       = req.get("style", "Moderno")
    energy      = req.get("energy", {})
    budget      = req.get("budget", 150000)

    # ── Costes reales mercado 2025 (base_costes_construccion_espana.md) ──────
    _BASE_M2 = {
        "Moderno": 1500, "Mediterráneo": 1500, "Contemporáneo": 1550,
        "Ecológico": 1600, "Rural": 1650, "Montaña": 1750, "Clásico": 1800,
    }
    _FLOOR_FACTOR = {
        "1 Planta": 1.00, "2 Plantas": 1.05,
        "Planta Baja + Semisótano": 1.18, "2 Plantas + Semisótano": 1.22,
    }
    floors_str   = req.get("floors", "1 Planta")
    base_m2      = _BASE_M2.get(style, 1500)
    # Override con precio configurado por el arquitecto (si existe)
    _custom_m2 = st.session_state.get("arch_cost_per_m2")
    if _custom_m2 and isinstance(_custom_m2, (int, float)) and _custom_m2 > 0:
        base_m2 = int(_custom_m2)
    floor_factor = _FLOOR_FACTOR.get(floors_str, 1.00)
    PEM          = int(total_area * base_m2 * floor_factor)
    # Tarifas configurables por el arquitecto (con fallback a valores de mercado)
    fee_pct  = float(st.session_state.get("arch_fee_pct",      8.0))
    exp_pct  = float(st.session_state.get("arch_expenses_pct", 5.0))
    iva_pct  = float(st.session_state.get("arch_iva_pct",     10.0))
    honorarios   = int(PEM * fee_pct / 100)
    licencias    = int(PEM * exp_pct / 100)
    # Extras reales (precio fijo de mercado)
    extras_req   = req.get("extras", {})
    extra_pool   = 25000 if extras_req.get("pool") else 0
    extra_solar  = 8000  if energy.get("solar") else 0
    extra_dom    = int(total_area * 80) if energy.get("domotic") else 0
    _STYLE_CHIMNEY = {"Montaña": 4500, "Rural": 4500, "Clásico": 3500}
    chimney_cost = _STYLE_CHIMNEY.get(style, 0)
    total_cost   = int(PEM * (1 + fee_pct/100 + exp_pct/100)) + 1200 + extra_pool + extra_solar + extra_dom + chimney_cost
    iva_amount   = int(total_cost * iva_pct / 100)
    total_con_iva = total_cost + iva_amount
    # Variables de compatibilidad (usadas en retorno y subvenciones)
    construction_cost = PEM
    foundation_cost   = int(PEM * 0.10)
    installation_cost = int(PEM * 0.13)
    architecture_cost = honorarios + licencias

    partidas = [
        ("1. Geotécnico y topografía",   1200,               "0%",  "Estudio del terreno previo a proyecto"),
        ("2. Movimiento de tierras",      int(PEM*0.03),      "3%",  "Excavación, nivelación y transporte"),
        ("3. Cimentación",               int(PEM*0.10),      "10%", "Zapatas / losa según geotécnico"),
        ("4. Estructura y forjados",      int(PEM*0.20),      "20%", "Hormigón armado, pilares, forjado"),
        ("5. Cerramientos y fachada",     int(PEM*0.12),      "12%", f"Fachada estilo {style}"),
        ("6. Cubierta",                  int(PEM*0.07),      "7%",  f"Cubierta {req.get('roof_type','inclinada')}"),
        ("7. Carpintería exterior",       int(PEM*0.06),      "6%",  "Ventanas, puertas acceso, persianas"),
        ("8. Particiones interiores",     int(PEM*0.05),      "5%",  "Tabiquería, trasdosados, pladur"),
        ("9. Instalaciones",             int(PEM*0.13),      "13%", "Eléctrica, fontanería, climatización"),
        ("10. Acabados interiores",       int(PEM*0.12),      "12%", "Suelos, techos, pintura interior"),
        ("11. Baños y cocina",            int(PEM*0.05),      "5%",  "Sanitarios, grifería, equip. básico cocina"),
        ("12. Urbanización parcela",      int(PEM*0.03),      "3%",  "Acceso vehículos, cerramiento, jardinería"),
        ("— Honorarios técnicos",        honorarios,         f"{fee_pct:.0f}%",  "Arquitecto, aparejador, coord. seguridad"),
        ("— Licencias y tasas",          licencias,          f"{exp_pct:.0f}%",  "Licencia de obras e ICIO municipal"),
    ]
    if chimney_cost:
        partidas.append(("★ Chimenea / hogar", chimney_cost, "—", f"Chimenea de leña/biomasa — estilo {style}"))
    if extra_pool:
        partidas.append(("★ Piscina",          extra_pool,   "—", "Piscina hormigón 8×4 m (precio medio)"))
    if extra_solar:
        partidas.append(("★ Inst. fotovoltaica 5 kW", extra_solar, "—", "Paneles + inversor + gestor energía"))
    if extra_dom:
        partidas.append(("★ Domótica",         extra_dom,    "—", "Sistema domótico medio (€80/m²)"))

    # Subvenciones
    subsidy = 0
    if energy.get("solar"):        subsidy += 3000
    if energy.get("aerotermia"):   subsidy += 5000
    if energy.get("geotermia"):    subsidy += 8000
    if energy.get("insulation"):   subsidy += 2000
    if energy.get("rainwater"):    subsidy += 1000
    if energy.get("domotic"):      subsidy += 500
    subsidy_total = min(subsidy, int(total_cost * 0.40))

    # Calificación energética
    energy_score = 0
    if energy.get("solar"):        energy_score += 25
    if energy.get("aerotermia"):   energy_score += 20
    if energy.get("geotermia"):    energy_score += 20
    if energy.get("insulation"):   energy_score += 15
    if energy.get("rainwater"):    energy_score += 10
    if energy.get("domotic"):      energy_score += 10

    if energy_score >= 60:   energy_label, energy_color = "A", "#2ECC71"
    elif energy_score >= 40: energy_label, energy_color = "B", "#27AE60"
    elif energy_score >= 20: energy_label, energy_color = "C", "#F39C12"
    else:                    energy_label, energy_color = "D", "#E74C3C"

    consumo_base  = total_area * 120
    ahorro_pct    = energy_score * 0.9
    consumo_real  = int(consumo_base * (1 - ahorro_pct / 100))
    ahorro_euros  = int((consumo_base - consumo_real) * 0.18)
    co2_evitado   = round((consumo_base - consumo_real) * 0.25 / 1000, 1)

    # Sincronizar design_data con el total calculado (consistencia ZIP/PDF)
    design_data["total_cost"] = total_cost
    design_data["cost_per_m2"] = base_m2

    return {
        "req": req, "design_data": design_data, "total_area": total_area,
        "rooms": rooms, "style": style, "energy": energy, "budget": budget,
        "construction_cost": construction_cost, "foundation_cost": foundation_cost,
        "installation_cost": installation_cost, "architecture_cost": architecture_cost,
        "chimney_cost": chimney_cost, "total_cost": total_cost,
        "iva_pct": iva_pct, "iva_amount": iva_amount, "total_con_iva": total_con_iva,
        "partidas": partidas, "subsidy_total": subsidy_total,
        "energy_score": energy_score, "energy_label": energy_label,
        "energy_color": energy_color, "consumo_real": consumo_real,
        "ahorro_euros": ahorro_euros, "co2_evitado": co2_evitado,
    }


# ══════════════════════════════════════════════════════════════════════════════
# PASO 4 — RESUMEN EJECUTIVO
# ══════════════════════════════════════════════════════════════════════════════

def render_step4_resumen():
    """Paso 4: Resumen limpio — KPIs + visor 3D + mapa satélite + Gemelo Digital (opcional)."""

    f = _get_financials()
    req, total_area, total_cost = f["req"], f["total_area"], f["total_cost"]
    subsidy_total = f["subsidy_total"]
    energy_label, energy_color = f["energy_label"], f["energy_color"]
    style = f["style"]

    if not req.get("ai_room_proposal"):
        st.warning("Primero completa los pasos anteriores.")
        if st.button("← Inicio"):
            st.session_state["ai_house_step"] = 1
            st.rerun()
        return

    st.markdown("## 🏠 Tu Proyecto en Resumen")
    st.caption("Un vistazo rápido a los números clave de tu casa. Cuando estés listo, continúa al Paso 5.")

    # ── Capturas 3D del editor Babylon ───────────────────────────────────────
    if st.session_state.get("babylon_captures"):
        _caps_r4 = st.session_state["babylon_captures"]
        st.success(f"✅ {len(_caps_r4)} vistas 3D vinculadas al proyecto")
        _cap_lbl = {"sur_fachada_principal": "Fachada Sur", "norte": "Norte",
                    "este": "Este", "oeste": "Oeste", "planta_cenital": "Cenital"}
        _ccols = st.columns(min(len(_caps_r4), 5))
        for _ci, (_ck, _cu) in enumerate(_caps_r4.items()):
            with _ccols[_ci % 5]:
                st.image(_cu, caption=_cap_lbl.get(_ck, _ck), use_container_width=True)
    else:
        st.info(
            "💡 Si editaste en el 3D: pulsa **📸 Capturar Vistas** (5 fotos) y **💾 Guardar JSON** "
            "en el editor. Súbelos aquí para incluirlos en tu proyecto."
        )
        _up_r4 = st.file_uploader(
            "ZIP de capturas · imágenes PNG · layout JSON",
            type=["zip", "png", "json"], accept_multiple_files=True,
            key="captures_s4_uploader", label_visibility="collapsed"
        )
        if _up_r4:
            import zipfile as _zf4, io as _io4, base64 as _b64r4, json as _json4
            _vlab = ["sur_fachada_principal", "norte", "este", "oeste", "planta_cenital"]
            _caps_new = {}
            for _uf in _up_r4:
                _fn = _uf.name.lower()
                if _fn.endswith(".zip"):
                    with _zf4.ZipFile(_io4.BytesIO(_uf.read())) as _z4:
                        for _i4, _pn in enumerate(sorted(_n for _n in _z4.namelist()
                                                          if _n.lower().endswith(".png"))):
                            _k4 = _vlab[_i4] if _i4 < len(_vlab) else f"vista_{_i4+1}"
                            _caps_new[_k4] = ("data:image/png;base64,"
                                              + _b64r4.b64encode(_z4.read(_pn)).decode())
                elif _fn.endswith(".png"):
                    _i4 = len(_caps_new)
                    _k4 = _vlab[_i4] if _i4 < len(_vlab) else f"vista_{_i4+1}"
                    _caps_new[_k4] = "data:image/png;base64," + _b64r4.b64encode(_uf.read()).decode()
                elif _fn.endswith(".json"):
                    try:
                        _lay = _json4.loads(_uf.read().decode("utf-8"))
                        if isinstance(_lay, list):
                            st.session_state["babylon_modified_layout"] = _lay
                    except Exception:
                        pass
            if _caps_new:
                st.session_state["babylon_captures"] = _caps_new
                st.rerun()
    st.markdown("---")

    # ── KPIs ────────────────────────────────────────────────────────────────
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Coste estimado",    f"€{total_cost:,}",
              delta=f"vs €{int(total_area*3500):,} piso ciudad")
    k2.metric("Subvenciones",      f"€{subsidy_total:,}",
              delta=f"Neto: €{total_cost-subsidy_total:,}")
    k3.metric("Superficie",        f"{total_area:.0f} m²",
              delta=f"Estilo {style.split()[-1]}")
    k4.metric("Ahorro vs piso",    f"€{int(total_area*3500)-(total_cost-subsidy_total):,}",
              delta="con subvenciones")

    # Badge energético
    st.markdown(f"""
<div style="display:inline-block;background:{energy_color};color:#fff;
            padding:8px 22px;border-radius:8px;font-size:20px;font-weight:900;
            margin:12px 0;">
  Calificación Energética: <b>{energy_label}</b>
  &nbsp;·&nbsp; <span style="font-size:14px;font-weight:400;">
  Ahorro ~€{f['ahorro_euros']:,}/año · CO₂ evitado {f['co2_evitado']} t/año
  </span>
</div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ── Visor 3D (viewer3d.py — Three.js) + Plano BIM ───────────────────────
    try:
        from .viewer3d import Viewer3D
        from .data_model import HouseDesign, Plot, RoomType, RoomInstance
        from .ifc_export import rooms_to_svg as _r2svg_s4
        import streamlit.components.v1 as _cmp_v3d

        _final_d_v3 = get_final_design()
        _rooms_v3   = _final_d_v3.get("rooms", [])
        if _rooms_v3:
            st.markdown("### 🏗️ Tu Casa en 3D — Vista Previa BIM")
            st.caption("Arrastra para rotar · Rueda para zoom · Plano de distribución a la derecha")
            _col_3d, _col_svg = st.columns([3, 2])

            with _col_3d:
                _plot_v3   = Plot(id="preview", area_m2=500, buildable_ratio=0.33)
                _design_v3 = HouseDesign(_plot_v3)
                _roof_v3   = req.get("roof_type", "Dos aguas (clásico, eficiente)")
                _shape_v3  = req.get("house_shape", "Rectangular (más común)")
                setattr(_design_v3, "request", {"house_shape": _shape_v3, "roof_type": _roof_v3})
                for _r3 in _rooms_v3:
                    _rt3 = RoomType(
                        code=_r3.get("code", _r3.get("name", "espacio")),
                        name=_r3.get("name", "Espacio"),
                        min_m2=2, max_m2=200, base_cost_per_m2=1000
                    )
                    _design_v3.rooms.append(
                        RoomInstance(room_type=_rt3, area_m2=float(_r3.get("area_m2", 10)))
                    )
                _viewer_v3 = Viewer3D(_design_v3, roof_type=_roof_v3)
                _html_v3d  = _viewer_v3.generate_html()
                _cmp_v3d.html(_html_v3d, height=520, scrolling=False)

            with _col_svg:
                # Preferir babylon_modified_layout (tiene x,z,width,depth → plano real)
                _bim_src = (
                    st.session_state.get("babylon_modified_layout")
                    or [{"name": k, "area": float(v)}
                        for k, v in req.get("ai_room_proposal", {}).items()
                        if isinstance(v, (int, float))]
                )
                _svg_bim = _r2svg_s4(_bim_src, px=360)
                if _svg_bim:
                    st.markdown("**📐 Plano de distribución BIM**")
                    st.markdown(_svg_bim, unsafe_allow_html=True)
                    st.caption("Cada espacio = objeto IFC2x3 · UE BIM Mandate")

            st.markdown("---")
    except Exception:
        pass  # nunca interrumpe el flujo principal

    # ── Mapa satélite ───────────────────────────────────────────────────────
    try:
        import folium as _fol, math as _m, streamlit.components.v1 as _cv
        _pd = st.session_state.get("design_plot_data", {})
        _lat, _lon = _pd.get("lat"), _pd.get("lon")
        if _lat and _lon:
            _ta = total_area or 80
            _tw = _m.sqrt(_ta * 1.3)
            _td = _m.sqrt(_ta / 1.3)
            _dlat = _td / 111000
            _dlon = _tw / (111000 * _m.cos(_m.radians(float(_lat))))
            _mm = _fol.Map(location=[float(_lat), float(_lon)], zoom_start=18,
                           tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
                           attr="Esri")
            _fol.Rectangle(
                bounds=[[float(_lat)-_dlat/2, float(_lon)-_dlon/2],
                        [float(_lat)+_dlat/2, float(_lon)+_dlon/2]],
                color="#0055FF", fill=True, fill_color="#0055FF",
                fill_opacity=0.35, weight=2,
                tooltip=f"Tu casa: {_tw:.1f}m × {_td:.1f}m ({_ta:.0f} m²)"
            ).add_to(_mm)
            _fol.Marker([float(_lat), float(_lon)],
                        icon=_fol.Icon(color="red", icon="home", prefix="fa"),
                        tooltip=_pd.get("title", "Tu parcela")).add_to(_mm)
            st.markdown("### 📍 Ubicación del proyecto")
            _cv.html(_mm._repr_html_(), height=380)
            st.caption(f"Parcela: {_pd.get('title', '')} · Huella aproximada {_tw:.1f}m × {_td:.1f}m")
            st.markdown("---")
    except Exception:
        pass

    # ── Gemelo Digital (expander opcional) ─────────────────────────────────
    with st.expander("🏗️ Ver Gemelo Digital — Análisis BIM / Energía / Carbono", expanded=False):
        try:
            import plotly.graph_objects as go
            tab_bim, tab_energia, tab_carbono, tab_whatif = st.tabs(
                ["📐 BIM / Espacios", "⚡ Energía", "🌿 Carbono", "🔮 What-If"])

            with tab_bim:
                st.markdown("#### Distribución de espacios IFC")
                _rooms = f["design_data"].get("rooms", [])
                if _rooms:
                    _names  = [r.get("name", "?") for r in _rooms]
                    _areas  = [r.get("area", r.get("area_m2", 0)) for r in _rooms]
                    _fig_bim = go.Figure(go.Pie(labels=_names, values=_areas, hole=0.35,
                                                marker_colors=["#1E3A5F","#2563EB","#3B82F6",
                                                               "#60A5FA","#93C5FD","#BFDBFE"]))
                    _fig_bim.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                           font_color="#F8FAFC", height=320, margin=dict(t=20,b=20))
                    st.plotly_chart(_fig_bim, use_container_width=True)
                    st.caption("Cada segmento = objeto IFC2x3 certificable · Compatible BIM Mandate UE")
                else:
                    st.info("Completa el diseño para ver la distribución de espacios.")

            with tab_energia:
                st.markdown("#### Consumo energético anual estimado")
                _e = f["energy"]
                _consumo_base = total_area * 120
                _sistemas = []
                _ahorros   = []
                if _e.get("solar"):        _sistemas.append("Solar PV"); _ahorros.append(_consumo_base*0.25)
                if _e.get("aerotermia"):   _sistemas.append("Aerotermia"); _ahorros.append(_consumo_base*0.20)
                if _e.get("geotermia"):    _sistemas.append("Geotermia"); _ahorros.append(_consumo_base*0.20)
                if _e.get("insulation"):   _sistemas.append("Aislamiento"); _ahorros.append(_consumo_base*0.15)
                if _e.get("rainwater"):    _sistemas.append("Agua lluvia"); _ahorros.append(_consumo_base*0.05)
                if _e.get("domotic"):      _sistemas.append("Domótica"); _ahorros.append(_consumo_base*0.05)
                if _sistemas:
                    _fig_e = go.Figure(go.Bar(x=_sistemas, y=[int(a) for a in _ahorros],
                                              marker_color="#2563EB", text=[f"{int(a):,} kWh" for a in _ahorros],
                                              textposition="outside"))
                    _fig_e.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                         font_color="#F8FAFC", height=300, yaxis_title="kWh/año ahorrado",
                                         margin=dict(t=20,b=40))
                    st.plotly_chart(_fig_e, use_container_width=True)
                else:
                    st.info("Activa sistemas energéticos en el Paso 2 para ver el análisis.")
                st.metric("Consumo estimado", f"{f['consumo_real']:,} kWh/año",
                          delta=f"Ahorro: €{f['ahorro_euros']:,}/año")

            with tab_carbono:
                st.markdown("#### Huella de carbono del proyecto")
                _carbon_construccion = int(total_area * 350)  # kg CO2 construcción típica
                _carbon_operacion    = int(f["consumo_real"] * 0.25)  # kg CO2/año operación
                _fig_c = go.Figure(go.Bar(
                    x=["Construcción (total)", "Operación (año 1)", "Operación (año 10)", "Operación (año 25)"],
                    y=[_carbon_construccion, _carbon_operacion,
                       _carbon_operacion*10, _carbon_operacion*25],
                    marker_color=["#DC2626","#F59E0B","#10B981","#059669"],
                    text=[f"{v:,} kg" for v in [_carbon_construccion, _carbon_operacion,
                                                  _carbon_operacion*10, _carbon_operacion*25]],
                    textposition="outside"))
                _fig_c.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                     font_color="#F8FAFC", height=300, yaxis_title="kg CO₂",
                                     margin=dict(t=20,b=40))
                st.plotly_chart(_fig_c, use_container_width=True)
                st.caption(f"CO₂ evitado vs casa convencional: **{f['co2_evitado']} t/año** gracias a tus sistemas energéticos")

            with tab_whatif:
                st.markdown("#### ¿Qué pasa si añado más superficie?")
                _extra = st.slider("m² adicionales", 0, 100, 20, 5, key="whatif_extra_m2_s4")
                _new_area  = total_area + _extra
                _new_cost  = int(_new_area * 1100) + int(_new_area * 180) + int(_new_area * 150) + int(_new_area * 1280 * 0.08)
                _new_kwh   = int(_new_area * 120 * (1 - f["energy_score"] * 0.009))
                _c1, _c2, _c3 = st.columns(3)
                _c1.metric("Nueva superficie", f"{_new_area:.0f} m²", delta=f"+{_extra} m²")
                _c2.metric("Coste estimado",   f"€{_new_cost:,}",      delta=f"+€{_new_cost-total_cost:,}")
                _c3.metric("Consumo",          f"{_new_kwh:,} kWh/año")
        except ImportError:
            st.info("Instala plotly para ver los gráficos del Gemelo Digital: `pip install plotly`")
        except Exception as _eg:
            st.warning(f"Gemelo Digital no disponible: {_eg}")

    # ── Botones de navegación ────────────────────────────────────────────────
    st.markdown("---")
    _c_back, _c_next = st.columns(2)
    with _c_back:
        if st.button("← Volver al Editor 3D", use_container_width=True, key="s4_back"):
            st.session_state["ai_house_step"] = 3
            st.rerun()
    with _c_next:
        if st.button("Continuar → Documentación y Servicios", type="primary",
                     use_container_width=True, key="s4_next"):
            st.session_state["ai_house_step"] = 5
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PASO 5 — DOCUMENTACIÓN + SERVICIOS
# ══════════════════════════════════════════════════════════════════════════════

def render_step5_docs():
    """Paso 5: Partidas + plano de planta + constructores + selección de servicios."""

    f = _get_financials()
    req, total_area, total_cost = f["req"], f["total_area"], f["total_cost"]
    partidas, subsidy_total = f["partidas"], f["subsidy_total"]
    energy_label, style = f["energy_label"], f["style"]

    if not req.get("ai_room_proposal"):
        st.warning("Primero completa los pasos anteriores.")
        if st.button("← Inicio"):
            st.session_state["ai_house_step"] = 1
            st.rerun()
        return

    st.markdown("## 📋 Documentación Técnica y Servicios")
    st.caption("Revisa el presupuesto detallado, el plano de planta y selecciona los servicios que necesitas.")

    # ── Presupuesto desglosado ───────────────────────────────────────────────
    st.markdown("### 💰 Presupuesto por Partidas")
    for partida_name, coste, pct, desc in partidas:
        c1, c2, c3 = st.columns([3, 1, 1])
        c1.markdown(f"**{partida_name}** · *{desc}*")
        c2.markdown(f"€{coste:,}")
        c3.markdown(pct)
    st.markdown("---")
    _tc1, _tc2, _tc3 = st.columns(3)
    _tc1.metric("Total ejecución material", f"€{total_cost:,}")
    _tc2.metric("Subvenciones estimadas",   f"-€{subsidy_total:,}")
    _tc3.metric("Coste neto",               f"€{total_cost-subsidy_total:,}")
    st.markdown("---")

    # ── Plano de Planta Profesional (FloorPlanSVG) ───────────────────────────
    st.markdown("### 🗺️ Plano de Planta")
    _final_design_p5 = get_final_design()
    _plan_source_p5  = _final_design_p5.get("source", "ai_original")
    _plan_label_p5   = {
        "babylon":       "🏗️ Plano desde tu diseño en Editor 3D",
        "step2_sliders": "📊 Plano con tus ajustes del Paso 2",
        "ai_original":   "🤖 Plano propuesta IA",
    }.get(_plan_source_p5, "Plano de distribución")

    if st.button("🗺️ Generar Plano de Planta", type="primary",
                 use_container_width=True, key="gen_plan_s5"):
        try:
            from .architect_layout import generate_layout as _gen_lay
            from .floor_plan_svg import FloorPlanSVG as _FPSVG
            from .data_model import HouseDesign as _HD, Plot as _Plt, RoomType as _RT, RoomInstance as _RI
            _rooms_p5 = [{"code": r.get("code", r.get("name", "espacio")),
                          "name": r.get("name", "Espacio"),
                          "area_m2": float(r.get("area_m2", 10))}
                         for r in _final_design_p5["rooms"]]
            _gen_lay(_rooms_p5, req.get("house_shape", "Rectangular"))
            _plot_p5 = _Plt(id="s5", area_m2=500, buildable_ratio=0.33)
            _des_p5  = _HD(_plot_p5)
            for _rp in _final_design_p5["rooms"]:
                _des_p5.rooms.append(_RI(
                    room_type=_RT(code=_rp.get("code", "espacio"), name=_rp.get("name", "Espacio"),
                                  min_m2=2, max_m2=200, base_cost_per_m2=1000),
                    area_m2=float(_rp.get("area_m2", 10))
                ))
            _planner_p5 = _FPSVG(_des_p5)
            st.session_state["final_floor_plan"] = _planner_p5.generate()
            st.success(f"✅ {_plan_label_p5}")
        except Exception as _ep5:
            st.error(f"Error generando plano: {_ep5}")

    if st.session_state.get("final_floor_plan"):
        st.image(st.session_state["final_floor_plan"],
                 caption=_plan_label_p5, use_container_width=True)
    elif st.session_state.get("current_floor_plan"):
        st.image(st.session_state["current_floor_plan"],
                 caption="Plano del Paso 2 (pulsa 'Generar' para actualizarlo con tu diseño final)",
                 use_container_width=True)
    else:
        st.info("Pulsa 'Generar Plano de Planta' para ver la distribución de tu casa")

    st.markdown("---")

    # ── Capturas 3D del Editor Babylon ───────────────────────────────────────
    st.markdown("### 📸 Vistas 3D de tu Diseño")
    if st.session_state.get("babylon_captures"):
        _caps5 = st.session_state["babylon_captures"]
        st.success(f"✅ {len(_caps5)} vistas 3D vinculadas — se incluirán en el ZIP del proyecto")
        _cap_labels5 = {
            "sur_fachada_principal": "Fachada Sur",
            "norte":                 "Vista Norte",
            "este":                  "Vista Este",
            "oeste":                 "Vista Oeste",
            "planta_cenital":        "Planta Cenital",
        }
        _ccols = st.columns(min(len(_caps5), 5))
        for _ci, (_ck, _cdurl) in enumerate(_caps5.items()):
            with _ccols[_ci % 5]:
                st.image(_cdurl, caption=_cap_labels5.get(_ck, _ck), use_container_width=True)
    else:
        st.info("💡 Si capturaste vistas en el Editor 3D, súbelas aquí para incluirlas en la documentación. "
                "En el editor: pulsa 📸 **Capturar Vistas** → se descargan automáticamente como ZIP.")
        _uploaded_caps5 = st.file_uploader(
            "Sube el ZIP de capturas (Vistas_3D_ArchiRapid.zip) o imágenes PNG sueltas",
            type=["zip", "png"],
            accept_multiple_files=True,
            key="captures_doc_uploader_s5"
        )
        if _uploaded_caps5:
            import zipfile as _zf5, io as _io5c, base64 as _b64c
            _vl5  = ["sur_fachada_principal", "norte", "este", "oeste", "planta_cenital"]
            _caps_new = {}
            for _uf in _uploaded_caps5:
                if _uf.name.lower().endswith(".zip"):
                    with _zf5.ZipFile(_io5c.BytesIO(_uf.read())) as _z5:
                        _pngs = sorted(n for n in _z5.namelist() if n.lower().endswith(".png"))
                        for _pi, _pn in enumerate(_pngs):
                            _k5 = _vl5[_pi] if _pi < len(_vl5) else f"vista_{_pi+1}"
                            _caps_new[_k5] = "data:image/png;base64," + _b64c.b64encode(_z5.read(_pn)).decode()
                elif _uf.name.lower().endswith(".png"):
                    _ki = len(_caps_new)
                    _k5 = _vl5[_ki] if _ki < len(_vl5) else f"vista_{_ki+1}"
                    _caps_new[_k5] = "data:image/png;base64," + _b64c.b64encode(_uf.read()).decode()
            if _caps_new:
                st.session_state["babylon_captures"] = _caps_new
                st.rerun()

    st.markdown("---")

    # ── Excel descargable con partidas presupuestarias ────────────────────────
    try:
        import io as _io5, openpyxl as _ox5
        from openpyxl.styles import Font as _Fnt5, PatternFill as _PF5, Alignment as _Al5
        _wb5 = _ox5.Workbook()
        _ws5 = _wb5.active
        _ws5.title = "Presupuesto"
        _hf5  = _Fnt5(bold=True, color="FFFFFF")
        _hfil = _PF5("solid", fgColor="1E3A5F")
        _tfil = _PF5("solid", fgColor="2C3E50")
        for _ci5, _ch in enumerate(["Partida", "Coste (€)", "% Total", "Descripción"], 1):
            _c5 = _ws5.cell(row=1, column=_ci5, value=_ch)
            _c5.font = _hf5; _c5.fill = _hfil
        for _ri5, (_pn, _pc, _pp, _pd) in enumerate(partidas, 2):
            _ws5.cell(row=_ri5, column=1, value=_pn)
            _ws5.cell(row=_ri5, column=2, value=_pc)
            _ws5.cell(row=_ri5, column=3, value=_pp)
            _ws5.cell(row=_ri5, column=4, value=_pd)
        _rf5 = len(partidas) + 2
        _ws5.cell(row=_rf5, column=1, value="TOTAL EJECUCIÓN").font = _Fnt5(bold=True, color="FFFFFF")
        _ws5.cell(row=_rf5, column=1).fill = _tfil
        _ws5.cell(row=_rf5, column=2, value=total_cost).font = _Fnt5(bold=True, color="FFFFFF")
        _ws5.cell(row=_rf5, column=2).fill = _tfil
        _ws5.cell(row=_rf5+1, column=1, value="Subvenciones estimadas")
        _ws5.cell(row=_rf5+1, column=2, value=-subsidy_total)
        _ws5.cell(row=_rf5+2, column=1, value="COSTE NETO").font = _Fnt5(bold=True, color="FFFFFF")
        _ws5.cell(row=_rf5+2, column=1).fill = _tfil
        _ws5.cell(row=_rf5+2, column=2, value=total_cost-subsidy_total).font = _Fnt5(bold=True, color="FFFFFF")
        _ws5.cell(row=_rf5+2, column=2).fill = _tfil
        _ws5.column_dimensions["A"].width = 35
        _ws5.column_dimensions["B"].width = 16
        _ws5.column_dimensions["D"].width = 38
        _xls_buf = _io5.BytesIO()
        _wb5.save(_xls_buf)
        st.download_button(
            label="📊 Descargar Presupuesto por Partidas (Excel)",
            data=_xls_buf.getvalue(),
            file_name="Presupuesto_ArchiRapid.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        pass

    # ── Constructores ─────────────────────────────────────────────────────────
    try:
        st.markdown("### 🏗️ Constructores Recomendados")
        _province = (req.get("province") or
                     st.session_state.get("design_plot_data", {}).get("province") or "")
        from modules.marketplace.service_providers import get_builders_for_province as _gbp
        _builders = _gbp(province=_province, style=style, max_results=3)
        if _builders:
            for _b in _builders:
                with st.container():
                    _bc1, _bc2 = st.columns([3, 1])
                    with _bc1:
                        st.markdown(f"**{_b.get('name', 'Constructor')}** — {_b.get('province', _province)}")
                        st.caption(f"{_b.get('specialty', '')} · {_b.get('rating', '⭐⭐⭐⭐')} · {_b.get('projects_done', '')} proyectos")
                    with _bc2:
                        st.button("Contactar", key=f"contact_{_b.get('id', _b.get('name', ''))}_s5",
                                  use_container_width=True)
        else:
            st.info(f"Constructores para {_province or 'tu zona'} disponibles tras registro.")
    except Exception:
        pass

    # ── Documentación base PDF / CAD ─────────────────────────────────────────
    st.markdown("### 📄 Documentación del Proyecto")
    _dcol1, _dcol2 = st.columns(2)
    with _dcol1:
        st.markdown("""
**📦 Paquete PDF**
- Memoria descriptiva
- Plano 2D profesional
- Presupuesto por partidas
- Calificación energética
- Datos cimientos y extras
""")
        _pdf_copias = st.number_input("Copias PDF — €1.800/copia",
                                      min_value=0, max_value=10, value=0, key="pdf_copias_s5")
    with _dcol2:
        st.markdown("""
**📐 Paquete CAD/BIM Profesional**
- Todo lo del PDF
- Archivos editables CAD
- Modelo 3D exportable (.glb)
- Apto para constructor y arquitecto
""")
        _cad_copias = st.number_input("Copias CAD — €2.500/copia",
                                      min_value=0, max_value=10, value=0, key="cad_copias_s5")

    _coste_doc = (_pdf_copias * 1800) + (_cad_copias * 2500)
    if _coste_doc > 0:
        st.info(f"💰 Documentación seleccionada: PDF ×{_pdf_copias} + CAD ×{_cad_copias} = **€{_coste_doc:,}**")

    st.markdown("---")

    # ── Servicios técnicos profesionales ─────────────────────────────────────
    st.markdown("### 🔧 Servicios Técnicos Profesionales")
    st.caption("Todos gestionados por ArchiRapid con profesionales colegiados")

    _hipoteca_pct   = int(total_cost * 0.005)   # 0,5% del coste de obra
    _hipoteca_total = _hipoteca_pct + 200        # + €200 contratación

    _SERVICIOS_S5 = {
        "visado":      {"label": "📋 Visado del Proyecto (Colegio Arquitectos)",  "precio": int(total_cost * 0.015), "desc": "Obligatorio para licencia de obra. ~1,5% presupuesto"},
        "revision":    {"label": "🔍 Revisión y Validación Técnica",              "precio": 450,                     "desc": "Arquitecto revisa y firma el diseño ArchiRapid"},
        "licencia":    {"label": "🏛️ Gestión Licencia Municipal",                 "precio": 800,                     "desc": "Tramitación completa ante el Ayuntamiento"},
        "constructor": {"label": "👷 Búsqueda y Contrato Constructor",            "precio": int(total_cost * 0.005), "desc": "0,5% del presupuesto. Selección, negociación y contrato"},
        "fontaneria":  {"label": "🚿 Proyecto Fontanería e Instalaciones",        "precio": 650,                     "desc": "Planos de agua, saneamiento y climatización"},
        "electricidad":{"label": "⚡ Proyecto Eléctrico BT",                      "precio": 550,                     "desc": "Esquema unifilar, cuadro eléctrico, domótica"},
        "energia":     {"label": "☀️ Auditoría Energética + CEE",                 "precio": 380,                     "desc": "Certificado Eficiencia Energética oficial"},
        "geotecnico":  {"label": "🔬 Estudio Geotécnico del Terreno",             "precio": 1200,                    "desc": "Obligatorio para cálculo de cimientos"},
        "topografia":  {"label": "📏 Levantamiento Topográfico",                  "precio": 900,                     "desc": "Medición precisa de la parcela"},
        "seguridad":   {"label": "🦺 Estudio Seguridad y Salud",                  "precio": 350,                     "desc": "Obligatorio para obras con proyecto"},
        "vr":          {"label": "🥽 Tour Virtual 360° / Realidad Virtual",       "precio": 290,                     "desc": "Recorrido inmersivo de tu vivienda antes de construir"},
        "hipoteca":    {"label": "🏦 Tramitación de Hipoteca",                    "precio": _hipoteca_total,         "desc": f"0,5% coste obra (€{_hipoteca_pct:,}) + €200 gestión contratación"},
    }

    _seleccionados_s5 = {}
    _coste_servicios  = 0
    _sv_cols = st.columns(2)
    for _i5, (_key5, _svc5) in enumerate(_SERVICIOS_S5.items()):
        with _sv_cols[_i5 % 2]:
            _chk = st.checkbox(
                f"{_svc5['label']} — **€{_svc5['precio']:,}**",
                key=f"svc_{_key5}_s5", help=_svc5["desc"]
            )
            if _chk:
                _seleccionados_s5[_key5] = _svc5["precio"]
                _coste_servicios += _svc5["precio"]

    st.session_state["selected_services_s5"] = list(_seleccionados_s5.keys())
    st.session_state["coste_doc_s5"]         = _coste_doc
    st.session_state["coste_servicios_s5"]   = _coste_servicios
    # Guardar detalles completos (label + precio) para paso 6
    _doc_detail = []
    if _pdf_copias > 0:
        _doc_detail.append({"label": f"📄 Documentación PDF ×{_pdf_copias} copia(s)", "precio": _pdf_copias * 1800})
    if _cad_copias > 0:
        _doc_detail.append({"label": f"📐 Documentación CAD ×{_cad_copias} copia(s)", "precio": _cad_copias * 2500})
    _svc_detail = [{"label": _SERVICIOS_S5[k]["label"], "precio": _SERVICIOS_S5[k]["precio"]}
                   for k in _seleccionados_s5]
    st.session_state["doc_detail_s5"] = _doc_detail
    st.session_state["svc_detail_s5"] = _svc_detail

    _total_s5 = _coste_doc + _coste_servicios
    st.markdown("---")
    st.markdown(f"""
<div style='background:linear-gradient(135deg,#1a1a2e,#2C3E50);
            padding:20px;border-radius:12px;color:white;text-align:center;
            border:2px solid #27AE60;'>
  <h3 style='margin:0;color:#27AE60;'>TOTAL SERVICIOS SELECCIONADOS</h3>
  <h1 style='margin:10px 0;color:white;'>€{_total_s5:,}</h1>
  <p style='margin:0;font-size:13px;color:#aaa;'>
    Documentación: €{_coste_doc:,} | Servicios técnicos: €{_coste_servicios:,}
  </p>
</div>""", unsafe_allow_html=True)

    # ── Botones de navegación ─────────────────────────────────────────────────
    st.markdown("---")
    _c_back5, _c_next5 = st.columns(2)
    with _c_back5:
        if st.button("← Volver al Resumen", use_container_width=True, key="s5_back"):
            st.session_state["ai_house_step"] = 4
            st.rerun()
    with _c_next5:
        if st.button("Continuar → Pago y Descarga", type="primary",
                     use_container_width=True, key="s5_next"):
            st.session_state["ai_house_step"] = 6
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MODO ESTUDIO — DESCARGA PARA ARQUITECTOS
# ══════════════════════════════════════════════════════════════════════════════

def _show_estudio_zip_button(f):
    """Genera y muestra el botón de descarga ZIP para Modo Estudio."""
    try:
        req        = f["req"]
        plot_data  = st.session_state.get("design_plot_data", {})
        zip_bytes, zip_filename = generar_zip_proyecto(
            req=req,
            design_data=f["design_data"],
            plot_data=plot_data,
            partidas=[(p[0], p[1], p[2], p[3]) for p in f["partidas"]],
            subsidy_total=f["subsidy_total"],
            energy_label=f["energy_label"],
        )
        st.download_button(
            label="📦 DESCARGAR PROYECTO COMPLETO (ZIP)",
            data=zip_bytes,
            file_name=zip_filename,
            mime="application/zip",
            type="primary",
            use_container_width=True,
            key="btn_dl_estudio",
        )
        st.caption("Incluye: Memoria PDF · Mediciones Excel · Plano 2D · BIM/IFC")
    except Exception as _ez:
        st.error(f"Error generando ZIP: {_ez}")


def _render_estudio_download():
    """Descarga del proyecto en Modo Estudio (gratis para PRO, €19 para Free)."""
    from modules.marketplace.architects import check_subscription as _check_sub
    arch_id = st.session_state.get("arch_id", "")
    sub     = _check_sub(arch_id) if arch_id else {"active": False}
    is_pro  = sub.get("active") and sub.get("plan") in ("PRO", "PRO_ANUAL", "ENTERPRISE")

    f   = _get_financials()
    req = f["req"]

    if not req.get("ai_room_proposal"):
        st.warning("Completa primero los pasos anteriores.")
        if st.button("← Inicio", key="est_back_s6"):
            st.session_state["ai_house_step"] = 1
            st.rerun()
        return

    arch_name = st.session_state.get("estudio_arch_name", "Arquitecto")
    st.markdown("## 📦 Descarga del Proyecto (Modo Estudio)")
    st.info(f"Proyecto diseñado por **{arch_name}**")

    st.markdown("""
<div style="background:rgba(37,99,235,0.12);border:1px solid rgba(37,99,235,0.35);
            border-radius:12px;padding:16px 20px;margin-bottom:16px;">
  <div style="font-weight:800;color:#F8FAFC;font-size:15px;margin-bottom:8px;">
    📦 Contenido del ZIP del Proyecto
  </div>
  <ul style="color:#CBD5E1;font-size:13px;margin:0;padding-left:20px;line-height:1.9;">
    <li>📄 Memoria descriptiva (PDF)</li>
    <li>📊 Mediciones y presupuesto (Excel)</li>
    <li>🗺️ Plano de planta (PNG)</li>
    <li>🏗️ Archivo BIM/IFC (IFC2x3)</li>
    <li>📋 Informe eficiencia energética</li>
  </ul>
</div>""", unsafe_allow_html=True)

    if is_pro:
        st.success("✅ Plan PRO activo — Descarga gratuita incluida")
        _show_estudio_zip_button(f)
        return

    # ── Plan Free: cobrar €19 ──────────────────────────────────────────────
    st.info("💡 **Plan Free**: Descarga por **€19** (pago único por proyecto)  |  Actualiza a PRO para descargas ilimitadas.")

    _stripe_ok = bool(os.getenv("STRIPE_SECRET_KEY", ""))
    try:
        if not _stripe_ok:
            _stripe_ok = bool(st.secrets.get("STRIPE_SECRET_KEY", ""))
    except Exception:
        pass

    if _stripe_ok and not st.session_state.get("stripe_session_id_estudio"):
        try:
            from modules.stripe_utils import create_custom_session as _ccs_est
            _url_est, _sid_est = _ccs_est(
                line_items=[{"name": "Descarga Proyecto Modo Estudio", "amount_cents": 1900, "quantity": 1}],
                client_email=st.session_state.get("arch_email", ""),
                success_url="https://archirapid.streamlit.app/?estudio_pago=ok",
                cancel_url="https://archirapid.streamlit.app/?estudio_pago=cancel",
                metadata={"arch_id": arch_id, "mode": "estudio"},
            )
            st.session_state["stripe_session_id_estudio"] = _sid_est
            st.session_state["stripe_checkout_url_estudio"] = _url_est
        except Exception:
            pass

    _url_s = st.session_state.get("stripe_checkout_url_estudio")
    _sid_s = st.session_state.get("stripe_session_id_estudio")

    _ec1, _ec2 = st.columns(2)
    with _ec1:
        if _url_s:
            import streamlit.components.v1 as _cv1_est
            _cv1_est.html(
                f"""<a href="{_url_s}" target="_blank"
                    style="display:block;text-align:center;background:#1E3A5F;color:#fff;
                           padding:12px 20px;border-radius:8px;font-weight:700;font-size:15px;
                           text-decoration:none;">
                    💳 Pagar €19 — Descargar Proyecto
                </a>
                <p style="color:#94A3B8;font-size:12px;text-align:center;margin-top:6px;">
                    Pago seguro con Stripe · Test: 4242 4242 4242 4242
                </p>""",
                height=80,
            )
        else:
            st.warning("Stripe no configurado. Contacta con soporte.")
    with _ec2:
        if st.button("🔑 Activar Plan PRO", key="btn_upgrade_estudio"):
            st.info("Ve a la pestaña 💎 Planes para activar Plan PRO con descargas ilimitadas.")

    if _sid_s:
        st.markdown("<br>", unsafe_allow_html=True)
        _ev1, _ev2, _ev3 = st.columns([1, 2, 1])
        with _ev2:
            if st.button("✅ Ya pagué — Verificar y Descargar", use_container_width=True, key="btn_verify_estudio"):
                try:
                    from modules.stripe_utils import verify_session as _vs_est
                    _sess_est = _vs_est(_sid_s)
                    if _sess_est.payment_status == "paid":
                        st.session_state["estudio_pago_completado"] = True
                        st.session_state.pop("stripe_session_id_estudio", None)
                        st.session_state.pop("stripe_checkout_url_estudio", None)
                        st.rerun()
                    else:
                        st.warning("⏳ Pago no confirmado. Completa el pago en Stripe y vuelve a verificar.")
                except Exception as _ve_est:
                    st.error(f"Error al verificar: {_ve_est}")

    if st.session_state.get("estudio_pago_completado"):
        _show_estudio_zip_button(f)


# ══════════════════════════════════════════════════════════════════════════════
# PASO 6 — PAGO + DESCARGA
# ══════════════════════════════════════════════════════════════════════════════

def render_step6_pago():
    """Paso 6: Pago, descarga ZIP, IFC/BIM y publicación en Tablón de Obras."""

    if st.session_state.get("estudio_mode"):
        _render_estudio_download()
        return

    f = _get_financials()
    req, total_area, total_cost = f["req"], f["total_area"], f["total_cost"]
    partidas, subsidy_total = f["partidas"], f["subsidy_total"]
    energy_label, style = f["energy_label"], f["style"]

    if not req.get("ai_room_proposal"):
        st.warning("Primero completa los pasos anteriores.")
        if st.button("← Inicio"):
            st.session_state["ai_house_step"] = 1
            st.rerun()
        return

    st.markdown("## 🛒 Pago y Descarga")

    # ── Resumen detallado ANTES del pago ────────────────────────────────────
    _doc_detail_6 = st.session_state.get("doc_detail_s5", [])
    _svc_detail_6 = st.session_state.get("svc_detail_s5", [])
    _coste_doc_6  = st.session_state.get("coste_doc_s5", 0)
    _coste_svc_6  = st.session_state.get("coste_servicios_s5", 0)
    _subtotal_6   = _coste_doc_6 + _coste_svc_6
    _iva_6        = int(_subtotal_6 * 0.21)
    _total_iva_6  = _subtotal_6 + _iva_6

    # Tabla resumen
    st.markdown("### 📋 Resumen de tu pedido")
    _all_items_6 = _doc_detail_6 + _svc_detail_6
    if _all_items_6:
        _rows_html = "".join(
            f"<tr><td style='padding:6px 12px;color:#CBD5E1;'>{it['label']}</td>"
            f"<td style='padding:6px 12px;text-align:right;color:#F8FAFC;font-weight:600;'>€{it['precio']:,}</td></tr>"
            for it in _all_items_6
        )
        st.markdown(f"""
<table style='width:100%;border-collapse:collapse;background:rgba(30,58,95,0.3);border-radius:10px;overflow:hidden;'>
  <thead>
    <tr style='background:rgba(37,99,235,0.4);'>
      <th style='padding:10px 12px;text-align:left;color:#F8FAFC;'>Concepto</th>
      <th style='padding:10px 12px;text-align:right;color:#F8FAFC;'>Importe</th>
    </tr>
  </thead>
  <tbody>
    {_rows_html}
    <tr style='background:rgba(0,0,0,0.2);border-top:1px solid rgba(255,255,255,0.1);'>
      <td style='padding:8px 12px;color:#94A3B8;'>Subtotal (sin IVA)</td>
      <td style='padding:8px 12px;text-align:right;color:#94A3B8;'>€{_subtotal_6:,}</td>
    </tr>
    <tr style='background:rgba(0,0,0,0.2);'>
      <td style='padding:8px 12px;color:#94A3B8;'>IVA 21%</td>
      <td style='padding:8px 12px;text-align:right;color:#94A3B8;'>€{_iva_6:,}</td>
    </tr>
    <tr style='background:rgba(39,174,96,0.2);border-top:2px solid #27AE60;'>
      <td style='padding:12px;color:#F8FAFC;font-weight:800;font-size:15px;'>TOTAL CON IVA</td>
      <td style='padding:12px;text-align:right;color:#27AE60;font-weight:800;font-size:18px;'>€{_total_iva_6:,}</td>
    </tr>
  </tbody>
</table>""", unsafe_allow_html=True)
    else:
        st.info("No has seleccionado documentación ni servicios en el paso anterior. Vuelve a Servicios para seleccionarlos.")

    st.caption("⚠️ Aviso MVP: Los importes mostrados son orientativos. El IVA real puede variar. "
               "Este documento no tiene validez fiscal.")

    # Proforma descargable antes de pagar
    try:
        import io as _io6p
        from reportlab.lib.pagesizes import A4 as _A4p
        from reportlab.platypus import SimpleDocTemplate as _SDp, Paragraph as _Pp, Spacer as _Sp, Table as _Tp, TableStyle as _TSp
        from reportlab.lib.styles import getSampleStyleSheet as _GSS
        from reportlab.lib.colors import HexColor as _HC
        from reportlab.lib.units import cm as _cmp
        import datetime as _dtp

        _pf_buf = _io6p.BytesIO()
        _pf_doc = _SDp(_pf_buf, pagesize=_A4p, topMargin=2*_cmp, bottomMargin=2*_cmp,
                       leftMargin=2*_cmp, rightMargin=2*_cmp)
        _pf_ss  = _GSS()
        _pf_story = []
        _pf_story.append(_Pp("<b>ARCHIRAPID — PRESUPUESTO PREVIO (PROFORMA)</b>", _pf_ss["Title"]))
        _pf_story.append(_Sp(0.3*_cmp))
        _pf_story.append(_Pp(f"Fecha: {_dtp.date.today().strftime('%d/%m/%Y')} · "
                             f"Ref: AR-{req.get('nombre_proyecto','PRJ').replace(' ','')[:8].upper()}-{_dtp.date.today().strftime('%Y%m%d')}",
                             _pf_ss["Normal"]))
        _pf_story.append(_Pp(f"Proyecto: {req.get('nombre_proyecto', 'Proyecto ArchiRapid')} · "
                             f"{total_area:.0f} m² · Estilo {style}", _pf_ss["Normal"]))
        _pf_story.append(_Sp(0.5*_cmp))
        _pf_data = [["Concepto", "Importe (€)"]] + \
                   [[it["label"].replace("📄","").replace("📐","").replace("🏦","").replace("📋","").replace("🔍","")
                                .replace("🏛️","").replace("👷","").replace("🚿","").replace("⚡","").replace("☀️","")
                                .replace("🔬","").replace("📏","").replace("🦺","").replace("🥽","").strip(),
                     f"{it['precio']:,}"] for it in _all_items_6] + \
                   [["Subtotal sin IVA", f"{_subtotal_6:,}"],
                    ["IVA 21%",          f"{_iva_6:,}"],
                    ["TOTAL CON IVA",    f"{_total_iva_6:,}"]]
        _pf_tbl = _Tp(_pf_data, colWidths=[13*_cmp, 4*_cmp])
        _pf_tbl.setStyle(_TSp([
            ("BACKGROUND", (0,0), (-1,0), _HC("#1E3A5F")),
            ("TEXTCOLOR",  (0,0), (-1,0), _HC("#FFFFFF")),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("BACKGROUND", (0,-3), (-1,-1), _HC("#EBF5FB")),
            ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0,1), (-1,-4), [_HC("#FFFFFF"), _HC("#F8F9FA")]),
            ("GRID", (0,0), (-1,-1), 0.5, _HC("#CCCCCC")),
            ("PADDING", (0,0), (-1,-1), 6),
        ]))
        _pf_story.append(_pf_tbl)
        _pf_story.append(_Sp(0.8*_cmp))
        _pf_story.append(_Pp("<i>AVISO LEGAL: Este documento es una estimación orientativa generada "
                             "automáticamente por ArchiRapid MVP. No tiene validez fiscal ni contractual. "
                             "Los precios finales pueden variar. IVA referencial 21%. "
                             "Para presupuesto oficial contacta con hola@archirapid.com</i>",
                             _pf_ss["Normal"]))
        _pf_doc.build(_pf_story)
        st.download_button(
            label="📄 Descargar Presupuesto Previo (PDF)",
            data=_pf_buf.getvalue(),
            file_name=f"Presupuesto_ArchiRapid_{_dtp.date.today().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    except Exception:
        pass  # sin PDF proforma no bloquea el flujo

    st.markdown("---")
    if not st.session_state.get("pago_completado"):
        # ── Stripe Checkout ───────────────────────────────────────────────────
        _client_email_6 = (st.session_state.get("user_email") or
                           st.session_state.get("client_email") or "")
        _stripe_key_ok  = bool(os.getenv("STRIPE_SECRET_KEY", ""))
        try:
            import streamlit as _stk6
            if not _stripe_key_ok:
                _stripe_key_ok = bool(_stk6.secrets.get("STRIPE_SECRET_KEY", ""))
        except Exception:
            pass

        # Crear sesión Stripe sólo una vez (almacenada en session_state)
        if _stripe_key_ok and _all_items_6 and not st.session_state.get("stripe_session_id_s6"):
            try:
                from modules.stripe_utils import create_custom_session as _ccs6
                _stripe_items_6 = [
                    {"name": it["label"][:80], "amount_cents": int(it["precio"] * 100), "quantity": 1}
                    for it in _all_items_6
                ]
                _s6_url, _s6_sid = _ccs6(
                    line_items=_stripe_items_6,
                    client_email=_client_email_6,
                    success_url="https://archirapid.streamlit.app/?pago=ok",
                    cancel_url="https://archirapid.streamlit.app/?pago=cancel",
                    metadata={
                        "project": req.get("nombre_proyecto", "ArchiRapid Project"),
                        "total_eur": str(_total_iva_6),
                        "style": str(style),
                    },
                )
                st.session_state["stripe_session_id_s6"] = _s6_sid
                st.session_state["stripe_checkout_url_s6"] = _s6_url
            except Exception as _e6:
                st.session_state["stripe_session_id_s6"] = None
                st.session_state["stripe_checkout_url_s6"] = None

        _stripe_url_6 = st.session_state.get("stripe_checkout_url_s6")
        _stripe_sid_6 = st.session_state.get("stripe_session_id_s6")

        _cp1, _cp2 = st.columns(2)
        with _cp1:
            if _stripe_url_6:
                # Abre Stripe en nueva pestaña — preserva session_state
                import streamlit.components.v1 as _cv1_s6
                _cv1_s6.html(
                    f"""<a href="{_stripe_url_6}" target="_blank"
                        style="display:block;text-align:center;background:#1E3A5F;color:#fff;
                               padding:12px 20px;border-radius:8px;font-weight:700;font-size:15px;
                               text-decoration:none;">
                        💳 Pagar con Tarjeta — €{_total_iva_6:,} (IVA incl.)
                    </a>
                    <p style="color:#94A3B8;font-size:12px;text-align:center;margin-top:6px;">
                        Pago seguro con Stripe · Tarjeta de prueba: 4242 4242 4242 4242
                    </p>""",
                    height=80,
                )
            else:
                # Fallback simulado si Stripe no está configurado
                if st.button("💳 Confirmar y Pagar", type="primary",
                             use_container_width=True, key="btn_pagar_s6"):
                    st.session_state["pago_completado"] = True
                    st.session_state["total_pagado"]    = _total_iva_6
                    st.rerun()
        with _cp2:
            if st.button("📞 Hablar con un Arquitecto", use_container_width=True, key="btn_arq_s6"):
                st.info("📧 hola@archirapid.com · Solicita llamada en el chat →")

        # Botón de verificación de pago (después de volver de Stripe)
        if _stripe_sid_6:
            st.markdown("<br>", unsafe_allow_html=True)
            _vp1, _vp2, _vp3 = st.columns([1, 2, 1])
            with _vp2:
                if st.button("✅ Ya he pagado — Verificar pago", use_container_width=True,
                             key="btn_verify_s6"):
                    try:
                        from modules.stripe_utils import verify_session as _vs6
                        _sess6 = _vs6(_stripe_sid_6)
                        if _sess6.payment_status == "paid":
                            st.session_state["pago_completado"] = True
                            st.session_state["total_pagado"]    = _total_iva_6
                            st.session_state.pop("stripe_session_id_s6", None)
                            st.session_state.pop("stripe_checkout_url_s6", None)
                            st.rerun()
                        else:
                            st.warning("⏳ Pago no confirmado todavía. Completa el pago en Stripe y vuelve a verificar.")
                    except Exception as _ve6:
                        st.error(f"Error al verificar: {_ve6}")
        return  # ← nada más se renderiza hasta que se pague

    # ── Bloque post-pago ──────────────────────────────────────────────────────
    if st.session_state.get("pago_completado"):
        st.balloons()
        st.success(f"✅ Pago de €{st.session_state.get('total_pagado', 0):,} (IVA incluido) procesado. ¡Tu proyecto está listo!")

        # Guardar en panel de cliente
        try:
            import sqlite3 as _sq6, json as _js6, datetime as _dt6
            from modules.marketplace.utils import DB_PATH as _DBP6
            _conn6 = _sq6.connect(_DBP6, timeout=15)
            _conn6.execute("PRAGMA journal_mode=WAL")
            # ensure_tables() in src/db.py already creates ai_projects with req_json column
            # DO NOT redefine the table here — it would create it without req_json on fresh SQLite
            _req_json_val = _js6.dumps(st.session_state.get("ai_house_requirements") or {})
            _conn6.execute("""
                INSERT OR REPLACE INTO ai_projects
                (client_email, project_name, total_area, total_cost, services_json, style, energy_label, created_at, status, req_json)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (st.session_state.get("client_email",""),
                 req.get("nombre_proyecto","Proyecto ArchiRapid"),
                 float(total_area), float(total_cost),
                 _js6.dumps({"doc": _doc_detail_6, "svc": _svc_detail_6, "total_iva": _total_iva_6}),
                 style, energy_label,
                 _dt6.datetime.now().isoformat(), "pagado", _req_json_val))
            _conn6.commit(); _conn6.close()
        except Exception:
            pass  # nunca interrumpe el flujo

        # Calculadora de hipoteca
        try:
            from modules.marketplace.hipoteca import render_calculadora as _rc_hip
            _plot_h       = st.session_state.get("design_plot_data", {})
            _precio_t     = float(_plot_h.get("price") or 0)
            _coste_m2h    = 1800 if "premium" in style.lower() or "lujo" in style.lower() \
                            else (1200 if "eco" in style.lower() or "madera" in style.lower() else 1400)
            _coste_obra_h = total_area * _coste_m2h
            with st.expander("🏦 Calculadora de Financiación — ¿Cuánto pagarías al mes?", expanded=True):
                _rc_hip(precio_terreno=_precio_t, coste_construccion=_coste_obra_h, key_prefix="s6_hip")
        except Exception:
            pass

        st.markdown("### 📥 Descarga Tu Proyecto Completo")

    # ── Botón de descarga ZIP (siempre visible, sin muro de pago) ─────────────
    st.markdown("""
<div style="background:rgba(37,99,235,0.12);border:1px solid rgba(37,99,235,0.35);
            border-radius:12px;padding:16px 20px;margin-bottom:16px;">
  <div style="font-weight:800;color:#F8FAFC;font-size:15px;margin-bottom:8px;">
    📦 Contenido del ZIP del Proyecto
  </div>
  <ul style="color:#CBD5E1;font-size:13px;margin:0;padding-left:20px;line-height:1.9;">
    <li>📄 Memoria descriptiva (PDF)</li>
    <li>📊 Mediciones y presupuesto (Excel)</li>
    <li>🗺️ Plano de planta (PNG)</li>
    <li>🏗️ Archivo BIM/IFC (IFC2x3)</li>
    <li>📸 Vistas 3D (si capturaste en el editor)</li>
    <li>📋 Informe eficiencia energética</li>
    <li>🌿 Estimación huella de carbono</li>
  </ul>
</div>""", unsafe_allow_html=True)

    try:
        _plot_data = st.session_state.get("design_plot_data", {})
        _design_d  = f["design_data"]
        _zip_bytes, _zip_filename = generar_zip_proyecto(
            req=req,
            design_data=_design_d,
            plot_data=_plot_data,
            partidas=[(p[0], p[1], p[2], p[3]) for p in partidas],
            subsidy_total=subsidy_total,
            energy_label=energy_label,
        )
        st.download_button(
            label="📦 DESCARGAR TODO EL PROYECTO (ZIP)",
            data=_zip_bytes,
            file_name=_zip_filename,
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )
        st.caption("Incluye: Memoria descriptiva · Mediciones Excel · Plano 2D · Datos catastro · Layout 3D")

        # Certificación blockchain del ZIP
        try:
            from modules.marketplace.blockchain_cert import certify, cert_badge_html
            _cert6 = certify(
                zip_bytes=_zip_bytes,
                doc_name=_zip_filename,
                user_email=st.session_state.get("client_email", ""),
                plot_id=str(st.session_state.get("design_plot_data", {}).get("id", ""))
            )
            st.markdown(cert_badge_html(_cert6), unsafe_allow_html=True)
        except Exception:
            pass

    except Exception as _ez:
        st.error(f"Error generando ZIP: {_ez}")

    # ── IFC / Gemelo Digital ──────────────────────────────────────────────────
    try:
        from modules.ai_house_designer.ifc_export import generate_ifc
        _rooms_ifc = (
            f["design_data"].get("rooms") or
            [{"name": k, "area": float(v)}
             for k, v in req.get("ai_room_proposal", {}).items()
             if isinstance(v, (int, float))]
        )
        if _rooms_ifc:
            _pname_ifc = (req.get("nombre_proyecto") or "ArchiRapid_Proyecto").replace(" ", "_")
            _ifc_bytes = generate_ifc(_rooms_ifc, project_name=_pname_ifc)
            st.markdown("---")
            st.markdown("""
<div style="background:linear-gradient(135deg,rgba(30,58,95,0.6),rgba(13,27,42,0.8));
            border:1px solid rgba(245,158,11,0.35);border-radius:12px;
            padding:16px 20px;margin-bottom:12px;">
  <div style="font-size:15px;font-weight:800;color:#F8FAFC;margin-bottom:4px;">
    🏗️ Gemelo Digital BIM/IFC
  </div>
  <div style="font-size:12px;color:#94A3B8;">
    Formato IFC2x3 · FreeCAD · Archicad · Revit · Navisworks · BIMvision
  </div>
  <div style="margin-top:8px;font-size:11px;color:#64748B;">
    ✅ Subvencionable UE (BIM Mandate) · Cada habitación = objeto IFC certificado
  </div>
</div>""", unsafe_allow_html=True)
            st.download_button(
                label="📐 Descargar Gemelo Digital (.ifc)",
                data=_ifc_bytes,
                file_name=f"{_pname_ifc}.ifc",
                mime="application/x-step",
                use_container_width=True,
            )
    except Exception:
        pass

    # ── Tablón de Obras ───────────────────────────────────────────────────────
    try:
        st.markdown("---")
        st.markdown("### 🏗️ ¿Quieres recibir ofertas de constructores?")
        _tab_key = f"tablon_published_s6_{req.get('nombre_proyecto','')}"
        if st.session_state.get(_tab_key):
            st.success("✅ Publicado en el Tablón. Los constructores de tu zona ya pueden enviarte ofertas.")
            st.info("Consulta las ofertas en tu **Panel de Cliente → Ofertas de Construcción**.")
        else:
            st.markdown("""
<div style="background:rgba(34,197,94,0.07);border:1px solid rgba(34,197,94,0.25);
            border-radius:10px;padding:14px 18px;margin-bottom:12px;">
  <div style="font-weight:700;color:#F8FAFC;font-size:14px;">🏗️ Red de constructores ArchiRapid</div>
  <div style="color:#94A3B8;font-size:12px;margin-top:4px;">
    Publicamos tu proyecto (anónimo). Los constructores de tu provincia te envían
    ofertas con precio, plazo y garantía. Tú comparas. Sin compromiso.
  </div>
</div>""", unsafe_allow_html=True)
            _prov_t  = (req.get("province") or
                        st.session_state.get("design_plot_data", {}).get("province") or "")
            _cmail   = (st.session_state.get("user_email") or
                        st.session_state.get("client_email") or "")
            _cname   = (st.session_state.get("user_name") or
                        st.session_state.get("client_name") or "Cliente")
            _pname_t = req.get("nombre_proyecto") or "Mi proyecto ArchiRapid"
            from modules.marketplace.service_providers import _ESP_LABELS as _ESP_L6
            _parts_sel6 = st.multiselect(
                "¿Qué partidas quieres contratar? (deja vacío = obra completa)",
                options=list(_ESP_L6.keys()),
                format_func=lambda x: _ESP_L6.get(x, x),
                key="tablon_partidas_sel_s6",
                help="Solo recibirás ofertas de profesionales con esas especialidades.",
            )
            if st.button("🏗️ Publicar en Tablón de Obras — Recibir ofertas",
                         type="primary", use_container_width=True, key="btn_tablon_s6"):
                try:
                    from modules.marketplace.service_providers import publish_to_tablon
                    _tid = publish_to_tablon(
                        client_email=_cmail, client_name=_cname,
                        project_name=_pname_t, province=_prov_t, style=style,
                        total_area=float(total_area), total_cost=float(total_cost),
                        partidas_list=[(p[0], p[1], p[2], p[3]) for p in partidas],
                        partidas_solicitadas=_parts_sel6 if _parts_sel6 else None,
                    )
                    st.session_state[_tab_key] = True
                    try:
                        from modules.marketplace.email_notify import _send
                        _parts_s = ", ".join(_parts_sel6) if _parts_sel6 else "obra completa"
                        _send(f"🏗️ Nuevo Tablón\nID:{_tid}\n{_cname} ({_cmail})\n"
                              f"{_pname_t} · {total_area:.0f}m² · €{total_cost:,}\nPartidas: {_parts_s}")
                    except Exception:
                        pass
                    st.success("✅ Publicado. Los constructores ya pueden enviarte ofertas.")
                    st.rerun()
                except Exception as _etab:
                    st.error(f"Error publicando: {_etab}")
    except Exception:
        pass

    st.markdown("---")
    _cb6, _ = st.columns(2)
    with _cb6:
        if st.button("← Volver a Servicios", use_container_width=True, key="s6_back"):
            st.session_state["ai_house_step"] = 5
            st.rerun()
    st.info("📬 Proyecto guardado en tu **Panel de Cliente**. Accede desde el menú para consultarlo, modificarlo o contactar con nosotros.")